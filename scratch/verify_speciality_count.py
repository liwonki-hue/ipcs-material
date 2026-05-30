# Speciality Item List 시트별 태그 수 vs BOM 카테고리별 태그 수 완전 대조
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BOM_PATH = 'Raw File/Speciality BOM.xlsx'
IL_PATH  = 'Raw File/Speciality Item List.xlsx'

SHEET_TAG_COL = {
    '1.STRAINER':             20,
    '2. STEAM TRAP':          20,
    '3. EXPANSION JOINT':     20,
    '4. SIGHT GLASS':         20,
    '5. FLEXIBLE JOINT':      20,
    '6. RESTRICTION ORIFICE': 21,
    '7. FLEXIBLE HOSE':       20,
    '8. SPRAY NOZZLE':        20,
    '9.EDUCTOR':              2,
    '10.AIR TRAP':            20,
    '11.BIRD SCREEN':         19,
}

# matcode prefix → Item List 시트 매핑
MATCODE_TO_SHEET = {
    'STR': '1.STRAINER',
    'STP': '2. STEAM TRAP',
    'EXJ': '3. EXPANSION JOINT',
    'SGG': '4. SIGHT GLASS',
    'FLX': '5. FLEXIBLE JOINT',
    'ROR': '6. RESTRICTION ORIFICE',
    'FLH': '7. FLEXIBLE HOSE',
    'SPN': '8. SPRAY NOZZLE',
    'EDC': '9.EDUCTOR',
    'ATP': '10.AIR TRAP',
    'BSC': '11.BIRD SCREEN',
}

# ── Item List: 시트별 태그 수집 ───────────────────────────────
wb_il = openpyxl.load_workbook(IL_PATH, data_only=True)
il_by_sheet = {}   # {sheet: [tags]}

for sh, tag_col in SHEET_TAG_COL.items():
    ws = wb_il[sh]
    tags = []
    for r in range(5, ws.max_row + 1):
        v = ws.cell(r, tag_col).value
        if not v:
            continue
        tag = str(v).strip()
        if not any(tag.startswith(p) for p in ('B0-', 'B1-', 'B2-')):
            continue
        tags.append(tag)
    il_by_sheet[sh] = tags

# ── BOM: matcode prefix 기준 시트별 태그 수집 ─────────────────
wb_bom = openpyxl.load_workbook(BOM_PATH, data_only=True)
ws_bom = wb_bom.active
bom_by_sheet = {sh: [] for sh in SHEET_TAG_COL}

for r in range(2, ws_bom.max_row + 1):
    mc  = str(ws_bom.cell(r, 2).value or '').strip()
    tag = str(ws_bom.cell(r, 3).value or '').strip()
    if not tag or not mc:
        continue
    prefix = mc.split('-')[0]
    sheet  = MATCODE_TO_SHEET.get(prefix)
    if sheet:
        bom_by_sheet[sheet].append(tag)

# ── 비교 ──────────────────────────────────────────────────────
print(f'{"시트":<28} {"IL 건수":>8} {"BOM 건수":>8} {"차이":>6} {"중복(IL)":>8} {"중복(BOM)":>9}')
print('-' * 75)

all_ok = True
for sh in SHEET_TAG_COL:
    il_tags  = il_by_sheet[sh]
    bom_tags = bom_by_sheet[sh]
    il_cnt   = len(il_tags)
    bom_cnt  = len(bom_tags)
    diff     = bom_cnt - il_cnt
    il_dup   = il_cnt - len(set(il_tags))
    bom_dup  = bom_cnt - len(set(bom_tags))
    flag     = '' if diff == 0 else '  ← 불일치'
    print(f'{sh:<28} {il_cnt:>8} {bom_cnt:>8} {diff:>+6} {il_dup:>8} {bom_dup:>9}{flag}')
    if diff != 0:
        all_ok = False

print('-' * 75)
il_total  = sum(len(v) for v in il_by_sheet.values())
bom_total = sum(len(v) for v in bom_by_sheet.values())
print(f'{"합계":<28} {il_total:>8} {bom_total:>8} {bom_total-il_total:>+6}')
print()

if all_ok:
    print('결과: 전체 시트 완전 일치.')
else:
    print('결과: 불일치 항목 존재 — 아래 상세 확인.')
    for sh in SHEET_TAG_COL:
        il_set  = set(il_by_sheet[sh])
        bom_set = set(bom_by_sheet[sh])
        missing = il_set - bom_set
        extra   = bom_set - il_set
        if missing or extra:
            print(f'\n[{sh}]')
            for t in sorted(missing):
                print(f'  IL에 있고 BOM 없음: {t}')
            for t in sorted(extra):
                print(f'  BOM에 있고 IL 없음: {t}')

# ── BOM 전체에서 matcode prefix 미인식 항목 ───────────────────
unrecog = []
for r in range(2, ws_bom.max_row + 1):
    mc  = str(ws_bom.cell(r, 2).value or '').strip()
    tag = str(ws_bom.cell(r, 3).value or '').strip()
    if not tag or not mc:
        continue
    prefix = mc.split('-')[0]
    if prefix not in MATCODE_TO_SHEET:
        unrecog.append((r, mc, tag))

if unrecog:
    print(f'\n[경고] matcode prefix 미인식 행 {len(unrecog)}건 (BOM 집계 누락 가능):')
    for row, mc, tag in unrecog:
        print(f'  row{row}: mc={mc}  tag={tag}')
