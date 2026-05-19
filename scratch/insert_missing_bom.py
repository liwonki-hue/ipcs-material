# B126/B226/B227/B127 ST/HS/LS/CH 시스템 누락 BOM 데이터 삽입
# Excel에서 파싱 후 Supabase bom 테이블에 INSERT
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import warnings
warnings.filterwarnings("ignore")
import openpyxl, requests, re, time
from collections import defaultdict

SUPABASE_URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
    'Content-Profile': 'material',
    'Accept-Profile': 'material',
    'Prefer': 'return=minimal'
}

# ─── DN→INCH 변환 맵 ─────────────────────────────────────────────────
DN_TO_INCH = {
    'DN6': '1/4"', 'DN8': '3/8"', 'DN10': '3/8"',
    'DN15': '1/2"', 'DN20': '3/4"', 'DN25': '1"',
    'DN32': '1-1/4"', 'DN40': '1-1/2"', 'DN50': '2"',
    'DN65': '2-1/2"', 'DN80': '3"', 'DN100': '4"',
    'DN125': '5"', 'DN150': '6"', 'DN200': '8"',
    'DN250': '10"', 'DN300': '12"', 'DN350': '14"',
    'DN400': '16"', 'DN450': '18"', 'DN500': '20"',
    'DN600': '24"', 'DN700': '28"', 'DN750': '30"',
    'DN800': '32"', 'DN900': '36"', 'DN1000': '40"',
}

def dn_to_inch(size_str):
    """DN 25 → 1" 변환"""
    s = re.sub(r'\s+', '', str(size_str).upper())  # "DN 25" → "DN25"
    if s in DN_TO_INCH:
        return DN_TO_INCH[s]
    # X DN Y 형태 (Reducer 등)
    m = re.match(r'(DN\d+)\s*[Xx]\s*(DN\d+)', s)
    if m:
        a = DN_TO_INCH.get(m.group(1), m.group(1))
        b = DN_TO_INCH.get(m.group(2), m.group(2))
        return f"{a} X {b}"
    return size_str  # 변환 못하면 원본 반환

def norm_matl(matl):
    """재질 정규화"""
    m = str(matl).strip()
    # SA → A (ASME prefix 제거)
    m = re.sub(r'^SA-?', 'A', m)
    m = re.sub(r'^SA(\d)', r'A\1', m)
    return m

def get_category(item):
    """ITEM 명에서 카테고리 분류"""
    item_u = str(item).upper()
    if any(x in item_u for x in ['PIPE SMLS', 'PIPE WELD', 'PIPE ERW', 'PIPE EFW', 'PIPE NIPPLE', 'PIPE SEAMLESS']):
        return 'Pipe'
    if any(x in item_u for x in ['VALVE', 'ACTUATOR']):
        return 'Valve'
    if any(x in item_u for x in ['GASKET', 'GSKT']):
        return 'Others'
    if any(x in item_u for x in ['STUD', 'BOLT', 'NUT', 'STB']):
        return 'Others'
    if any(x in item_u for x in ['ELBOW', 'TEE', 'CAP', 'COUPLING', 'FLANGE', 'REDUCER',
                                   'SOCKET', 'SWAGE', 'UNION', 'CROSS', 'NIPPO', 'OLET',
                                   'FLG', 'FLN', 'FLB', 'FLA', 'FLS', 'WOL', 'SOL',
                                   'RDC', 'RDE', 'EL9', 'EL4', 'TER', 'CPF', 'SCN']):
        return 'Fitting'
    if 'PIPE' in item_u:
        return 'Pipe'
    return 'Fitting'  # 기본값

def build_full_description(item, matl1, matl2, size, thick, et):
    """full_description 빌드: ITEM, MATL, SIZE_INCH, CLASS/SCH, ET"""
    parts = []
    parts.append(str(item).strip())
    matl = norm_matl(matl1) if matl1 else ''
    if matl: parts.append(matl)
    size_norm = dn_to_inch(size) if size else ''
    if size_norm: parts.append(size_norm)
    if thick: parts.append(str(thick).strip())
    if et: parts.append(str(et).strip())
    return ', '.join(filter(None, parts))

def extract_system(line_no):
    """line_no에서 system 코드 추출"""
    m = re.search(r'^\d+["-]([A-Z]+)-', str(line_no))
    return m.group(1) if m else ''

# ─── 1. DB에서 현재 ISO 목록 수집 ────────────────────────────────────

print("=== DB ISO 목록 수집 중... ===")
db_isos = set()
offset = 0
while True:
    resp = requests.get(
        f'{SUPABASE_URL}/rest/v1/bom?select=iso_dwg_no&iso_dwg_no=neq.&limit=1000&offset={offset}',
        headers=HEADERS
    )
    rows = resp.json()
    if not rows: break
    for r in rows: db_isos.add(r['iso_dwg_no'].strip())
    offset += 1000
    if len(rows) < 1000: break

# 테스트 행 삭제 시도 (이전 INSERT 테스트)
requests.delete(f'{SUPABASE_URL}/rest/v1/bom?iso_dwg_no=eq.__TEST_ISO__', headers=HEADERS)

print(f"DB ISO: {len(db_isos)}")

# ─── 2. Excel 파싱: 550개 미등록 ISO 데이터 추출 ─────────────────────

def parse_bom_sheet(filepath, sheet_name, header_row, uom_col, qty_col, target_isos):
    """target_isos에 속하는 행만 추출"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    # (iso, line, full_desc_key) -> {qty, category, system}
    result = {}
    rows_read = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row: continue
        if row[1] is None or not isinstance(row[1], (int, float)): continue
        iso = str(row[8] or '').strip()
        if iso not in target_isos: continue

        line = str(row[9] or '').strip().replace('\r', '').replace('\n', '')
        system_full = str(row[6] or '').strip()
        item  = str(row[14] or '').strip()
        et    = str(row[15] or '').strip()
        matl1 = str(row[16] or '').strip()
        matl2 = str(row[17] or '').strip()
        size  = str(row[18] or '').strip()
        thick = str(row[19] or '').strip()
        uom_raw = str(row[uom_col] or '').strip() if len(row) > uom_col else ''
        qty_raw = row[qty_col] if len(row) > qty_col else None

        if not item: continue
        try:
            qty = float(qty_raw) if qty_raw is not None else 0.0
        except: qty = 0.0
        if uom_raw == 'MM':
            qty = qty / 1000.0
            uom_raw = 'M'
        if qty <= 0: continue

        full_desc = build_full_description(item, matl1, matl2, size, thick, et)
        system = extract_system(line)
        category = get_category(item)

        key = (iso, line, full_desc, uom_raw)
        if key not in result:
            result[key] = {'qty': 0.0, 'category': category, 'system': system}
        result[key]['qty'] += qty
        rows_read += 1

    wb.close()
    print(f"  {filepath.split('/')[-1]}: {rows_read}행")
    return result

# Excel에만 있는 ISO 목록
lb_isos_excel = set()
sb_isos_excel = set()

wb_lb = openpyxl.load_workbook("Raw File/LARGE BORE BOM(251223).xlsm", read_only=True, data_only=True)
ws_lb = wb_lb["Piping&Fitting"]
for i, row in enumerate(ws_lb.iter_rows(values_only=True)):
    if i < 8: continue
    if row[1] is None or not isinstance(row[1], (int, float)): continue
    iso = str(row[8] or '').strip()
    if iso: lb_isos_excel.add(iso)
wb_lb.close()

wb_sb = openpyxl.load_workbook("Raw File/SB BOM(20260128) (002).xlsx", read_only=True, data_only=True)
ws_sb = wb_sb["MERGED_Piping&Fitting"]
for i, row in enumerate(ws_sb.iter_rows(values_only=True)):
    if i < 1: continue
    if row[1] is None or not isinstance(row[1], (int, float)): continue
    iso = str(row[8] or '').strip()
    if iso: sb_isos_excel.add(iso)
wb_sb.close()

all_excel_isos = lb_isos_excel | sb_isos_excel
target_isos = all_excel_isos - db_isos
print(f"\n삽입 대상 ISO: {len(target_isos)}개")

print("\n=== Excel 파싱 중... ===")
# SB BOM 먼저 파싱 (우선순위)
sb_data = parse_bom_sheet("Raw File/SB BOM(20260128) (002).xlsx", "MERGED_Piping&Fitting", 1, 24, 25, target_isos)
# LB BOM 파싱 (SB에 없는 ISO/LINE 항목만 추가)
lb_data = parse_bom_sheet("Raw File/LARGE BORE BOM(251223).xlsm", "Piping&Fitting", 8, 23, 24, target_isos)

# 합치기: SB 우선, LB는 SB에 없는 key만
combined = dict(sb_data)
lb_only = 0
for key, val in lb_data.items():
    if key not in combined:
        combined[key] = val
        lb_only += 1

print(f"\nSB 데이터: {len(sb_data)}건")
print(f"LB 추가: {lb_only}건")
print(f"전체 삽입 대상 행: {len(combined)}건")

# ─── 3. Supabase INSERT ──────────────────────────────────────────────

print("\n=== Supabase bom INSERT 중... ===")

# 배치 INSERT (50건씩)
rows_to_insert = []
for (iso, line, full_desc, uom), val in combined.items():
    rows_to_insert.append({
        'mat_code': None,
        'category': val['category'],
        'tag': None,
        'system': val['system'],
        'iso_dwg_no': iso,
        'line_no': line,
        'full_description': full_desc,
        'uom': uom,
        'qty': round(val['qty'], 4)
    })

BATCH = 50
total_inserted = 0
errors = []

for i in range(0, len(rows_to_insert), BATCH):
    batch = rows_to_insert[i:i+BATCH]
    resp = requests.post(
        f'{SUPABASE_URL}/rest/v1/bom',
        json=batch,
        headers=HEADERS
    )
    if resp.status_code in (200, 201):
        total_inserted += len(batch)
    else:
        errors.append((i, resp.status_code, resp.text[:100]))
        print(f"  오류 배치 {i//BATCH+1}: {resp.status_code} {resp.text[:100]}")

    if (i // BATCH + 1) % 10 == 0:
        print(f"  진행: {min(i+BATCH, len(rows_to_insert))}/{len(rows_to_insert)}")

print(f"\n삽입 완료: {total_inserted}건")
if errors:
    print(f"오류: {len(errors)}건")
    for e in errors[:5]:
        print(f"  배치#{e[0]//BATCH+1}: {e[1]} {e[2]}")

# ─── 4. 결과 확인 ────────────────────────────────────────────────────

print("\n=== 삽입 후 DB 확인 ===")
resp_check = requests.get(
    f'{SUPABASE_URL}/rest/v1/bom?select=count&iso_dwg_no=neq.',
    headers={**HEADERS, 'Prefer': 'count=exact', 'Range-Unit': 'items', 'Range': '0-0'}
)
print(f"bom 전체 행 수: {resp_check.headers.get('Content-Range')}")

# 삽입된 ISO 샘플 확인
sample_iso = sorted(target_isos)[0]
resp_sample = requests.get(
    f'{SUPABASE_URL}/rest/v1/bom?select=iso_dwg_no,line_no,full_description,uom,qty&iso_dwg_no=eq.{sample_iso}&limit=3',
    headers=HEADERS
)
print(f"\n샘플 ISO ({sample_iso}):")
for r in resp_sample.json():
    print(f"  {r}")
