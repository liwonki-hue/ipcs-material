# Flexible Hose & Joint BOM Excel → material.bom INSERT SQL 생성
import openpyxl

BOM_PATH = 'Raw File/Flexible Hose & Joint BOM.xlsx'
OUT_SQL  = 'scratch/insert_flexible_bom.sql'

def sq(v):
    if v is None:
        return 'NULL'
    return "'" + str(v).strip().replace("'", "''") + "'"

wb = openpyxl.load_workbook(BOM_PATH, data_only=True)
ws = wb['TAG_NO']

lines = [
    '-- Flexible Hose & Joint BOM 등록 (TAG_NO 시트)',
    '-- 원본: Raw File/Flexible Hose & Joint BOM.xlsx',
    '',
]

cnt = 0
skipped = 0
for r in range(2, ws.max_row + 1):
    tag = ws.cell(r, 6).value
    if not tag:
        skipped += 1
        continue

    mc   = ws.cell(r, 1).value          # MATCODE
    cat  = ws.cell(r, 2).value or 'Speciality'
    sys_ = ws.cell(r, 3).value          # SYSTEM
    iso  = ws.cell(r, 4).value          # ISO DRAWING
    lno  = ws.cell(r, 5).value          # LINE NO
    desc = ws.cell(r, 9).value          # DESCRIPTION
    uom  = ws.cell(r, 10).value or 'EA' # UNIT
    qty  = ws.cell(r, 11).value
    if qty is None:
        qty = 1

    vals = ', '.join([
        sq(mc), sq(cat), sq(str(tag).strip()),
        sq(sys_), sq(iso), sq(lno), sq(desc),
        sq(uom), str(qty)
    ])
    lines.append(
        f'INSERT INTO material.bom'
        f' (mat_code, category, tag, system, iso_dwg_no, line_no, full_description, uom, qty)'
        f' VALUES ({vals});'
    )
    cnt += 1

with open(OUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'생성 완료: {OUT_SQL}')
print(f'  INSERT: {cnt}건, 스킵(빈 TAG): {skipped}건')
