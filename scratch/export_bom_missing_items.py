# BOM 미등록 발주항목 엑셀 출력 (PO 발주·입고됐으나 BOM에 없는 품목)
import sys, re, requests, openpyxl
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

URL      = 'https://ognhvfvlboqblueuldlm.supabase.co'
KEY      = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
HDR      = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Accept-Profile': 'material'}
IN_FILE  = 'c:/Users/PCLOVE/Downloads/ipcs-material/Raw File/PO_List_Final_v22.xlsx'
OUT_FILE = 'c:/Users/PCLOVE/Downloads/ipcs-material/scratch/BOM_미등록_발주항목.xlsx'

# ── 변환 테이블 ───────────────────────────────────────────────────────────────
ITEM_TO_PREFIX = {
    'PIPE SEAMLESS':'PIS','PIPE WELDED':'PIW','PIPE WELDED/SEAMLESS':'PIW','PIPE':'PIW',
    'NIPPLE':'PIN',
    'ELBOW LR 90D':'EL9L','ELBOW SR 90D':'EL9S','ELBOW 45D':'EL4L','ELBOW':'EL9L',
    'TEE':'TEE','TEE-RED':'TER',
    'REDUCER-CON':'RDC','REDUCER-ECC':'RDE',
    'CAP':'CAP',
    'FLANGE':'FLN','FLANGE-BLIND':'FLB','FLANGE-SOCKET':'FLS',
    'FLANGE-SLIP':'FLA','FLANGE-WELD':'FLW',
    'WELDOLET':'WOL','SOCKOLET':'SOL','THREADOLET':'TOL',
    'COUPLING-FULL':'CPF','COUPLING-HALF':'CPH','COUPLING':'CPF',
    'SWAGE-CON':'SCN','SWAGE-ECC':'SCE','SWAGE':'SCN',
    'UNION':'UNI',
}
MATL_TO_CODE = {
    'A105':'CS05','A234-WPB':'CS05',
    'A182-F304':'SS04','A403-WP304':'SS04','A403-WP304W':'SS04',
    'A312-TP304':'SS04','A312-TP304W':'SS04',
    'A182-F316':'SS16','A403-WP316':'SS16','A403-WP316W':'SS16',
    'A312-TP316':'SS16','A312-TP316W':'SS16',
    'A182-F91':'AS91','A234-WP91':'AS91',
    'A182-F22':'AS22','A234-WP22':'AS22',
    'A106-B':'CS06','A106-GRB':'CS06',
    'A672-B60-CL22':'B060','A672-B60 CL22':'B060','A672-B60-22':'B060',
    'A420-WPL6':'LT06',
}
DN_TO_DCODE = {
    15:'D005',20:'D008',25:'D010',32:'D013',40:'D015',50:'D020',
    65:'D025',80:'D030',100:'D040',125:'D050',150:'D060',200:'D080',
    250:'D100',300:'D120',350:'D140',400:'D160',450:'D180',500:'D200',
    550:'D220',600:'D240',650:'D260',700:'D280',750:'D300',800:'D320',
}
FLANGE_PREFIXES = {'FLN','FLB','FLS','FLA','FLW'}
SMALL_BORE      = {'D005','D008','D010','D013','D015','D020'}

def normalize_class(raw):
    s = re.sub(r'\s', '', str(raw)).upper()
    s = re.sub(r'^S-', 'S', s)
    s = re.sub(r'^CL', 'C', s)
    return s

def parse_desc(desc):
    parts = [p.strip() for p in desc.split(',')]
    if len(parts) < 2:
        return None
    item_raw    = parts[0].upper()
    item_prefix = None
    for key in sorted(ITEM_TO_PREFIX, key=len, reverse=True):
        if item_raw.startswith(key):
            item_prefix = ITEM_TO_PREFIX[key]
            break
    if not item_prefix:
        return None
    matl_raw  = parts[1].strip().upper()
    matl_code = MATL_TO_CODE.get(matl_raw)
    if not matl_code:
        for k, v in MATL_TO_CODE.items():
            if k.upper() in matl_raw:
                matl_code = v
                break
    if not matl_code:
        matl_code = matl_raw
    dns_raw = ','.join(parts[2:])
    dcodes  = [DN_TO_DCODE.get(int(d)) for d in re.findall(r'DN\s*(\d+)', dns_raw.upper())
               if DN_TO_DCODE.get(int(d))]
    if not dcodes:
        return None
    class_raw = ''
    et_raw    = ''
    for p in parts[3:]:
        p = p.strip().upper()
        if re.match(r'^(BW|RF|FF|SW|PE|BE|TH|WNRF|WNFF)$', p):
            et_raw = 'RF' if 'RF' in p else ('SW' if p == 'SW' else 'BW')
            continue
        tok = p.split('X')[0].strip() if 'X' in p else p
        if not class_raw and tok and not re.match(r'^DN', tok):
            class_raw = normalize_class(tok)
    if not et_raw:
        if item_prefix in FLANGE_PREFIXES:
            et_raw = 'RF'
        elif item_prefix in {'CPF','CPH','SCN','SCE','EL9L','EL9S','EL4L','TEE','TER','CAP'} \
                and dcodes and dcodes[0] in SMALL_BORE:
            et_raw = 'SW'
        else:
            et_raw = 'BW'
    return item_prefix, matl_code, dcodes, class_raw, et_raw

def build_mc(parsed):
    item_prefix, matl_code, dcodes, class_desc, et_desc = parsed
    size_part = dcodes[0] if len(dcodes) == 1 else ''.join(dcodes[:2])
    return f'{item_prefix}-{matl_code}-{size_part}-{class_desc}-{et_desc}'

# ── Supabase 조회 ─────────────────────────────────────────────────────────────
def fetch_all(table, select):
    rows, offset, limit = [], 0, 1000
    while True:
        params = {'select': select, 'offset': str(offset), 'limit': str(limit)}
        data   = requests.get(f'{URL}/rest/v1/{table}', headers=HDR, params=params).json()
        if not isinstance(data, list) or not data:
            break
        rows.extend(data)
        if len(data) < limit:
            break
        offset += limit
    return rows

print('Supabase 조회 중...')
bom_rows = fetch_all('bom', 'mat_code,full_description,system,iso_dwg_no,uom,qty')
bom_matcodes = {(r['mat_code'] or '').strip().upper() for r in bom_rows if r.get('mat_code')}

recv_rows = fetch_all('receiving', 'mat_code,qty,unit')
recv_qty  = {}   # mat_code → {qty, unit}
for r in recv_rows:
    mc = (r.get('mat_code') or '').strip().upper()
    if not mc:
        continue
    if mc not in recv_qty:
        recv_qty[mc] = {'qty': 0.0, 'unit': r.get('unit') or 'EA'}
    recv_qty[mc]['qty'] += float(r.get('qty') or 0)

# ── 미매칭 분析 로드 및 분류 ──────────────────────────────────────────────────
wb_in = openpyxl.load_workbook(IN_FILE, data_only=True)
ws_mm = next(ws for ws in wb_in.worksheets if '미매칭' in ws.title)

result_rows = []
for row in range(4, ws_mm.max_row + 1):
    no      = ws_mm.cell(row, 1).value
    tag     = ws_mm.cell(row, 2).value
    if not no or not tag:
        continue
    po_no    = ws_mm.cell(row, 3).value
    bom_desc = ws_mm.cell(row, 4).value
    po_desc  = ws_mm.cell(row, 5).value

    bom_parsed = parse_desc(str(bom_desc).strip()) if bom_desc else None
    po_parsed  = parse_desc(str(po_desc).strip())  if po_desc  else None
    bom_mc     = build_mc(bom_parsed) if bom_parsed else None
    po_mc      = build_mc(po_parsed)  if po_parsed  else None

    # 동일 mat_code이거나 BOM에 이미 있으면 제외
    if bom_mc and po_mc and bom_mc == po_mc:
        continue
    if bom_mc and bom_mc in bom_matcodes:
        continue
    # 입고 확인된 항목만 포함
    if not po_mc or po_mc not in recv_qty:
        continue

    result_rows.append({
        'no':      no,
        'tag':     str(tag).strip(),
        'po_no':   str(po_no).strip() if po_no else '',
        'po_desc': str(po_desc).strip() if po_desc else '',
        'po_mc':   po_mc,
    })

print(f'BOM 미등록 항목 (입고 확인): {len(result_rows)}건')

# ── 스타일 상수 ───────────────────────────────────────────────────────────────
HDR_F  = PatternFill('solid', fgColor='1F4E79')
OK_F   = PatternFill('solid', fgColor='C6EFCE')
HDR2_F = PatternFill('solid', fgColor='833C00')
OK2_F  = PatternFill('solid', fgColor='C6EFCE')
WARN2_F= PatternFill('solid', fgColor='FFEB9C')
ERR_F  = PatternFill('solid', fgColor='FFC7CE')
THIN   = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)

def apply_hdr(ws, headers, widths, fill):
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.fill = fill
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

# ── Sheet 1: BOM 미등록 발주항목 ──────────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws1 = wb_out.active
ws1.title = 'BOM 미등록 발주항목'

S1_HEADERS = ['No', 'TAG NO', 'PO NO', 'PO List 표기 (발주)', 'mat_code', '입고 수량', '단위']
S1_WIDTHS  = [4, 34, 24, 52, 34, 12, 8]
apply_hdr(ws1, S1_HEADERS, S1_WIDTHS, HDR_F)

for r, row in enumerate(result_rows, 2):
    rq   = recv_qty.get(row['po_mc'].upper(), {})
    vals = [row['no'], row['tag'], row['po_no'], row['po_desc'],
            row['po_mc'], rq.get('qty', 0), rq.get('unit', 'EA')]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(r, c, v)
        cell.border = THIN
        cell.font   = Font(size=9)
        cell.alignment = Alignment(vertical='center', wrap_text=(c == 4))
        if c in (5, 6, 7):
            cell.fill = OK_F
    ws1.row_dimensions[r].height = 30

# ── Sheet 2: mat_code 변경 검토 ───────────────────────────────────────────────
CHANGE_PAIRS = [
    ('EL4L-SS04-D080-S40S-BW', 'EL9L-SS04-D080-S40S-BW',
     'ELBOW 타입 변경', '⚠️ 엔지니어링 확인', '45D → LR 90D 타입 변경. 설계 원본 확인 필요'),
    ('EL9L-CS05-D060-S40-BW',  '(파싱 오류 — 수동 처리)',
     '재질 코드 비표준', '🔴 수동 처리', 'PO 재질 A234-WP22 CL1 비표준 표기. 설계팀 원본 도면 확인 필요'),
    ('FLN-CS05-D140-C150-RF',  'FLN-CS05-D140-C300-RF',
     '등급 변경 (CL150→CL300)', '⚠️ 설계팀 승인 필요', 'FLANGE DN350: 전체 BOM 12행 영향. 설계 의도 확인 후 실행'),
    ('FLN-CS05-D200-C150-RF',  'FLN-CS05-D200-C300-RF',
     '등급 변경 (CL150→CL300)', '⚠️ 설계팀 승인 필요', 'FLANGE DN500: 전체 BOM 19행 영향. 설계 의도 확인 후 실행'),
]

bom_by_mc = defaultdict(list)
for r in bom_rows:
    mc = (r.get('mat_code') or '').strip().upper()
    if mc:
        bom_by_mc[mc].append(r)

ws2 = wb_out.create_sheet('mat_code 변경 검토')

S2_HEADERS = ['변경 유형', '판단', '현행 mat_code (BOM)', '변경 mat_code (PO)',
              'System', 'ISO Drawing No', 'Full Description', 'UOM', 'Qty (BOM)', '비고']
S2_WIDTHS  = [22, 18, 30, 30, 12, 28, 52, 6, 8, 46]
apply_hdr(ws2, S2_HEADERS, S2_WIDTHS, HDR2_F)

row_idx = 2
for old_mc, new_mc, chg_type, judgement, note in CHANGE_PAIRS:
    affected = bom_by_mc.get(old_mc.upper(), [])
    r_fill   = OK2_F if '즉시' in judgement else (ERR_F if '수동' in judgement or '오류' in judgement else WARN2_F)

    if not affected:
        vals = [chg_type, judgement, old_mc, new_mc, '-', '(BOM에 없음)', '-', '-', '-', note]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row_idx, c, v)
            cell.border = THIN
            cell.font   = Font(size=9)
            cell.fill   = r_fill
            cell.alignment = Alignment(vertical='center', wrap_text=(c in (1, 7, 10)))
        ws2.row_dimensions[row_idx].height = 30
        row_idx += 1
    else:
        for i, b in enumerate(sorted(affected, key=lambda x: (x.get('system') or '', x.get('iso_dwg_no') or ''))):
            vals = [
                chg_type  if i == 0 else '',
                judgement if i == 0 else '',
                old_mc    if i == 0 else '',
                new_mc    if i == 0 else '',
                b.get('system')     or '-',
                b.get('iso_dwg_no') or '-',
                (b.get('full_description') or '-').replace('_', '-'),
                b.get('uom') or 'EA',
                float(b.get('qty') or 0),
                note if i == 0 else '',
            ]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row_idx, c, v)
                cell.border = THIN
                cell.font   = Font(size=9)
                cell.fill   = r_fill
                cell.alignment = Alignment(vertical='center', wrap_text=(c in (1, 7, 10)))
            ws2.row_dimensions[row_idx].height = 28
            row_idx += 1
        # 그룹 구분선
        for c in range(1, len(S2_HEADERS) + 1):
            ws2.cell(row_idx, c).border = THIN
        row_idx += 1

wb_out.save(OUT_FILE)
print(f'\n저장 완료: {OUT_FILE}')
print(f'[Sheet1] BOM 미등록 발주항목: {len(result_rows)}건')
print(f'[Sheet2] mat_code 변경 검토: {len(CHANGE_PAIRS)}쌍')
