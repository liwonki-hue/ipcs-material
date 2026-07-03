# Valve (Receiving) 업그레이드 버전 추천 포맷 - 샘플 엑셀 생성 (v6: Tag 유무로만 구분 - Valves(관리 대상) / Untagged Items(비관리))
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

wb = openpyxl.Workbook()

header_fill = PatternFill(start_color='0A2540', end_color='0A2540', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=10)
key_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
future_fill = PatternFill(start_color='E2D9F3', end_color='E2D9F3', fill_type='solid')
note_fill = PatternFill(start_color='D9EDF7', end_color='D9EDF7', fill_type='solid')
band_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_banner(ws, text, ncols, height=60):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = Font(size=9, italic=True, color='0A2540')
    cell.fill = note_fill
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = height


def write_header(ws, headers, row=2):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


# ══════════════════════════════════════════════════════════════════════════
# Sheet 1: Valves - Tag가 있는 밸브만. 재고/부족자재 관리 대상은 이 시트뿐.
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Valves'
V_HEADERS = ['PKG', 'PKG NO', 'CATEGORY', 'TAG NO', 'VALVE TYPE', 'OPERATION TYPE',
             'ISO DWG NO', 'LINE NO', 'MAT', 'SIZE', 'RATING', 'UNIT', 'QTY', 'STATUS']
V_KEY_COLS = {'TAG NO'}
V_FUTURE_COLS = {'ISO DWG NO', 'LINE NO'}

V_NOTE = ('※ Valves 시트 = Tag가 있는 밸브만. 재고/부족자재 관리는 Tag 있는 것만 대상이므로 이 시트가 관리 대상 전체임. '
          'Tag 없이 예비 완제품으로 공급되는 밸브(Spare Valve)나 부속품/소모품/공구는 전부 Untagged Items 시트로 - 여기 섞지 않음. '
          'VALVE TYPE/OPERATION TYPE/ISO DWG NO/LINE NO는 이 밸브의 속성이므로 한 번만 기재. ISO DWG NO/LINE NO는 추후 입력 예정.')
write_banner(ws1, V_NOTE, len(V_HEADERS))
write_header(ws1, V_HEADERS)

# 각 튜플: (TAG NO, VALVE TYPE, OPERATION TYPE, ISO DWG NO, LINE NO, MAT, SIZE, RATING, UNIT, QTY, STATUS)
valve_groups = [
    ('PGU-DE-0072', 'PGU-DE-072-MOV-GLV-001', [
        ('B0-MOV-28021', 'Gate Valve', 'MOV', '', '', 'A216-WCB', '6"', 'CL150', 'EA', 1, 'On-Site'),
        ('B0-MOV-35001', 'Gate Valve', 'MOV', '', '', 'A351-CF8', '8"', 'CL150', 'EA', 1, 'On-Site'),
        ('B0-MOV-35002', 'Gate Valve', 'MOV', '', '', 'A351-CF8', '8"', 'CL150', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0125', 'PGU-DE-125-MOV-BFV-001', [
        ('B1-MOV-32001', 'Butterfly Valve', 'MOV', '', '', 'CS05', '14"', 'C150', 'EA', 1, 'On-Site'),
        ('B1-MOV-32002', 'Butterfly Valve', 'MOV', '', '', 'CS05', '14"', 'C150', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0138', 'PGU-DE-0138-BYPS-001', [
        ('B1-PV-26051', 'Globe Valve', 'AOV', '', '', '', '', '', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0138', 'PGU-DE-0138-BYPS-051', [
        ('B1/B2-PV-26051/2', 'Globe Valve', 'AOV', '', '', '', '', '', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0191', 'PGU-DE-0191-SRV-VLV-001', [
        ('B1-PSV-34006', 'Safety Valve', 'Manual', '', '', 'CS05', '6" X 8"', 'C150', 'EA', 1, ''),
        ('B1-PSV-28191', 'Safety Valve', 'Manual', '', '', 'CS05', '4" X 6"', 'C150', 'EA', 1, ''),
    ]),
    ('PGU-DE-0363', 'PGU-DE-0363-CV-001', [
        ('B2-LCV-33081', 'Globe Valve', 'AOV', '', '', '', '3"', 'CL150', 'SET', 1, 'On-Site'),
        ('B2-LCV-35083', 'Globe Valve', 'AOV', '', '', '', '3"', 'CL150', 'SET', 1, 'On-Site'),
        ('B1-LCV-34083', 'Globe Valve', 'AOV', '', '', '', '4"', 'CL150', 'SET', 1, 'On-Site'),
        ('B0-PCV-37017', 'Globe Valve', 'AOV', '', '', '', '4"', 'CL150', 'SET', 1, 'On-Site'),
    ]),
    ('PGU-DE-0390', 'PGU-DE-390-CBV-009', [
        ('B0-MV-46119', 'Ball Valve', 'Manual', '', '', '', '12"', 'CL150', 'EA', 1, ''),
        ('B0-MV-46120', 'Ball Valve', 'Manual', '', '', '', '12"', 'CL150', 'EA', 1, ''),
        ('B0-MV-46121', 'Ball Valve', 'Manual', '', '', '', '12"', 'CL150', 'EA', 1, ''),
    ]),
    ('PGU-DE-0391', 'PGU-DE-0391-CMV-001', [
        ('B1-MV-31106', 'Gate Valve', 'Manual', '', '', 'A216-WCB', '3"', 'CL600', 'EA', 1, 'On-Site'),
        ('B1-MV-31107', 'Gate Valve', 'Manual', '', '', 'A216-WCB', '3"', 'CL600', 'EA', 1, 'On-Site'),
        ('B1-MV-31110', 'Gate Valve', 'Manual', '', '', 'A216-WCC', '3"', 'CL1500', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0454', 'PGU-DE-0454-BOP-VLV-001', [
        ('B1-MV-30101A', 'Gate Valve', 'Manual', '', '', 'WCB-13CR', '10"', 'CL150', 'EA', 1, 'On-Site'),
        ('B2-MV-30101A', 'Gate Valve', 'Manual', '', '', 'WCB-13CR', '10"', 'CL150', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0474', 'PGU-DE-0474-BOP-MBV-001', [
        ('B0-MV-46164', 'Ball Valve', 'Manual', '', '', 'A105', '2"', '', 'EA', 1, ''),
        ('B0-MV-46301', 'Ball Valve', 'Manual', '', '', 'A105', '1"', 'CL150', 'EA', 1, ''),
    ]),
    ('PGU-DE-0510', 'PGU-DE-0510-BOP-VLV-001', [
        ('B1-NV-32201', 'Check Valve', 'Manual', '', '', 'WCB-304', '14"', 'CL150', 'EA', 1, ''),
        ('B2-NV-32201', 'Check Valve', 'Manual', '', '', 'WCB-304', '14"', 'CL150', 'EA', 1, ''),
    ]),
]

r = 3
band = False
for pkg, pkgno, rows in valve_groups:
    band = not band
    for row in rows:
        tag_no, vtype, optype, iso, lineno, mat, size, rating, unit, qty, status = row
        full = [pkg, pkgno, 'Valve', tag_no, vtype, optype, iso, lineno, mat, size, rating, unit, qty, status]
        for c, val in enumerate(full, start=1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = border
            if V_HEADERS[c - 1] in V_FUTURE_COLS:
                cell.fill = future_fill
            elif V_HEADERS[c - 1] in V_KEY_COLS:
                cell.fill = key_fill
            elif band:
                cell.fill = band_fill
        r += 1

v_widths = [14, 24, 10, 18, 15, 14, 14, 10, 12, 10, 9, 7, 6, 10]
for c, w in enumerate(v_widths, start=1):
    ws1.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws1.freeze_panes = 'A3'

v_comments = {
    'D': 'TAG NO [식별자]: 실제 설계 Tag. 이 시트는 재고/부족자재 관리 대상이므로 전 행에 반드시 존재. Tag 없는 항목은 여기 넣지 않음(Untagged Items 시트로).',
    'E': 'VALVE TYPE [속성]: 실제 밸브 종류. "MOV"는 여기 쓰지 않음(액추에이터 방식이지 밸브 종류가 아님).',
    'F': 'OPERATION TYPE [속성]: 조작 방식 (MOV/AOV/Manual 등).',
    'G': 'ISO DWG NO [속성 - 추후 입력]: 이 밸브가 설치되는 배관 ISO Drawing 번호.',
    'H': 'LINE NO [속성 - 추후 입력]: 이 밸브가 설치되는 배관 Line No(공정 라인 번호).',
}
for col, text in v_comments.items():
    ws1[f'{col}2'].comment = Comment(text, 'Claude')

# ══════════════════════════════════════════════════════════════════════════
# Sheet 2: Untagged Items - Tag 없는 모든 것(예비 밸브/부속품/소모품/공구). 재고/부족자재 관리 대상 아님.
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Untagged Items')
A_HEADERS = ['PKG', 'PKG NO', 'TAG NO (참조)', 'SEQ NO', 'DESCRIPTION', 'MAT', 'SIZE', 'UNIT', 'QTY', 'STATUS']
A_KEY_COLS = {'TAG NO (참조)', 'SEQ NO'}

A_NOTE = ('※ Untagged Items 시트 = Tag가 없는 모든 항목 (예비 완제품 밸브, 부속품, 소모품, 공구 - 전부 구분 없이 한 목록). '
          '재고/부족자재는 Tag 있는 것만 관리하므로 이 시트 전체는 관리 대상 아님 - 설치 시 필요한 참고 정보로만 기록. '
          '"TAG NO (참조)"는 실제로 속한 Tag를 알 때만 채우고(예: 특정 밸브의 소모품), 예비 완제품 밸브처럼 특정 Tag가 없으면 공란. '
          'SEQ NO는 TAG NO (참조)가 채워진 행에서만 사용 (01, 02...). 무엇인지는 DESCRIPTION으로 충분히 구분되므로 별도 분류 컬럼 없음.')
write_banner(ws2, A_NOTE, len(A_HEADERS), height=60)
write_header(ws2, A_HEADERS)

# 각 튜플: (TAG NO 참조, SEQ NO, DESCRIPTION, MAT, SIZE, UNIT, QTY, STATUS)
untagged_groups = [
    ('PGU-DE-0125', 'PGU-DE-125-MOV-BFV-001', [
        ('', '', 'POWER BOARD for Actuator (특정 Tag 확인 불가 - B1-MOV-32001/32002 공용 재고로 추정)', '', '', 'EA', 1, 'On-Site'),
        ('', '', 'POSITION SENSOR for Actuator (특정 Tag 확인 불가)', '', '', 'EA', 1, 'On-Site'),
        ('', '', 'OPERATING BOARD for Actuator (특정 Tag 확인 불가)', '', '', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0138', 'PGU-DE-0138-BYPS-051', [
        ('B1/B2-PV-26051/2', '01', 'Stem packing φ45/61 t=8', '', '', 'EA', 4, 'On-Site'),
        ('B1/B2-PV-26051/2', '02', 'Pressure seal φ168 φ144 t=18', '', '', 'EA', 1, 'On-Site'),
        ('', '', 'Soft and commissioning spare (특정 Tag 확인 불가, 커미셔닝 공용 예비품)', '', '', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0138', 'PGU-DE-0138-BYPS-049', [
        ('', '', 'Blow tool (Steam Blow 시운전용 - 특정 Tag 아님, PKG 내 다수 밸브 공용)', '', '', 'EA', 1, 'On-Site'),
        ('', '', 'Blow out house complete (Steam Blow 시운전용 공용 Tool)', '', '', 'EA', 1, 'On-Site'),
    ]),
    ('PGU-DE-0363', 'PGU-DE-0363-CV-019', [
        ('B0-PCV-37017', '01', 'COMMISSIONING SPARE PARTS - PACKING SET', '', '', 'SET', 2, ''),
        ('B0-PCV-37017', '02', 'COMMISSIONING SPARE PARTS - GASKET SET', '', '', 'SET', 2, ''),
        ('B0-PCV-37017', '03', 'COMMISSIONING SPARE PARTS - BALANCE SEAL', '', '', 'EA', 2, ''),
        ('B0-PCV-37017', '04', 'SPECIAL TOOLS - BLIND FLANGE', '', '', 'EA', 1, ''),
        ('B0-PCV-37017', '05', 'SPECIAL TOOLS - BONNET GASKET', '', '', 'EA', 2, ''),
    ]),
    ('PGU-DE-0474', 'PGU-DE-0474-BOP-MBV-001', [
        ('', '', 'BALL 150# DN25 FLGD-RF A105 (예비 완제품 밸브, 특정 Tag 미지정)', 'A105', '1"', 'EA', 53, ''),
        ('', '', 'BALL 150# DN50 FLGD-RF A105 (예비 완제품 밸브, 특정 Tag 미지정)', 'A105', '2"', 'EA', 5, ''),
    ]),
]

r = 3
band = False
for pkg, pkgno, rows in untagged_groups:
    band = not band
    for row in rows:
        tag_ref, seq_no, desc, mat, size, unit, qty, status = row
        full = [pkg, pkgno, tag_ref, seq_no, desc, mat, size, unit, qty, status]
        for c, val in enumerate(full, start=1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = border
            if A_HEADERS[c - 1] in A_KEY_COLS:
                cell.fill = key_fill
            elif band:
                cell.fill = band_fill
        r += 1

a_widths = [14, 24, 20, 8, 55, 12, 10, 7, 6, 10]
for c, w in enumerate(a_widths, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws2.freeze_panes = 'A3'

a_comments = {
    'C': 'TAG NO (참조) [연결]: 이 항목이 속한 Valves 시트의 TAG NO. 예비 완제품 밸브처럼 특정 Tag가 없으면 공란 (PKG NO는 이미 별도 컬럼이라 대신 넣지 않음).',
    'D': 'SEQ NO [식별자 - TAG NO (참조)가 있을 때만]: "그 Tag의 몇 번째 항목" 일련번호. TAG NO (참조)가 공란이면 SEQ NO도 공란.',
}
for col, text in a_comments.items():
    ws2[f'{col}2'].comment = Comment(text, 'Claude')

# ══════════════════════════════════════════════════════════════════════════
# Sheet 3: 컬럼 설명
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('컬럼 설명')
guide = [
    ('시트', '컬럼', '성격', '설명 / 작성 규칙'),
    ('Valves', 'PKG / PKG NO / CATEGORY', '속성', '기존과 동일'),
    ('Valves', 'TAG NO', '식별자 (필수)', '실제 설계 Tag. 이 시트는 재고/부족자재 관리 대상이므로 전 행 필수'),
    ('Valves', 'VALVE TYPE', '속성', '실제 밸브 종류 (Gate/Globe/Ball/Check/Butterfly/Safety Valve 등). "MOV"는 여기 아님'),
    ('Valves', 'OPERATION TYPE', '속성', 'MOV/AOV/Manual/Manual+Gear 등. "MOV"는 밸브 종류가 아니라 여기'),
    ('Valves', 'ISO DWG NO / LINE NO', '속성 (추후 입력)', '이 밸브가 설치되는 배관 ISO Drawing/Line No. 현재는 미확보라 공란'),
    ('Valves', 'MAT/SIZE/RATING/UNIT/QTY/STATUS', '속성', '기존과 동일'),
    ('Untagged Items', 'PKG / PKG NO', '속성', '기존과 동일 - 특정 Tag를 모를 때 소속 파악용으로도 씀'),
    ('Untagged Items', 'TAG NO (참조)', '연결', 'Valves 시트의 TAG NO를 가리킴. 특정 불가하면 공란 (PKG NO를 대신 넣지 않음)'),
    ('Untagged Items', 'SEQ NO', '식별자 (TAG NO 참조가 있을 때만)', '그 Tag 안에서 항목 순번 (01, 02...). TAG NO (참조)가 공란이면 SEQ NO도 공란'),
    ('Untagged Items', 'DESCRIPTION', '속성', '품명/규격 자유 서술. 예비 밸브/부속품/소모품/공구를 별도로 분류하지 않고 여기 텍스트로만 구분'),
    ('Untagged Items', 'MAT/SIZE/UNIT/QTY/STATUS', '속성', '기존과 동일'),
]
write_header(ws3, guide[0], row=1)
for r3, row in enumerate(guide[1:], start=2):
    for c, val in enumerate(row, start=1):
        cell = ws3.cell(row=r3, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws3.column_dimensions['A'].width = 16
ws3.column_dimensions['B'].width = 30
ws3.column_dimensions['C'].width = 26
ws3.column_dimensions['D'].width = 62
ws3.freeze_panes = 'A2'

ws3.insert_rows(1)
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
principle_cell = ws3.cell(row=1, column=1, value=
    '원칙: 재고/부족자재 관리는 Tag가 있는 것만 대상 - Valves 시트가 관리 대상 전체. '
    'Tag 없는 것(예비 완제품 밸브/부속품/소모품/공구)은 전부 Untagged Items 시트에 구분 없이 기록(설치 시 참고용, 관리 대상 아님).')
principle_cell.font = Font(size=10, bold=True, italic=True, color='0A2540')
principle_cell.fill = note_fill
principle_cell.alignment = Alignment(wrap_text=True, vertical='center')
ws3.row_dimensions[1].height = 40

# ══════════════════════════════════════════════════════════════════════════
# Sheet 4: 원본 확인 필요
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('원본 확인 필요')
issues = [
    ('구분', '내용'),
    ('중복 Tag 의심',
     'PGU-DE-0363(Control Valve)에서 동일 TAG NO가 서로 다른 Size/Rating으로 2개 패키지에 등록됨: '
     'B1-LCV-34083, B2-LCV-34083, B0-PCV-37017 (예: B0-PCV-37017 → CV-001 4"/CL150, CV-002 3"/CL600). '
     '같은 밸브가 두 번 실린 게 아니라 서로 다른 실물에 같은 Tag가 잘못 붙었을 가능성 - 원본 재확인 필요. '
     '이 템플릿의 Valves 시트에는 CV-001 쪽 값(4"/CL150)만 반영함.'),
]
write_header(ws4, issues[0], row=1)
for r4, row in enumerate(issues[1:], start=2):
    for c, val in enumerate(row, start=1):
        cell = ws4.cell(row=r4, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws4.column_dimensions['A'].width = 16
ws4.column_dimensions['B'].width = 90

# ══════════════════════════════════════════════════════════════════════════
# Sheet 5: 프로그램 활용 및 효과
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('프로그램 활용 및 효과')
usage = [
    ('항목', '현재 프로그램의 활용 방식', '이 포맷 적용 시 좋아지는 점'),
    ('Valves / Untagged Items 시트 분리',
     '재고/부족자재 계산에 Tag 있는 밸브와 부속품/소모품/공구/예비 밸브가 구분 없이 섞여 들어갈 위험이 있음',
     'Valves 시트만 재고/부족자재 관리 대상으로 명확히 분리됨 - Untagged Items는 애초에 계산 대상에서 제외되어 정확도가 올라감'),
    ('TAG NO (참조) + SEQ NO',
     'Material Finding "Item 검색"(Mode C)이 receiving.tag로 밸브/부속품을 그룹핑해 BOM/입고/재고/PKG 현황을 보여줌',
     '유니크한 tag가 만들어져 서로 무관한 부속품이 한 행으로 뭉치는 문제가 없어짐 (이미 실제 DB 1,117건에 적용·검증 완료)'),
    ('ISO DWG NO / LINE NO',
     'BOM에 없는 Valve Tag는 화면에서 "ISO 지정" 버튼으로 Tag 하나씩 수동 입력해야 함 (tag_overrides 테이블)',
     '등록 시점에 값을 일괄 반영할 수 있어 수백 건을 화면에서 클릭해서 입력하던 작업이 사라짐'),
    ('VALVE TYPE / OPERATION TYPE',
     'Item 필터 목록에 "MOV"처럼 실제 밸브 종류가 아닌 값이 밸브 종류와 섞여서 나옴',
     'Item 필터가 진짜 밸브 종류만 깔끔하게 보여주고, "MOV 밸브만 보기" 같은 조작방식 기준 필터를 새로 만들 수 있음 (예: 시운전 액추에이터 점검 대상 추출)'),
]
write_header(ws5, usage[0], row=1)
for r5, row in enumerate(usage[1:], start=2):
    for c, val in enumerate(row, start=1):
        cell = ws5.cell(row=r5, column=c, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws5.row_dimensions[r5].height = 45
ws5.column_dimensions['A'].width = 24
ws5.column_dimensions['B'].width = 55
ws5.column_dimensions['C'].width = 65
ws5.freeze_panes = 'A2'

OUT = r'Raw File\Valve (Receiving)_Format_Template.xlsx'
wb.save(OUT)
print('saved:', OUT)
