# Piping BOM vs Receiving 매칭 분석 → 미입고/부족 항목 엑셀 출력 (Support 제외)
import requests, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Accept-Profile': 'material'}

OUT_PATH = 'Raw File/BOM_vs_Receiving.xlsx'

# ── 1. BOM 전체 로드 (페이지네이션) ──
print('Loading BOM...')
bom_rows = []
step = 1000
offset = 0
while True:
    r = requests.get(f'{URL}/rest/v1/bom', headers=H, params={
        'select': 'mat_code,category,full_description,uom,qty',
        'limit': step, 'offset': offset
    })
    batch = r.json()
    if not batch:
        break
    bom_rows.extend(batch)
    offset += step
    if len(batch) < step:
        break
print(f'  BOM rows: {len(bom_rows):,}')

# ── 2. Receiving 전체 로드 ──
print('Loading Receiving...')
rec_rows = []
offset = 0
while True:
    r = requests.get(f'{URL}/rest/v1/receiving', headers=H, params={
        'select': 'mat_code,category,qty,full_description,pkg_no',
        'limit': step, 'offset': offset
    })
    batch = r.json()
    if not batch:
        break
    rec_rows.extend(batch)
    offset += step
    if len(batch) < step:
        break
print(f'  Receiving rows: {len(rec_rows):,}')

# ── 3. mat_code 기준으로 집계 ──
# BOM: mat_code → {category, description, uom, total_qty}
bom_agg = {}
for b in bom_rows:
    mc = b['mat_code'] or ''
    if not mc:
        continue
    if mc not in bom_agg:
        bom_agg[mc] = {
            'mat_code': mc,
            'category': b['category'] or '',
            'description': b['full_description'] or '',
            'uom': b['uom'] or 'EA',
            'bom_qty': 0.0,
        }
    bom_agg[mc]['bom_qty'] += float(b['qty'] or 0)

# Receiving: mat_code → total received qty
rec_agg = defaultdict(float)
for r in rec_rows:
    mc = r['mat_code'] or ''
    if mc:
        rec_agg[mc] += float(r['qty'] or 0)

# ── 4. 매칭 분류 ──
not_received  = []  # BOM에 있으나 Receiving 전혀 없음
short         = []  # 일부 입고 (rec < bom)
full          = []  # 완전 입고 (rec >= bom)

for mc, b in sorted(bom_agg.items(), key=lambda x: (x[1]['category'], x[0])):
    rec_qty = rec_agg.get(mc, 0.0)
    shortage = b['bom_qty'] - rec_qty
    row = {**b, 'rec_qty': rec_qty, 'shortage': shortage}
    if rec_qty == 0:
        not_received.append(row)
    elif shortage > 0.001:
        short.append(row)
    else:
        full.append(row)

print(f'\n결과 (mat_code 기준):')
print(f'  미입고:   {len(not_received):,}건')
print(f'  부분입고: {len(short):,}건')
print(f'  완전입고: {len(full):,}건')
print(f'  전체:     {len(bom_agg):,}건')

# ── 5. 엑셀 스타일 정의 ──
HEADER_FILL  = PatternFill('solid', fgColor='1F3864')
RED_FILL     = PatternFill('solid', fgColor='FFCCCC')
YELLOW_FILL  = PatternFill('solid', fgColor='FFF2CC')
GREEN_FILL   = PatternFill('solid', fgColor='E2EFDA')
header_font  = Font(bold=True, color='FFFFFF', size=10)
bold         = Font(bold=True, size=10)
normal       = Font(size=10)
center       = Alignment(horizontal='center', vertical='center')
thin         = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))

COLS      = ['CATEGORY', 'MAT CODE', 'DESCRIPTION', 'UOM', 'BOM QTY', 'REC QTY', 'SHORTAGE']
COL_WIDTHS = [12, 30, 50, 8, 12, 12, 12]

def write_header(ws):
    for col, (h, w) in enumerate(zip(COLS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 18

def write_data(ws, rows, fill):
    for r_no, row in enumerate(rows, 2):
        vals = [row['category'], row['mat_code'], row['description'],
                row['uom'], row['bom_qty'], row['rec_qty'], row['shortage']]
        for c_no, val in enumerate(vals, 1):
            cell = ws.cell(row=r_no, column=c_no, value=val)
            cell.fill = fill
            cell.font = normal
            cell.alignment = center
            cell.border = thin
            if c_no in (5, 6, 7) and isinstance(val, float):
                cell.number_format = '#,##0.00'

wb = openpyxl.Workbook()

# Summary
ws_sum = wb.active
ws_sum.title = 'Summary'
summary = [
    ('항목', '건수(mat_code)', '비고'),
    ('BOM 전체 (unique mat_code)', len(bom_agg), ''),
    ('Receiving 전체 (unique mat_code)', len(rec_agg), ''),
    ('', '', ''),
    ('미입고 (Receiving 없음)', len(not_received), ''),
    ('부분입고 (부족)', len(short), ''),
    ('완전입고 (충족)', len(full), ''),
]
fills_sum = [HEADER_FILL, None, None, None, RED_FILL, YELLOW_FILL, GREEN_FILL]
for r_no, (row, fill) in enumerate(zip(summary, fills_sum), 1):
    for c_no, val in enumerate(row, 1):
        cell = ws_sum.cell(row=r_no, column=c_no, value=val)
        cell.border = thin
        cell.alignment = center
        if fill:
            cell.fill = fill
            cell.font = header_font if fill == HEADER_FILL else bold
        else:
            cell.font = normal
ws_sum.column_dimensions['A'].width = 35
ws_sum.column_dimensions['B'].width = 18
ws_sum.column_dimensions['C'].width = 18

# 미입고
ws1 = wb.create_sheet('미입고 (Receiving 없음)')
write_header(ws1)
write_data(ws1, not_received, RED_FILL)

# 부분입고
ws2 = wb.create_sheet('부분입고 (부족)')
write_header(ws2)
write_data(ws2, short, YELLOW_FILL)

# 완전입고
ws3 = wb.create_sheet('완전입고')
write_header(ws3)
write_data(ws3, full, GREEN_FILL)

wb.save(OUT_PATH)
print(f'\n저장 완료: {OUT_PATH}')
