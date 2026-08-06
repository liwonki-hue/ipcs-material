# Valve (Receiving)_Format_Template.xlsx(Valves/Untagged Items) 기준 receiving category='Valve' 전체 재적재
import openpyxl
import psycopg2
import psycopg2.extras

DB_URL = open('.env').read().strip().split('=', 1)[1]
XLSX_PATH = 'Raw File/Valve (Receiving)_Format_Template.xlsx'

# PGU-DE-0593의 B1-MV-29109/B2-MV-29109 — PGU-DE-0454와 중복되는 Tag, 사용자 지시로 임시 -01/-02 접미사
DUP_TAG_PKG = 'PGU-DE-0593'
DUP_TAG_SUFFIX = {'B1-MV-29109': '01', 'B2-MV-29109': '02'}

conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

# --- BOM tag -> (mat1, mat2) 참조맵 (BOM 우선 원칙, project_valve_bom_registration 관례 재사용) ---
cur.execute("SELECT tag, mat1, mat2 FROM bom WHERE category='Valve'")
bom_map = {row[0]: (row[1], row[2]) for row in cur.fetchall()}
print(f'BOM Valve tag map: {len(bom_map)}건')

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# ============ Valves 시트 ============
ws = wb['Valves']
valve_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
print(f'Valves 시트: {len(valve_rows)}행')

blank_tag_seq = {}  # pkg_no -> counter, 공란 TAG NO 임시 번호용
out_rows = []
bom_backfilled = 0

for r in valve_rows:
    pkg, pkg_no, raw_tag, op_type, vtype, mat1, mat2, size, rating, conn_type, unit, qty = r
    vtype = (vtype or '').strip()

    if raw_tag:
        raw_tag = raw_tag.strip()
        if pkg == DUP_TAG_PKG and raw_tag in DUP_TAG_SUFFIX:
            tag = f'{raw_tag}-{DUP_TAG_SUFFIX[raw_tag]}'
            parent_tag = raw_tag
        else:
            tag = raw_tag
            parent_tag = raw_tag
    else:
        # TAG NO 공란 — PKG NO 기준 임시 Tag 부여 (사용자 확인, 추후 실제 Tag 확인 필요)
        n = blank_tag_seq.get(pkg_no, 0) + 1
        blank_tag_seq[pkg_no] = n
        tag = f'{pkg_no}-{n:02d}'
        parent_tag = pkg_no
        raw_tag = None

    # BOM 우선 원칙: BOM에 해당 raw tag가 있으면 mat1/mat2를 BOM 값으로 덮어씀
    if raw_tag and raw_tag in bom_map:
        b_mat1, b_mat2 = bom_map[raw_tag]
        if b_mat1:
            mat1 = b_mat1
        if b_mat2:
            mat2 = b_mat2
        bom_backfilled += 1

    full_desc = f'{vtype} VALVE, {mat2 or "-"}, {size or "-"}, {rating or "-"}, {conn_type or "-"}'

    out_rows.append({
        'doc_no': pkg, 'pkg_no': pkg_no, 'tag': tag, 'parent_tag': parent_tag,
        'op_type': op_type, 'valve_type': vtype or None,
        'mat1': mat1, 'mat2': mat2, 'size': size, 'rating': rating,
        'full_description': full_desc, 'unit': unit or 'EA', 'qty': qty or 1,
    })

print(f'BOM 매칭으로 mat1/mat2 보정된 행: {bom_backfilled}')

# ============ Untagged Items 시트 ============
ws2 = wb['Untagged Items']
untagged_rows = [r for r in ws2.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r[:8])]
print(f'Untagged Items 시트: {len(untagged_rows)}행')

acc_seq = {}  # pkg_no -> counter (Tag 유니크화용, 원본 중복 행 포함 그대로 반영)

for r in untagged_rows:
    pkg, pkg_no, ref_tag, seq_no, item, desc, unit, qty = r[:8]
    item = (item or '').strip() or None
    ref_tag = (ref_tag or '').strip() or None

    n = acc_seq.get(pkg_no, 0) + 1
    acc_seq[pkg_no] = n
    # ITEM='SPARE'는 기존 Spare 탭 분류 로직(isSpareBodyRow, tag에 'SPARE' 포함 여부로 판별)과의
    # 호환을 위해 태그에 SPARE를 유지 — 그 외는 일반 부속 항목(ACC)으로 접미사
    infix = 'SPARE' if item and item.upper() == 'SPARE' else 'ACC'
    tag = f'{pkg_no}-{infix}-{n:03d}'
    parent_tag = ref_tag or pkg_no

    out_rows.append({
        'doc_no': pkg, 'pkg_no': pkg_no, 'tag': tag, 'parent_tag': parent_tag,
        'op_type': None, 'valve_type': item,
        'mat1': None, 'mat2': None, 'size': None, 'rating': None,
        'full_description': desc, 'unit': unit or 'EA', 'qty': qty or 1,
    })

print(f'총 적재 대상 행: {len(out_rows)}')

# --- Tag 유니크성 최종 검증 ---
tags = [r['tag'] for r in out_rows]
assert len(tags) == len(set(tags)), f'Tag 중복 발견: {len(tags) - len(set(tags))}건'
print('Tag 100% 유니크 확인 완료')

# ============ DB 반영 ============
cur.execute("SELECT COUNT(*) FROM receiving WHERE category='Valve'")
before = cur.fetchone()[0]
print(f'기존 receiving category=Valve 행 수: {before}')

cur.execute("DELETE FROM receiving WHERE category='Valve'")
print(f'삭제된 행: {cur.rowcount}')

cur.execute("SELECT COALESCE(MAX(id), 0) FROM receiving")
next_id = cur.fetchone()[0] + 1

insert_sql = """
    INSERT INTO receiving (id, doc_no, pkg_no, category, tag, parent_tag, op_type, valve_type,
                            mat1, mat2, size, rating, full_description, unit, qty)
    VALUES %s
"""
values = []
for r in out_rows:
    values.append((
        next_id, r['doc_no'], r['pkg_no'], 'Valve', r['tag'], r['parent_tag'],
        r['op_type'], r['valve_type'], r['mat1'], r['mat2'], r['size'], r['rating'],
        r['full_description'], r['unit'], r['qty']
    ))
    next_id += 1

psycopg2.extras.execute_values(cur, insert_sql, values)
conn.commit()
print(f'삽입된 행: {len(values)}')

cur.execute("SELECT COUNT(*), COUNT(DISTINCT tag) FROM receiving WHERE category='Valve'")
total, distinct = cur.fetchone()
print(f'최종 확인: 총 {total}행, Tag 유니크 {distinct}개')

cur.close()
conn.close()
