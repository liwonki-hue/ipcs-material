# TAG 항목 BOM vs Receiving 매칭 분석 (tag IS NOT NULL인 항목만)
import requests, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Accept-Profile': 'material'}

OUT_PATH = 'Raw File/TAG_BOM_vs_Receiving.xlsx'

# ── 1. BOM: tag IS NOT NULL 항목 전체 로드 ──
print('Loading BOM (tag 항목만)...')
bom_rows = []
step = 1000
offset = 0
while True:
    r = requests.get(f'{URL}/rest/v1/bom', headers=H, params={
        'select': 'tag,category,mat_code,full_description,uom,qty,system,iso_dwg_no',
        'tag': 'not.is.null',
        'limit': step, 'offset': offset
    })
    batch = r.json()
    if not batch:
        break
    bom_rows.extend(batch)
    offset += step
    if len(batch) < step:
        break
print(f'  BOM tag 항목: {len(bom_rows):,}건')

# ── 2. Receiving 전체 로드 ──
print('Loading Receiving...')
rec_rows = []
offset = 0
while True:
    r = requests.get(f'{URL}/rest/v1/receiving', headers=H, params={
        'select': 'tag,mat_code,qty,category,doc_no,pkg_no,full_description',
        'limit': step, 'offset': offset
    })
    batch = r.json()
    if not batch:
        break
    rec_rows.extend(batch)
    offset += step
    if len(batch) < step:
        break
print(f'  Receiving rows: {len(rec_rows):,}건')

# ── 3. Receiving을 tag 기준으로 인덱싱 ──
rec_by_tag = defaultdict(list)
for r in rec_rows:
    tag = (r['tag'] or '').strip()
    if tag and tag != 'BULK':
        rec_by_tag[tag].append(r)

print(f'  Receiving 고유 tag(BULK 제외): {len(rec_by_tag):,}개')

# ── 4. BOM 각 행 매칭 분류 ──
not_received = []   # Receiving에 tag 없음
matched      = []   # Receiving에 tag 있음

for b in bom_rows:
    tag = (b['tag'] or '').strip()
    bom_qty = float(b['qty'] or 0)
    rec_list = rec_by_tag.get(tag, [])
    rec_qty  = sum(float(r['qty'] or 0) for r in rec_list)

    row = {
        'tag':         tag,
        'category':    b['category'] or '',
        'mat_code':    b['mat_code'] or '',
        'description': b['full_description'] or '',
        'uom':         b['uom'] or 'EA',
        'system':      b['system'] or '',
        'iso_dwg_no':  b['iso_dwg_no'] or '',
        'bom_qty':     bom_qty,
        'rec_qty':     rec_qty,
        'shortage':    bom_qty - rec_qty,
        'rec_doc':     ', '.join(set(r['doc_no'] for r in rec_list if r.get('doc_no'))) if rec_list else '',
        'rec_pkg':     ', '.join(set(r['pkg_no'] for r in rec_list if r.get('pkg_no'))) if rec_list else '',
    }
    if rec_qty == 0:
        not_received.append(row)
    else:
        matched.append(row)

print(f'\n결과:')
print(f'  미입고: {len(not_received):,}건')
print(f'  입고됨: {len(matched):,}건')
print(f'  전체:   {len(bom_rows):,}건')

# ── 5. 엑셀 출력 ──
HEADER_FILL = PatternFill('solid', fgColor='1F3864')
RED_FILL    = PatternFill('solid', fgColor='FFCCCC')
GREEN_FILL  = PatternFill('solid', fgColor='E2EFDA')
header_font = Font(bold=True, color='FFFFFF', size=10)
bold        = Font(bold=True, size=10)
normal      = Font(size=10)
center      = Alignment(horizontal='center', vertical='center')
thin        = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

COLS = ['TAG', 'CATEGORY', 'MAT CODE', 'DESCRIPTION', 'SYSTEM', 'UOM', 'BOM QTY', 'REC QTY', 'SHORTAGE', 'DOC NO', 'PKG NO']
WIDTHS = [22, 12, 28, 45, 12, 8, 10, 10, 10, 22, 30]

def write_header(ws):
    for col, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 18

def write_data(ws, rows, fill):
    for r_no, row in enumerate(rows, 2):
        vals = [row['tag'], row['category'], row['mat_code'], row['description'],
                row['system'], row['uom'], row['bom_qty'], row['rec_qty'],
                row['shortage'], row['rec_doc'], row['rec_pkg']]
        for c_no, val in enumerate(vals, 1):
            cell = ws.cell(row=r_no, column=c_no, value=val)
            cell.fill = fill
            cell.font = normal
            cell.alignment = center
            cell.border = thin
            if c_no in (7, 8, 9) and isinstance(val, float):
                cell.number_format = '#,##0.00'

wb = openpyxl.Workbook()

# Summary
ws_sum = wb.active
ws_sum.title = 'Summary'
summary = [
    ('항목', '건수', ''),
    ('BOM TAG 전체', len(bom_rows), ''),
    ('미입고 (Receiving 없음)', len(not_received), ''),
    ('입고됨', len(matched), ''),
]
sum_fills = [HEADER_FILL, None, RED_FILL, GREEN_FILL]
for r_no, (row, fill) in enumerate(zip(summary, sum_fills), 1):
    for c_no, val in enumerate(row, 1):
        cell = ws_sum.cell(row=r_no, column=c_no, value=val)
        cell.border = thin
        cell.alignment = center
        if fill:
            cell.fill = fill
            cell.font = header_font if fill == HEADER_FILL else bold
        else:
            cell.font = normal
ws_sum.column_dimensions['A'].width = 28
ws_sum.column_dimensions['B'].width = 12

# 미입고
ws1 = wb.create_sheet('미입고')
write_header(ws1)
write_data(ws1, sorted(not_received, key=lambda x: (x['category'], x['tag'])), RED_FILL)

# 입고됨
ws2 = wb.create_sheet('입고됨')
write_header(ws2)
write_data(ws2, sorted(matched, key=lambda x: (x['category'], x['tag'])), GREEN_FILL)

wb.save(OUT_PATH)
print(f'\n저장 완료: {OUT_PATH}')
