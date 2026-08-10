# Speciality BOM<->Receiving Tag 불일치 항목 확인용 Excel 생성 (검토용, DB 변경 없음)
import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB_URL = open('.env').read().strip().split('=', 1)[1]
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

cur.execute("SELECT tag FROM receiving WHERE category='Speciality' AND parent_tag=tag")
recv_tags = set(r[0] for r in cur.fetchall())

cur.execute("""
    SELECT tag, system, iso_dwg_no, line_no, full_description, mat1, mat2, uom, qty
    FROM bom WHERE category='Speciality' ORDER BY tag
""")
bom_rows = cur.fetchall()
bom_only = [r for r in bom_rows if r[0] not in recv_tags]

cur.execute("SELECT tag FROM bom WHERE category='Speciality'")
bom_tags = set(r[0] for r in cur.fetchall())

cur.execute("""
    SELECT doc_no, pkg_no, tag, full_description, mat1, mat2, size, rating, unit, qty
    FROM receiving WHERE category='Speciality' AND parent_tag=tag ORDER BY tag
""")
recv_rows = cur.fetchall()
recv_only = [r for r in recv_rows if r[2] not in bom_tags]

cur.close()
conn.close()

print(f'BOM Only(Receiving 미등록): {len(bom_only)}건')
print(f'Receiving Only(BOM 미등록): {len(recv_only)}건')

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill('solid', fgColor='0A2540')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'


ws1 = wb.active
ws1.title = 'BOM Only (미입고)'
headers1 = ['TAG', 'SYSTEM', 'ISO DWG NO', 'LINE NO', 'DESCRIPTION', 'MAT1', 'MAT2', 'UOM', 'QTY']
ws1.append(headers1)
for r in bom_only:
    tag, system, iso, line_no, desc, mat1, mat2, uom, qty = r
    ws1.append([tag, system, iso, line_no, desc, mat1, mat2, uom, float(qty) if qty is not None else None])
style_header(ws1, len(headers1))
for i, w in enumerate([20, 14, 20, 14, 45, 10, 12, 8, 8], start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws2 = wb.create_sheet('Receiving Only (BOM 미등록)')
headers2 = ['TAG', 'PKG', 'PKG NO', 'DESCRIPTION', 'MAT1', 'MAT2', 'SIZE', 'RATING', 'UNIT', 'QTY']
ws2.append(headers2)
for r in recv_only:
    doc_no, pkg_no, tag, desc, mat1, mat2, size, rating, unit, qty = r
    ws2.append([tag, doc_no, pkg_no, desc, mat1, mat2, size, rating, unit, float(qty) if qty is not None else None])
style_header(ws2, len(headers2))
for i, w in enumerate([20, 16, 24, 45, 10, 12, 10, 10, 8, 8], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

OUT_PATH = 'scratch/Speciality_Tag_Mismatch.xlsx'
wb.save(OUT_PATH)
print(f'저장 완료: {OUT_PATH}')
