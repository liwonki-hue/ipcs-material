# 편집된 0574 PL 파일 기준으로 DB receiving 검증 후 pl_db_comparison.xlsx 업데이트
import sys, io, re, requests
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from collections import defaultdict

SUPABASE_URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
HEADERS = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}', 'Accept-Profile': 'material'}

PL_PATH      = 'Raw File/PGU-DE-0574_BOP Piping_General System_(편집).xlsx'
COMPARE_PATH = 'scratch/pl_db_comparison.xlsx'

ITEM_PREFIX = {
    'PIPE':          ['PIS', 'PIW', 'PIN'],
    'ELBOW LR 90':   ['EL9L'],
    'ELBOW SR 90':   ['EL9S'],
    'ELBOW LR 45':   ['EL4L'],
    'ELBOW 90':      ['EL9L', 'EL9S'],
    'ELBOW 45':      ['EL4L'],
    'ELBOW':         ['EL9L', 'EL9S', 'EL4L'],
    'TEE-RED':       ['TER'],
    'TEE':           ['TEE', 'TER'],
    'REDUCER':       ['RDC', 'RDE'],
    'CAP':           ['CAP'],
    'FLANGE-BLIND':  ['FLB'],
    'FLANGE':        ['FLN', 'FLA', 'FLB', 'FLS'],
    'COUPLING-FULL': ['CPF'],
    'COUPLING-HALF': ['CPH'],
    'COUPLING':      ['CPF', 'CPH'],
    'WELDOLET':      ['WOL'],
    'SOCKOLET':      ['SOL'],
    'SWAGED CON':    ['SCN'],
    'SWAGE':         ['SCN'],
    'NOZZLE':        ['NOZ'],
}

def check_matcode(desc, mat_code):
    if not mat_code:
        return 'NULL'
    desc_u = desc.upper()
    for kw, prefixes in ITEM_PREFIX.items():
        if kw in desc_u:
            if any(mat_code.startswith(p) for p in prefixes):
                return 'OK'
            return 'MISMATCH'
    return 'OK'

def norm_desc(s):
    return re.sub(r'\s+', ' ', str(s).upper().strip())

def norm_unit(u):
    if not u:
        return ''
    s = str(u).strip().upper()
    if 'EA' in s:
        return 'EA'
    if s == 'M':
        return 'M'
    return s

def is_pipe(desc):
    d = str(desc).upper().strip()
    return d.startswith('PIPE') and 'NIPPLE' not in d

# ── 편집 PL 파일 파싱 ─────────────────────────────────────────────────
# 컬럼: col0=No, col1=Package No, col2=Description, col3=Q'ty, col4=Unit
# PKG NO: FDR → DB에서는 FRD 로 저장됨
print('편집 PL 파일 파싱...')
wb = openpyxl.load_workbook(PL_PATH, read_only=True)
ws = wb.active

# (pkg_no, norm_desc) 기준으로 수량 합산 (중복 행 처리)
pl_grouped = defaultdict(lambda: {'desc': '', 'qty_pl': 0.0, 'unit_pl': 'EA'})

for row in ws.iter_rows(min_row=2, values_only=True):
    pkg_raw = row[1]
    desc    = row[2]
    qty     = row[3]
    if not pkg_raw or not desc or not isinstance(qty, (int, float)):
        continue
    # FDR → FRD 변환 + 공백 제거
    pkg_db = str(pkg_raw).strip().replace('-BOP-FDR-', '-BOP-FRD-')
    key = (pkg_db, norm_desc(desc))
    pl_grouped[key]['desc']   = str(desc).strip()
    pl_grouped[key]['qty_pl'] += float(qty)
    pl_grouped[key]['unit_pl'] = norm_unit(row[4]) or 'EA'

print(f'  → {len(pl_grouped)}개 항목 (PKG별 중복 합산 완료)')

# ── DB 0574 receiving 조회 ───────────────────────────────────────────
print('DB 조회...')
r = requests.get(
    f'{SUPABASE_URL}/rest/v1/receiving?select=id,pkg_no,mat_code,qty,unit,full_description'
    f'&doc_no=eq.PGU-DE-0574&order=pkg_no,id',
    headers=HEADERS
)
db_rows = r.json()
print(f'  → {len(db_rows)}행')

# DB dict: (pkg_no, norm_desc) → row
db_by_key = {}
for row in db_rows:
    key = (row['pkg_no'], norm_desc(row['full_description'] or ''))
    db_by_key[key] = row

# ── 비교 ──────────────────────────────────────────────────────────────
results = []

# PL 항목 → DB 매칭
pl_matched_keys = set()
db_matched_keys = set()

for (pkg_db, pl_desc_n), pl_info in sorted(pl_grouped.items()):
    pl_qty  = pl_info['qty_pl']
    pl_unit = pl_info['unit_pl']
    pipe    = is_pipe(pl_info['desc'])
    pl_qty_db = pl_qty * 6.0 if pipe else pl_qty  # M 환산
    db_unit_expect = 'M' if pipe else 'EA'

    # 완전 일치 우선 시도
    exact_key = (pkg_db, pl_desc_n)
    if exact_key in db_by_key:
        dr = db_by_key[exact_key]
        db_qty  = dr['qty'] or 0
        db_unit = norm_unit(dr['unit'])
        mat_code = dr['mat_code'] or ''
        mc_status = check_matcode(pl_info['desc'], mat_code)
        qty_diff = round(db_qty - pl_qty_db, 4)
        status   = 'OK' if abs(qty_diff) < 0.01 else ('DB↑' if qty_diff > 0 else 'DB↓')
        if mc_status != 'OK':
            status = mc_status + ('/' + status if status != 'OK' else '')
        pl_matched_keys.add(exact_key)
        db_matched_keys.add(exact_key)
        results.append({
            'pkg_no': pkg_db, 'pl_desc': pl_info['desc'],
            'pl_qty': pl_qty, 'pl_unit': pl_unit,
            'pl_qty_db': pl_qty_db, 'db_mat_code': mat_code,
            'db_qty': db_qty, 'db_unit': db_unit,
            'qty_diff': qty_diff, 'mc_status': mc_status, 'status': status
        })
        continue

    # 단어 유사도 fallback
    pl_words = set(pl_desc_n.split())
    best_key, best_score = None, 0
    for dk in db_by_key:
        if dk in db_matched_keys or dk[0] != pkg_db:
            continue
        db_words = set(dk[1].split())
        score = len(pl_words & db_words)
        if score > best_score:
            best_score, best_key = score, dk

    if best_key and best_score >= 3:
        dr = db_by_key[best_key]
        db_qty  = dr['qty'] or 0
        db_unit = norm_unit(dr['unit'])
        mat_code = dr['mat_code'] or ''
        mc_status = check_matcode(pl_info['desc'], mat_code)
        qty_diff = round(db_qty - pl_qty_db, 4)
        status   = 'OK' if abs(qty_diff) < 0.01 else ('DB↑' if qty_diff > 0 else 'DB↓')
        if mc_status != 'OK':
            status = mc_status + ('/' + status if status != 'OK' else '')
        pl_matched_keys.add((pkg_db, pl_desc_n))
        db_matched_keys.add(best_key)
        results.append({
            'pkg_no': pkg_db, 'pl_desc': pl_info['desc'],
            'pl_qty': pl_qty, 'pl_unit': pl_unit,
            'pl_qty_db': pl_qty_db, 'db_mat_code': mat_code,
            'db_qty': db_qty, 'db_unit': db_unit,
            'qty_diff': qty_diff, 'mc_status': mc_status, 'status': status
        })
    else:
        results.append({
            'pkg_no': pkg_db, 'pl_desc': pl_info['desc'],
            'pl_qty': pl_qty, 'pl_unit': pl_unit,
            'pl_qty_db': pl_qty_db, 'db_mat_code': '',
            'db_qty': '', 'db_unit': '',
            'qty_diff': '', 'mc_status': '', 'status': 'DB없음'
        })

# DB에만 있고 PL에 없는 항목
for dk, dr in sorted(db_by_key.items(), key=lambda x: x[0]):
    if dk in db_matched_keys:
        continue
    mat_code  = dr['mat_code'] or ''
    mc_status = check_matcode(dr['full_description'] or '', mat_code)
    results.append({
        'pkg_no': dr['pkg_no'], 'pl_desc': '(PL 미매칭)',
        'pl_qty': '', 'pl_unit': '',
        'pl_qty_db': '', 'db_mat_code': mat_code,
        'db_qty': dr['qty'], 'db_unit': norm_unit(dr['unit']),
        'qty_diff': '', 'mc_status': mc_status, 'status': 'PL없음'
    })

print(f'\n0574 비교 결과: {len(results)}행')
ok = sum(1 for r in results if r['status'] == 'OK')
diff = sum(1 for r in results if 'DB↑' in str(r['status']) or 'DB↓' in str(r['status']))
null_ = sum(1 for r in results if 'NULL' in str(r['status']))
mism = sum(1 for r in results if 'MISMATCH' in str(r['status']))
dbmiss = sum(1 for r in results if r['status'] == 'DB없음')
plmiss = sum(1 for r in results if r['status'] == 'PL없음')
print(f'  OK={ok}, 수량불일치={diff}, MatCode NULL={null_}, MISMATCH={mism}, DB없음={dbmiss}, PL없음={plmiss}')
for r in results:
    s = r['status']
    if s != 'OK':
        print(f'  [{s}] {r["pkg_no"]} | {str(r["pl_desc"])[:45]} | PL_DB={r["pl_qty_db"]} / DB={r["db_qty"]} {r["db_unit"]} | mat={r["db_mat_code"]}')

# ── pl_db_comparison.xlsx 업데이트 ──────────────────────────────────
print(f'\npl_db_comparison.xlsx 업데이트...')

FILLS = {
    'OK':       PatternFill('solid', fgColor='C6EFCE'),
    'DB↑':      PatternFill('solid', fgColor='FFEB9C'),
    'DB↓':      PatternFill('solid', fgColor='FFC7CE'),
    'DB없음':   PatternFill('solid', fgColor='FFC7CE'),
    'PL없음':   PatternFill('solid', fgColor='DDEBF7'),
    'NULL':     PatternFill('solid', fgColor='FFC7CE'),
    'MISMATCH': PatternFill('solid', fgColor='FF9900'),
}

def row_fill(status):
    s = str(status)
    if s == 'OK': return FILLS['OK']
    if 'DB↑' in s: return FILLS['DB↑']
    if 'DB↓' in s: return FILLS['DB↓']
    if 'NULL' in s: return FILLS['NULL']
    if 'MISMATCH' in s: return FILLS['MISMATCH']
    if s == 'DB없음': return FILLS['DB없음']
    if s == 'PL없음': return FILLS['PL없음']
    return None

wb_cmp = openpyxl.load_workbook(COMPARE_PATH)
ws_cmp = wb_cmp['항목별 비교']

# 기존 0574 행 번호 수집 (헤더=1)
rows_to_delete = []
for i in range(2, ws_cmp.max_row + 1):
    val = ws_cmp.cell(row=i, column=1).value
    if val and '0574' in str(val):
        rows_to_delete.append(i)

# 뒤에서부터 삭제
for i in reversed(rows_to_delete):
    ws_cmp.delete_rows(i)

print(f'  기존 0574 행 {len(rows_to_delete)}행 삭제 완료')

# 새 0574 결과 추가 (기존 파일 맨 끝에 삽입)
for r in results:
    row_vals = [
        r['pkg_no'], r['pl_desc'],
        r['pl_qty'], r['pl_unit'],
        r['pl_qty_db'], r['db_mat_code'],
        r['db_qty'], r['db_unit'],
        r['qty_diff'], r['mc_status'], r['status']
    ]
    ws_cmp.append(row_vals)
    fill = row_fill(r['status'])
    if fill:
        for cell in ws_cmp[ws_cmp.max_row]:
            cell.fill = fill

print(f'  새 0574 결과 {len(results)}행 추가 완료')

# 수정 방안 열(12번째)이 있는지 확인 후 추가
header_row = [ws_cmp.cell(row=1, column=c).value for c in range(1, ws_cmp.max_column + 1)]
if len(header_row) < 12 or header_row[11] != '수정 방안':
    print('  수정 방안 열 없음 → 스킵 (add_suggestion_col.py 별도 실행 필요)')

wb_cmp.save(COMPARE_PATH)
print(f'저장 완료: {COMPARE_PATH}')
