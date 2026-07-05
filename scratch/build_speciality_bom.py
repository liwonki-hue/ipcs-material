# Speciality.xlsx 11개 카테고리 원본 데이터를 Speciality BOM.xlsx의 VALVELIST_BOP 포맷으로 추출
import openpyxl

SRC = 'Raw File/Speciality.xlsx'
DST = 'Raw File/Speciality BOM.xlsx'

MAT1_GRADE = {
    'A106-B': 'CS', 'A53-B': 'CS', 'A672-B60-CL.22': 'CS',
    'A234-WPB': 'CS', 'A216-WCB': 'CS',
    'A312-TP304': 'SS', 'A312-TP304_S': 'SS', '316SS': 'SS',
    'SUS304': 'SS', 'STAINLESS STEEL': 'SS',
    'A335-P91': 'ALLOY (P91)',
    'RUBBER': 'RUBBER', 'ALUMINUM': 'ALUMINUM', 'METALLIC': 'METALLIC',
}

def classify_mat1(mat2):
    if mat2 is None:
        return None
    key = str(mat2).strip().upper()
    for k, v in MAT1_GRADE.items():
        if k.upper() == key:
            return v
    return None

# sheet_name: (line_no_col, tag_col, qty_col, item_label, body_mat_col, pipe_mat_col, size_col, end_type_col, rating_col)
CFG = {
    '1.STRAINER':             (2, 20, 21, 'STRAINER',             23, 6, 7, 24, 25),
    '2. STEAM TRAP':          (2, 20, 21, 'STEAM TRAP',           23, 6, 7, 24, 25),
    '3. EXPANSION JOINT':     (2, 20, 21, 'EXPANSION JOINT',      23, 6, 7, 24, 25),
    '4. SIGHT GLASS':         (2, 20, 21, 'SIGHT GLASS',          23, 6, 7, 24, 25),
    '5. FLEXIBLE JOINT':      (2, 20, 21, 'FLEXIBLE JOINT',       23, 6, 7, 24, 25),
    '6. RESTRICTION ORIFICE': (2, 21, 22, 'RESTRICTION ORIFICE',  24, 6, 7, 25, 26),
    '7. FLEXIBLE HOSE':       (2, 20, 21, 'FLEXIBLE HOSE',        23, 6, 7, 25, None),
    '8. SPRAY NOZZLE':        (2, 20, 21, 'SPRAY NOZZLE',         23, 6, 7, 26, None),
    '10.AIR TRAP':            (2, 20, 21, 'AIR TRAP',             23, 6, 7, 24, 25),
    '11.BIRD SCREEN':         (2, 19, 20, 'BIRD SCREEN',          22, 6, 7, 25, 26),
}

SAME_AS_PIPE = 'same or equivalent material w/ pipe'

wb = openpyxl.load_workbook(SRC, data_only=True)
rows = []

for sheet_name, (lc, tc, qc, item, bmc, pmc, sc, ec, rc) in CFG.items():
    ws = wb[sheet_name]
    for r in range(5, ws.max_row + 1):
        tag = ws.cell(r, tc).value
        if tag is None or str(tag).strip() == '':
            continue
        line_no = ws.cell(r, lc).value
        qty = ws.cell(r, qc).value
        if qty is None:
            qty = 1
        body_mat = ws.cell(r, bmc).value
        pipe_mat = ws.cell(r, pmc).value
        if body_mat is not None and str(body_mat).strip().lower() == SAME_AS_PIPE:
            mat2 = pipe_mat
        else:
            mat2 = body_mat
        mat1 = classify_mat1(mat2)
        size = ws.cell(r, sc).value
        end_type = ws.cell(r, ec).value if ec else None
        rating = ws.cell(r, rc).value if rc else None
        rows.append([None, None, line_no, tag, item, mat1, mat2, size, rating, end_type, qty])

# EDUCTOR - 별도 구조 (TAG NO 컬럼2, DISCHARGE 라인 정보 사용)
ws = wb['9.EDUCTOR']
for r in range(4, ws.max_row + 1):
    tag = ws.cell(r, 2).value
    if tag is None or str(tag).strip() == '':
        continue
    line_no = ws.cell(r, 15).value  # DISCHARGE LINE NO
    size = ws.cell(r, 17).value     # DISCHARGE SIZE
    qty = ws.cell(r, 38).value
    if qty is None:
        qty = 1
    mat2 = ws.cell(r, 39).value
    mat1 = classify_mat1(mat2)
    end_type = ws.cell(r, 40).value
    rating = ws.cell(r, 41).value
    rows.append([None, None, line_no, tag, 'EDUCTOR', mat1, mat2, size, rating, end_type, qty])

print(f'Total rows: {len(rows)}')

# 카테고리별 건수
from collections import Counter
cnt = Counter(r[4] for r in rows)
for k, v in cnt.items():
    print(f'  {k}: {v}')

# Tag 중복 체크
tags = [r[3] for r in rows]
dup = [t for t, c in Counter(tags).items() if c > 1]
print(f'중복 Tag: {len(dup)}건', dup[:10])

# mat1 미분류 값 체크
unclassified = sorted(set(r[6] for r in rows if r[5] is None and r[6] is not None))
print('MAT1 미분류 MAT2 값:', unclassified)

# 결과를 파일로 저장 (검토용 텍스트)
import json
with open('scratch/speciality_bom_preview.json', 'w', encoding='utf-8') as f:
    json.dump(rows, f, ensure_ascii=False, indent=1, default=str)
