# Safety Valve Tag 비교: BOM DB vs Received DB → 차이 엑셀 출력
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from collections import defaultdict
import os

SUPABASE_URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Accept-Profile': 'material'
}

def fetch_all(table, params='', limit=10000):
    rows = []
    offset = 0
    while True:
        url = f"{SUPABASE_URL}/rest/v1/{table}?{params}&limit={limit}&offset={offset}"
        r = requests.get(url, headers={**HEADERS, 'Range-Unit': 'items'})
        r.raise_for_status()
        data = r.json()
        if not data:
            break
        rows.extend(data)
        if len(data) < limit:
            break
        offset += limit
    return rows

print("BOM Safety Valve 태그 조회 중...")
bom_rows = fetch_all(
    'bom_detail',
    'select=tag,mat_code,full_description,system,iso_dwg_no,qty,uom'
    '&tag=not.is.null'
    '&or=(mat_code.ilike.PSV*,mat_code.ilike.PRV*,full_description.ilike.*Safety Valve*,full_description.ilike.*PSV*)'
)
print(f"  BOM PSV 행: {len(bom_rows)}건")

print("Received Safety Valve 태그 조회 중...")
recv_rows = fetch_all(
    'receiving',
    'select=tag,mat_code,full_description,category,doc_no,pkg_no,qty'
    '&category=eq.Valve'
    '&tag=not.is.null'
    '&or=(full_description.ilike.*Safety Valve*,full_description.ilike.*PSV*,tag.ilike.*PSV*,tag.ilike.*PRV*)'
)
print(f"  Received PSV 행: {len(recv_rows)}건")

# ── Tag 기준 집계 ─────────────────────────────────────────────────────────
bom_by_tag = defaultdict(list)
for r in bom_rows:
    tag = (r.get('tag') or '').strip().upper()
    if tag:
        bom_by_tag[tag].append(r)

recv_by_tag = defaultdict(list)
for r in recv_rows:
    tag = (r.get('tag') or '').strip().upper()
    if tag:
        recv_by_tag[tag].append(r)

bom_tags  = set(bom_by_tag.keys())
recv_tags = set(recv_by_tag.keys())

bom_only  = sorted(bom_tags - recv_tags)   # BOM에만 있음 (미입고)
recv_only = sorted(recv_tags - bom_tags)   # Received에만 있음 (BOM 없음)
matched   = sorted(bom_tags & recv_tags)   # 양쪽 모두

print(f"\n[비교 결과]")
print(f"  BOM에만 있음 (미입고): {len(bom_only)}건")
print(f"  Received에만 있음:     {len(recv_only)}건")
print(f"  양쪽 매칭:             {len(matched)}건")

# ── 엑셀 생성 ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# 색상 정의
RED_FILL    = PatternFill("solid", fgColor="FFCCCC")   # BOM만 (미입고)
GREEN_FILL  = PatternFill("solid", fgColor="CCFFCC")   # Received만
BLUE_FILL   = PatternFill("solid", fgColor="CCE5FF")   # 매칭
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BOLD        = Font(bold=True, size=10)
THIN        = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

def write_header(ws, cols):
    for c, title in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[1].height = 28

def style_row(ws, row_idx, fill, cols_count):
    for c in range(1, cols_count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill
        cell.border = THIN
        cell.alignment = Alignment(vertical='center')

# ── Sheet 1: 전체 요약 ────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Summary"
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 20

def s(r, c, v, bold=False, fill=None):
    cell = ws1.cell(row=r, column=c, value=v)
    if bold: cell.font = Font(bold=True)
    if fill: cell.fill = fill
    cell.border = THIN
    cell.alignment = Alignment(vertical='center')

s(1,1,"구분",bold=True); s(1,2,"건수",bold=True)
ws1.cell(1,1).fill = HEADER_FILL; ws1.cell(1,1).font = HEADER_FONT
ws1.cell(1,2).fill = HEADER_FILL; ws1.cell(1,2).font = HEADER_FONT

s(2,1,"BOM 전체 PSV Tags"); s(2,2,len(bom_tags))
s(3,1,"Received 전체 PSV Tags"); s(3,2,len(recv_tags))
s(4,1,"매칭 (양쪽 모두)"); s(4,2,len(matched), fill=BLUE_FILL)
s(5,1,"BOM에만 있음 (미입고)"); s(5,2,len(bom_only), fill=RED_FILL)
s(6,1,"Received에만 있음 (BOM 없음)"); s(6,2,len(recv_only), fill=GREEN_FILL)

# ── Sheet 2: BOM에만 있음 (미입고) ───────────────────────────────────────
ws2 = wb.create_sheet("BOM Only (미입고)")
cols2 = ["TAG", "MAT CODE", "DESCRIPTION", "SYSTEM", "ISO DWG NO", "QTY", "UOM"]
write_header(ws2, cols2)
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 8
ws2.column_dimensions['G'].width = 8

row = 2
for tag in bom_only:
    for b in bom_by_tag[tag]:
        ws2.cell(row,1,tag)
        ws2.cell(row,2,b.get('mat_code') or '')
        ws2.cell(row,3,b.get('full_description') or '')
        ws2.cell(row,4,b.get('system') or '')
        ws2.cell(row,5,b.get('iso_dwg_no') or '')
        ws2.cell(row,6,b.get('qty') or '')
        ws2.cell(row,7,b.get('uom') or 'EA')
        style_row(ws2, row, RED_FILL, len(cols2))
        row += 1

# ── Sheet 3: Received에만 있음 ────────────────────────────────────────────
ws3 = wb.create_sheet("Recv Only (BOM없음)")
cols3 = ["TAG", "MAT CODE", "DESCRIPTION", "DOC NO", "PKG NO", "QTY"]
write_header(ws3, cols3)
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 45
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 25
ws3.column_dimensions['F'].width = 8

row = 2
for tag in recv_only:
    for r in recv_by_tag[tag]:
        ws3.cell(row,1,tag)
        ws3.cell(row,2,r.get('mat_code') or '')
        ws3.cell(row,3,r.get('full_description') or '')
        ws3.cell(row,4,r.get('doc_no') or '')
        ws3.cell(row,5,r.get('pkg_no') or '')
        ws3.cell(row,6,r.get('qty') or '')
        style_row(ws3, row, GREEN_FILL, len(cols3))
        row += 1

# ── Sheet 4: 매칭 ─────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Matched")
cols4 = ["TAG", "BOM MAT CODE", "BOM DESCRIPTION", "BOM SYSTEM", "BOM ISO",
         "BOM QTY", "RECV MAT CODE", "RECV DESCRIPTION", "RECV PKG NO", "RECV QTY"]
write_header(ws4, cols4)
for i, w in enumerate([18,25,35,12,18,8,20,35,25,8], 1):
    ws4.column_dimensions[chr(64+i)].width = w

row = 2
for tag in matched:
    b_list = bom_by_tag[tag]
    r_list = recv_by_tag[tag]
    b = b_list[0]
    r = r_list[0]
    ws4.cell(row,1,tag)
    ws4.cell(row,2,b.get('mat_code') or '')
    ws4.cell(row,3,b.get('full_description') or '')
    ws4.cell(row,4,b.get('system') or '')
    ws4.cell(row,5,b.get('iso_dwg_no') or '')
    ws4.cell(row,6,b.get('qty') or '')
    ws4.cell(row,7,r.get('mat_code') or '')
    ws4.cell(row,8,r.get('full_description') or '')
    ws4.cell(row,9,r.get('pkg_no') or '')
    ws4.cell(row,10,r.get('qty') or '')
    style_row(ws4, row, BLUE_FILL, len(cols4))
    row += 1

# ── 저장 ──────────────────────────────────────────────────────────────────
out_path = os.path.join(os.path.dirname(__file__), '..', 'Raw File', 'SV_Tag_Comparison.xlsx')
os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)
print(f"\n저장 완료: {os.path.abspath(out_path)}")
