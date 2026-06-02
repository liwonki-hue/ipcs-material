# 신규 Packing List 4건 → material.receiving INSERT SQL 생성
# 0454: Casting Manual Valve (Valve BOM 매칭)
# 0504: Orifice Plate + 부속 (Speciality, description에서 tag 추출)
# 0363: Control Valve (Speciality)
# 0138: Bypass Steam System (Speciality)

import openpyxl, re, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

START_ID = 4065
OUT_PATH = 'scratch/insert_pl_0454_0504_0363_0138.sql'
VALVE_BOM_PATH = 'Raw File/BOM Data/Valve BOM.xlsx'

FILES = [
    ('PGU-DE-0454', 'Raw File/Packing List/PGU-DE-0454_BOP_Casting Manual Valve_CIPL(Rev.1).xlsx',                             'Valve'),
    ('PGU-DE-0504', 'Raw File/Packing List/PGU-DE-0504_BOP Piping Specialties Item_Orifice_CIPL(Rev.2).xlsx',                  'Speciality'),
    ('PGU-DE-0363', 'Raw File/Packing List/PGU-DE-0363_Control Valve_CIPL(Final)_IM-70 No.20.xlsx',                            'Speciality'),
    ('PGU-DE-0138', 'Raw File/Packing List/PGU-DE-0138_Bypass Steam System_CIPL(Rev.7).xlsx',                                  'Speciality'),
]

# Valve BOM 로드: Tag → {mat_code, desc}
bom_wb = openpyxl.load_workbook(VALVE_BOM_PATH)
bom_ws = bom_wb.active
tag_map = {}
for r in range(2, bom_ws.max_row + 1):
    mc   = bom_ws.cell(r, 1).value
    tag  = bom_ws.cell(r, 7).value
    desc = bom_ws.cell(r, 6).value
    if tag and mc:
        tag_map[str(tag).strip()] = {
            'mat_code': str(mc).strip(),
            'desc':     str(desc).strip() if desc else '',
        }
print(f'Valve BOM 로드: {len(tag_map)}건')

TAG_RE = re.compile(r'\(([A-Z][0-9]-[A-Z0-9]+-\d+[^)]*)\)\s*$')

def extract_tag_from_desc(desc):
    """description 말미 괄호에서 tag 추출 (0504 서브 아이템용)"""
    if not desc:
        return None
    m = TAG_RE.search(str(desc))
    return m.group(1).strip() if m else None

def esc(v):
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def normalize_unit(u):
    if not u:
        return 'EA'
    s = str(u).strip()
    if '/' in s:
        s = s.split('/')[0].strip()
    return s if s else 'EA'

SKIP_TAGS = {'-', 'spare', 'later', 'n/a'}

all_rows = []

for doc_no, fp, category in FILES:
    wb = openpyxl.load_workbook(fp)
    ws = wb['Detail PL'] if 'Detail PL' in wb.sheetnames else wb.active

    pkg_no     = None
    file_rows  = 0
    unmatched  = 0

    for r in range(2, ws.max_row + 1):
        pkg  = ws.cell(r, 1).value
        desc = ws.cell(r, 2).value
        qty  = ws.cell(r, 3).value
        unit = ws.cell(r, 4).value
        tag  = ws.cell(r, 5).value

        if not (pkg or desc or tag or qty):
            continue

        if pkg:
            pkg_no = str(pkg).strip()

        desc_str = str(desc).strip() if desc else ''
        tag_str  = str(tag).strip() if tag else ''

        if tag_str.lower() in SKIP_TAGS:
            continue
        if not desc_str:
            continue

        # 0504: tag 없으면 description 말미 괄호에서 추출
        if not tag_str and category == 'Speciality':
            tag_str = extract_tag_from_desc(desc_str) or ''

        rec_qty  = float(qty) if qty is not None else 1.0
        unit_str = normalize_unit(unit)

        if category == 'Valve' and tag_str:
            bom = tag_map.get(tag_str)
            if bom:
                mat_code  = bom['mat_code']
                full_desc = bom['desc']
            else:
                mat_code  = None
                full_desc = desc_str
                unmatched += 1
                print(f'  [미매칭] {doc_no} tag={tag_str}')
        else:
            mat_code  = None
            full_desc = desc_str

        all_rows.append({
            'doc_no':   doc_no,
            'pkg_no':   pkg_no,
            'mat_code': mat_code,
            'unit':     unit_str,
            'qty':      rec_qty,
            'category': category,
            'desc':     full_desc,
            'tag':      tag_str if tag_str else None,
        })
        file_rows += 1

    suffix = f' (미매칭 {unmatched}건)' if unmatched else ''
    print(f'{doc_no}: {file_rows}행{suffix}')

print(f'\n총 {len(all_rows)}행 → {OUT_PATH}')
print(f'id 범위: {START_ID} ~ {START_ID + len(all_rows) - 1}')

with open(OUT_PATH, 'w', encoding='utf-8') as f:
    f.write('-- 신규 Packing List receiving INSERT\n')
    f.write('-- 0454/0504/0363/0138\n')
    f.write(f'-- START_ID={START_ID}, 총 {len(all_rows)}행\n\n')

    for idx, row in enumerate(all_rows):
        row_id = START_ID + idx
        f.write(
            f"INSERT INTO material.receiving "
            f"(id, doc_no, pkg_no, mat_code, unit, qty, category, full_description, tag) VALUES "
            f"({row_id}, {esc(row['doc_no'])}, {esc(row['pkg_no'])}, {esc(row['mat_code'])}, "
            f"{esc(row['unit'])}, {row['qty']}, {esc(row['category'])}, "
            f"{esc(row['desc'])}, {esc(row['tag'])});\n"
        )

print('완료.')
