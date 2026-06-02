# Safety Valve BOM + Control Valve BOM → Supabase bom 테이블 INSERT SQL 생성
# 실행: python scratch/insert_sv_cv_bom.py
# 생성된 SQL → Supabase SQL Editor에서 실행

import openpyxl, sys, io, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── 헬퍼 ──────────────────────────────────────────────────────────────────
def esc(val):
    if val is None:
        return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

def clean(val):
    if val is None:
        return None
    return re.sub(r'\s+', ' ', str(val)).strip() or None

INCH_TO_DN = {
    '1/2"': 'DN 15', '3/4"': 'DN 20', '1"': 'DN 25', '1.5"': 'DN 40',
    '1-1/2"': 'DN 40', '2"': 'DN 50', '2.5"': 'DN 65', '3"': 'DN 80',
    '4"': 'DN 100', '5"': 'DN 125', '6"': 'DN 150', '8"': 'DN 200',
    '10"': 'DN 250', '12"': 'DN 300', '14"': 'DN 350', '16"': 'DN 400',
}

def to_dn(val):
    if val is None:
        return None
    s = str(val).strip()
    return INCH_TO_DN.get(s, s)

def tag_system(tag):
    """태그 접두사 B0/B1/B2 → 시스템 코드"""
    if not tag:
        return None
    m = re.match(r'^(B[012])-', tag)
    return m.group(1) if m else None

# ── 1. Safety Valve BOM ───────────────────────────────────────────────────
sv_rows = []
SV_PATH = 'Raw File/BOM Data/Safety Valve BOM.xlsx'
wb = openpyxl.load_workbook(SV_PATH)
ws = wb.active
for r in range(2, ws.max_row + 1):
    sys_  = ws.cell(r, 1).value   # System
    # col 2 = ISO Drawing, col 3 = LINE No  → 모두 NULL
    desc  = ws.cell(r, 4).value   # Description
    tag   = ws.cell(r, 5).value   # Tag No.
    size  = ws.cell(r, 7).value   # Size
    mat   = ws.cell(r, 8).value   # Body Mat'l
    conn  = ws.cell(r, 9).value   # Connection
    qty   = ws.cell(r, 11).value  # Q'ty

    tag = clean(tag)
    if not tag:
        continue

    full_desc = ', '.join(filter(None, [
        clean(desc), to_dn(size), clean(mat), clean(conn)
    ]))
    sv_rows.append({
        'mat_code': None,
        'category': 'Valve',
        'tag':      tag,
        'system':   clean(sys_),
        'iso_dwg_no': None,
        'line_no':  None,
        'full_description': full_desc or None,
        'uom':      'EA',
        'qty':      float(qty) if qty is not None else 1.0,
    })

print(f'Safety Valve: {len(sv_rows)}행')

# ── 2. Control Valve BOM (C01A 시트) ──────────────────────────────────────
cv_rows = []
CV_PATH = 'Raw File/BOM Data/Control Valve BOM.xlsx'
wb2 = openpyxl.load_workbook(CV_PATH)
ws2 = wb2['C01A']
for r in range(2, ws2.max_row + 1):
    # col 1=System(NULL), 2=ISO(NULL), 3=LineNo(NULL), 4=Description
    desc  = ws2.cell(r, 4).value
    tag   = ws2.cell(r, 5).value   # Tag
    style = ws2.cell(r, 6).value   # Body Style
    size  = ws2.cell(r, 7).value   # Size
    mat   = ws2.cell(r, 8).value   # Body Material
    rating= ws2.cell(r, 9).value   # Rating (numeric)
    conn  = ws2.cell(r, 10).value  # Conn. Type
    qty   = ws2.cell(r, 11).value  # Qty.

    tag = clean(tag)
    if not tag:
        continue

    rating_str = f'{int(rating)}#' if rating is not None else None
    full_desc = ', '.join(filter(None, [
        clean(desc), clean(style), to_dn(size), clean(mat), rating_str, clean(conn)
    ]))
    cv_rows.append({
        'mat_code': None,
        'category': 'Valve',
        'tag':      tag,
        'system':   tag_system(tag),   # B0/B1/B2 파생 (소스에 system 없음)
        'iso_dwg_no': None,
        'line_no':  None,
        'full_description': full_desc or None,
        'uom':      'EA',
        'qty':      float(qty) if qty is not None else 1.0,
    })

print(f'Control Valve: {len(cv_rows)}행')

# ── 3. SQL 생성 ────────────────────────────────────────────────────────────
all_rows = sv_rows + cv_rows

# 삭제 조건: PSV/TCV/LCV/FCV/FV/PCV 패턴 태그
DELETE_SQL = (
    "-- Safety Valve / Control Valve BOM 기존 데이터 삭제\n"
    "DELETE FROM material.bom\n"
    "WHERE category = 'Valve'\n"
    "  AND tag ~ '^B[012w][-]?(PSV|TCV|LCV|FCV|FV|PCV)';\n\n"
)

def make_insert(row):
    return (
        "INSERT INTO material.bom "
        "(mat_code, category, tag, system, iso_dwg_no, line_no, full_description, uom, qty) VALUES "
        f"(NULL, 'Valve', {esc(row['tag'])}, {esc(row['system'])}, "
        f"NULL, NULL, {esc(row['full_description'])}, 'EA', {row['qty']});"
    )

out_path = 'scratch/insert_sv_cv_bom.sql'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('-- Safety Valve + Control Valve BOM INSERT\n')
    f.write('-- Supabase SQL Editor에서 실행\n\n')
    f.write(DELETE_SQL)
    f.write(f'-- Safety Valve ({len(sv_rows)}행)\n')
    for row in sv_rows:
        f.write(make_insert(row) + '\n')
    f.write(f'\n-- Control Valve ({len(cv_rows)}행)\n')
    for row in cv_rows:
        f.write(make_insert(row) + '\n')

print(f'\n→ {out_path} 생성 완료 (총 {len(all_rows)}행)')
print('\n[Safety Valve 샘플 3건]')
for row in sv_rows[:3]:
    print(f"  tag={row['tag']}, sys={row['system']}, desc={row['full_description']}")
print('\n[Control Valve 샘플 3건]')
for row in cv_rows[:3]:
    print(f"  tag={row['tag']}, sys={row['system']}, desc={row['full_description']}")
