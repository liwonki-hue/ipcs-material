# MOV Excel Tag 분해: 멀티 Tag → 1행 1Tag (qty=1)

import re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 공통 스타일 ────────────────────────────────────────────────
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=9)
GRP_FILL = PatternFill('solid', fgColor='D6E4F0')
GRP_FONT = Font(bold=True, size=9)
CELL_FONT = Font(size=9)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, n):
    for c in range(1, n+1):
        ws.cell(row, c).fill = HDR_FILL
        ws.cell(row, c).font = HDR_FONT
        ws.cell(row, c).border = BORDER
        ws.cell(row, c).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

def style_group(ws, row, n, text):
    ws.cell(row, 1, text).fill = GRP_FILL
    ws.cell(row, 1).font = GRP_FONT
    ws.cell(row, 1).border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)

def style_cell(ws, row, n):
    for c in range(1, n+1):
        ws.cell(row, c).font = CELL_FONT
        ws.cell(row, c).border = BORDER
        ws.cell(row, c).alignment = Alignment(wrap_text=False, vertical='center')

# ── Tag 분해 함수 ──────────────────────────────────────────────
def expand_tag_group(s):
    """'B1-MOV-32001/002/003' → ['B1-MOV-32001','B1-MOV-32002','B1-MOV-32003']"""
    s = s.strip()
    if not s or s.upper() in ('LATER', '-'):
        return [s]

    # 범위 ~ 처리: 'B1-MOV-34001~34003'
    if '~' in s:
        left, right = s.split('~', 1)
        m = re.match(r'^(.*?)(\d+)([A-Za-z]*)$', left.strip())
        if m:
            prefix, start_num, alpha = m.group(1), m.group(2), m.group(3)
            end_num = int(re.sub(r'[^\d]', '', right))
            width = len(start_num)
            return [f"{prefix}{n:0{width}d}{alpha}" for n in range(int(start_num), end_num+1)]

    # 슬래시 / 처리: 'B1-MOV-32001/002/003'
    parts = s.split('/')
    if len(parts) == 1:
        return [s]

    first = parts[0].strip()
    result = [first]
    m = re.match(r'^(.*?)(\d+)([A-Za-z]*)$', first)
    if not m:
        return [s]
    prefix, full_num, alpha = m.group(1), m.group(2), m.group(3)
    full_with_alpha = full_num + alpha

    for sfx in parts[1:]:
        sfx = sfx.strip()
        if len(sfx) >= len(full_with_alpha):
            result.append(prefix + sfx)
        else:
            new_tail = full_with_alpha[:-len(sfx)] + sfx
            result.append(prefix + new_tail)
    return result

def expand_all_tags(tag_str):
    """멀티라인 + 슬래시 tag 문자열 → 개별 tag 리스트"""
    all_tags = []
    for line in str(tag_str).split('\n'):
        line = line.strip()
        if line:
            all_tags.extend(expand_tag_group(line))
    return all_tags

# ══════════════════════════════════════════════════════════════
# 1. PDF 13 → MOV_List_Rev01_tags.xlsx
# ══════════════════════════════════════════════════════════════
COLS1 = ['System','P&ID No.','Description','GA Dwg\nSheet No.',
         'Tag No.','Q\'ty','Size\n[DN]','Valve Type','Valve Detail Type',
         'ANSI\nClass','Medium','Design\nPress.\n[bar.a]',
         'Design\nTemp.\n[°C]','Body Material\n(ASTM)']
N1 = len(COLS1)

# (System, P&ID, Description, GA, Tags, DN, Type, Detail, Class, Medium, DP, DT, Mat)
RAW1 = [
    # ── HP & LP Steam System
    ('HP & LP Steam System', None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B126-PR-007-0001-001','HRSG #1 HP Steam Drain MOV (upstream NRV)',5,'B1-MOV-26001\nB2-MOV-26001',25,'GLOBE','T-type, Plug Disc','1500#','Steam',96.4,530,'SA182-F91'),
    ('','CCP-PW-B126-PR-007-0001-001','HRSG #2 HP Steam Drain MOV (upstream NRV)',5,'B1-MOV-26002\nB2-MOV-26002',25,'GLOBE','T-type, Plug Disc','1500#','Steam',96.4,530,'SA182-F91'),
    ('','CCP-PW-B126-PR-007-0001-001','HRSG #1 HP Steam Drain MOV (downstream NRV)',6,'B1-MOV-26003\nB2-MOV-26003',25,'GLOBE','T-type, Plug Disc','1500#','Steam',96.4,530,'A182-F91'),
    ('','CCP-PW-B126-PR-007-0001-001','HRSG #2 HP Steam Drain MOV (downstream NRV)',6,'B1-MOV-26004\nB2-MOV-26004',25,'GLOBE','T-type, Plug Disc','1500#','Steam',96.4,530,'A182-F91'),
    ('','CCP-PW-B126-PR-007-0001-001','Auxiliary Steam Supply MOV',7,'B1-MOV-26011\nB2-MOV-26011',100,'GATE','Parallel Slide Disc','1500#','Steam',96.4,530,'A182-F91 or A217-C12A'),
    ('','CCP-PW-B126-PR-007-0001-001','HP Steam Dynamic Strainer Drain MOV',8,'B1-MOV-26012\nB2-MOV-26012',50,'GLOBE','T-type, Plug Disc','1500#','Steam',96.4,530,'A182-F91'),
    ('','CCP-PW-B126-PR-007-0001-001','HP Steam Start-up Warming Drain MOV',9,'B1-MOV-26021\nB2-MOV-26021',80,'GLOBE','T-type, Plug Disc','1500#','Steam',96.4,530,'A182-F91 or A217-C12A'),
    ('','CCP-PW-B127-PR-007-0001-002','HRSG #1 LP Steam Drain MOV (upstream NRV)',10,'B1-MOV-27001\nB2-MOV-27001',25,'GLOBE','T-type, Plug Disc','600#','Steam',10,292,'SA105'),
    ('','CCP-PW-B127-PR-007-0001-002','HRSG #2 LP Steam Drain MOV (upstream NRV)',10,'B1-MOV-27002\nB2-MOV-27002',25,'GLOBE','T-type, Plug Disc','600#','Steam',10,292,'SA105'),
    ('','CCP-PW-B127-PR-007-0001-002','HRSG #1 LP Steam Drain MOV (downstream NRV)',10,'B1-MOV-27003\nB2-MOV-27003',25,'GLOBE','T-type, Plug Disc','600#','Steam',10,292,'SA105'),
    ('','CCP-PW-B127-PR-007-0001-002','HRSG #2 LP Steam Drain MOV (downstream NRV)',10,'B1-MOV-27004\nB2-MOV-27004',25,'GLOBE','T-type, Plug Disc','600#','Steam',10,292,'SA105'),
    ('','CCP-PW-B127-PR-007-0001-002','LP Steam Common Drain MOV',29,'B1-MOV-27011\nB2-MOV-27011',25,'GLOBE','T-type, Plug Disc','600#','Steam',10,292,'A105'),
    # ── Steam Blowing System
    ('Steam Blowing System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','N/A','HP STEAM BLOWING MOV (Martyr valve)',11,'Later',400,'GATE','Flexible Wedge Disc','900#','Steam',51,400,'A216-WCC'),
    ('','N/A','LP STEAM BLOWING MOV (Martyr valve)',12,'Later',550,'GATE','Flexible Wedge Disc','600#','Steam',10,292,'A216-WCC'),
    # ── Auxiliary Steam System
    ('Auxiliary Steam System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B128-PR-007-0001-001','Start-up Auxiliary Boiler Supply MOV',13,'B0-MOV-28021',150,'GATE','Flexible Wedge Disc','300#','Steam',21,326,'A216-WCB'),
    ('','CCP-PW-B128-PR-007-0001-001','Neighboring Block Auxiliary Steam Header Supply MOV',13,'B1-MOV-28011\nB2-MOV-28011',150,'GATE','Flexible Wedge Disc','300#','Steam',21,326,'A216-WCB'),
    ('','CCP-PW-B128-PR-007-0001-001','Turbine Gland Seal Steam Header Supply MOV',14,'B1-MOV-28012\nB2-MOV-28012',50,'GLOBE','T-type, Plug Disc','600#','Steam',21,326,'A105'),
    ('','CCP-PW-B128-PR-007-0001-001','Turbine Gland Seal Steam Header Warming Drain MOV',15,'B1-MOV-28013\nB2-MOV-28013',50,'GLOBE','T-type, Plug Disc','600#','Steam',21,326,'A105'),
    ('','CCP-PW-B128-PR-007-0001-001','Aux. PRDS Pressure Control Valve Drain MOV',16,'B1-MOV-28001\nB2-MOV-28001',25,'GLOBE','T-type, Plug Disc','600#','Steam',21,326,'A105'),
    # ── Condensate System
    ('Condensate System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B129-PR-007-0001-001','CEP DISCHARGE MOV',17,'B1-MOV-29001~29003\nB2-MOV-29001~29003',200,'GATE','Flexible Wedge Disc','300#','Condensate',36,100,'A216-WCB'),
    # ── Feedwater System
    ('Feedwater System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B130-PR-007-0001-001','HRSG #11 BFP A IP Discharge MOV',18,'B1-MOV-30001A/31001A\nB2-MOV-30001A/31001A',80,'GATE','Parallel Slide Disc','600#','Feedwater',60,195,'A216-WCB'),
    ('','CCP-PW-B130-PR-007-0001-001','HRSG #11 BFP A HP Discharge MOV',19,'B1-MOV-30002A/31002A\nB2-MOV-30002A/31002A',150,'GATE','Parallel Slide Disc','1500#','Feedwater',143,198,'A216-WCC'),
    ('','CCP-PW-B130-PR-007-0001-001','BYPASS MOV FOR HRSG #11 BFP A HP Discharge MOV',20,'B1-MOV-30003A/31003A\nB2-MOV-30003A/31003A',25,'GLOBE','T-type, Plug Disc','1500#','Feedwater',143,198,'A105'),
    ('','CCP-PW-B130-PR-007-0001-001','HRSG #11 BFP B IP Discharge MOV',18,'B1-MOV-30001B/31001B\nB2-MOV-30001B/31001B',80,'GATE','Parallel Slide Disc','600#','Feedwater',60,195,'A216-WCB'),
    ('','CCP-PW-B130-PR-007-0001-001','HRSG #11 BFP B HP Discharge MOV',19,'B1-MOV-30002B/31002B\nB2-MOV-30002B/31002B',150,'GATE','Parallel Slide Disc','1500#','Feedwater',143,198,'A216-WCC'),
    ('','CCP-PW-B130-PR-007-0001-001','BYPASS MOV FOR HRSG #11 BFP B HP Discharge MOV',20,'B1-MOV-30003B/31003B\nB2-MOV-30003B/31003B',25,'GLOBE','T-type, Plug Disc','1500#','Feedwater',143,198,'A105'),
    # ── Closed Cooling Water System
    ('Closed Cooling Water System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B132-PR-007-0001-001','Closed Cooling Water Head Tank Make-up Line',21,'B1-MOV-32005\nB2-MOV-32005',25,'GLOBE','T-type, Plug Disc','600#','DM water',13,55,'A182-F304'),
    # ── Hot Water Distribution System
    ('Hot Water Distribution System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B134-PR-007-0001-001','HOT WATER HEAD TANK MAKE UP LINE MOV',22,'B1-MOV-34004\nB2-MOV-34004',25,'GLOBE','T-type, Plug Disc','600#','DM water',13,55,'A182-F304'),
    ('','CCP-PW-B134-PR-007-0001-001','HOT WATER SUPPLY PUMP DISCHARGE MOV',23,'B1-MOV-34001~34003\nB2-MOV-34001~34003',150,'GLOBE','T-type, Plug Disc','150#','Hot Water',16,120,'A216-WCB'),
    ('','CCP-PW-B134-PR-007-0001-002','HOT WATER HEATER TUBE SIDE INLET MOV',24,'B1-MOV-34005/34008\nB2-MOV-34005/34008',150,'GATE','Flexible Wedge Disc','150#','Hot Water',16,120,'A216-WCB'),
    ('','CCP-PW-B134-PR-007-0001-002','HOT WATER HEATER TUBE SIDE OUTLET MOV',24,'B1-MOV-34011/34013\nB2-MOV-34011/34013',150,'GATE','Flexible Wedge Disc','150#','Hot Water',16,120,'A216-WCB'),
    ('','CCP-PW-B134-PR-007-0001-002','HOT WATER HEATER SHELL SIDE INLET MOV',25,'B1-MOV-34006/34009\nB2-MOV-34006/34009',250,'GATE','Flexible Wedge Disc','300#','Steam',5,326,'A216-WCB'),
    ('','CCP-PW-B134-PR-007-0001-002','HOT WATER HEATER SHELL SIDE VENT LINE BYPASS MOV',26,'B1-MOV-34007/34010\nB2-MOV-34007/34010',25,'GLOBE','T-type, Plug Disc','600#','Steam',5,326,'A105'),
    # ── Demineralized Water System
    ('Demineralized Water System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B035-PR-007-0001-001','Demi plant to Demi tank A inlet MOV',27,'B0-MOV-35001',200,'GATE','Flexible Wedge Disc','150#','DM water',5,55,'A351-CF8'),
    ('','CCP-PW-B035-PR-007-0001-001','Demi plant to Demi tank B inlet MOV',27,'B0-MOV-35002',200,'GATE','Flexible Wedge Disc','150#','DM water',5,55,'A351-CF8'),
    # ── Fuel Oil System
    ('Fuel Oil System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B046-PR-007-0001-002','Fuel Oil Transfer Pump-A Discharge MOV',28,'B0-MOV-46001',400,'GATE','Flexible Wedge Disc','150#','Diesel Oil',16,55,'A216-WCB'),
    ('','CCP-PW-B046-PR-007-0001-002','Fuel Oil Transfer Pump-B Discharge MOV',28,'B0-MOV-46002',400,'GATE','Flexible Wedge Disc','150#','Diesel Oil',16,55,'A216-WCB'),
    # ── Service Water System
    ('Service Water System',None,None,None,None,None,None,None,None,None,None,None,None),
    ('','CCP-PW-B133-PR-007-0001-001','Drain Quenching water MOV',30,'B1-MOV-33001\nB2-MOV-33001',50,'GLOBE','T-type, Plug Disc','600#','Service Water',10,50,'A105'),
]

wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = 'MOV List'
ws1.row_dimensions[1].height = 40
for c, h in enumerate(COLS1, 1):
    ws1.cell(1, c, h)
style_header(ws1, 1, N1)

r = 2
for row in RAW1:
    sys_name = row[0]
    if sys_name and row[1] is None:
        style_group(ws1, r, N1, sys_name)
        r += 1
        continue
    tags = expand_all_tags(row[4]) if row[4] else ['']
    for tag in tags:
        # System P&ID Desc GA Tag DN Type Detail Class Medium DP DT Mat
        vals = [row[0], row[1], row[2], row[3], tag, 1,
                row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12]]
        for c, v in enumerate(vals, 1):
            ws1.cell(r, c, v)
        style_cell(ws1, r, N1)
        r += 1

widths1 = [22, 30, 50, 8, 22, 5, 7, 8, 18, 7, 12, 8, 7, 20]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'
ws1.row_dimensions[1].height = 40

total_rows1 = r - 2
wb1.save('Raw File/MOV_List_Rev01_tags.xlsx')
print(f'MOV_List_Rev01_tags.xlsx 저장 ({total_rows1}행)')

# ══════════════════════════════════════════════════════════════
# 2. PDF 12 → MOV_Butterfly_Valve_Rev_C01_tags.xlsx
# ══════════════════════════════════════════════════════════════
COLS2 = ['NO.','REV.','Tag No.','P&ID No.','Valve Description',
         'Valve\nModel','Rating','Size\n(inch)','Size\n(DN)','Q\'ty',
         'Type','Drilling','End','Body','Disc','Seat','Stem',
         'Design\nPress.\n[bar]','Work\nPress.\n[bar]',
         'Design\nTemp.\n[°C]','Work\nTemp.\n[°C]','Operator']
N2 = len(COLS2)

# (NO, REV, Tags, P&ID, Desc, Model, Rating, Inch, DN, Type, Drilling, End, Body, Disc, Seat, Stem, DP, WP, DT, WT, Op)
RAW2 = [
    (1,'C01','B1-MOV-32001/002/003\nB2-MOV-32001/002/003','CCP-PW-B132-PR-007-0001-001','Each CCWP Outlet Line','SRS912','150LB',14,350,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A351 CF8','EPDM','A479 T410',10,7,80,55,'MOTOR'),
    (2,'C01','B1-MOV-32004\nB2-MOV-32004','CCP-PW-B132-PR-007-0001-001','Each CCW FFC Bypass Line','SRS912','150LB',20,500,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A351 CF8','EPDM','A479 T410',10,7,80,48,'MOTOR'),
    (3,'C01','B1-MOV-34015/014\nB2-MOV-34015/017','CCP-PW-B134-PR-007-0001-003','Each GT Air Preheater Supply MOV','SRS912','150LB',6,150,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A216 WCB','EPDM','A479 T410',16,10,120,90,'MOTOR'),
    (4,'C01','B0-MOV-38001/002','CCP-PW-B038-PR-007-0001-001','External raw water supply to raw water tank inlet MOV','SRS912','150LB',10,250,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A351 CF8','A351 CF8','EPDM','A479 T410',10,3,50,44.2,'MOTOR'),
    (5,'C01','B0-MOV-38003/004/005/006','CCP-PW-B038-PR-007-0001-001','Raw Water Discharge to Raw Water Distribution Line Common','SRS912','150LB',6,150,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A351 CF8','EPDM','A479 T410',10,4.1,55,44.2,'MOTOR'),
    (6,'C01','B0-MOV-36001/002','CCP-PW-B036-PR-007-0001-001','External Service Water Supply to Service Water Tank Inlet MOV','SRS913','150LB',3,80,'FLANGE','ASME B16.5 150LB RF','MSS SP67 NARROW','A351 CF8','A351 CF8','EPDM','A479 T410',10,3,50,44.2,'MOTOR'),
    (7,'C01','B0-MOV-36003/004/005/006','CCP-PW-B036-PR-007-0001-001','Service Water Discharge to Service Water Distribution Line Common','SRS913','150LB',4,100,'FLANGE','ASME B16.5 150LB RF','MSS SP67 NARROW','A216 WCB','A351 CF8','EPDM','A479 T410',10,4.7,55,44.2,'MOTOR'),
    (8,'C01','LATER','LATER','N/A','SRS912','150LB',8,200,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A351 CF8','EPDM','A479 T410',10,3,55,35,'MOTOR'),
    (9,'C01','B0-MOV-37003/004/005/006','CCP-PW-B037-PR-007-0001-001','Potable Water Discharge to Potable Water Distribution Line Common','SRS912','150LB',6,150,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A351 CF8','A351 CF8','EPDM','A479 T410',10,6.5,55,35,'MOTOR'),
]

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Butterfly Valve'
ws2.row_dimensions[1].height = 45
for c, h in enumerate(COLS2, 1):
    ws2.cell(1, c, h)
style_header(ws2, 1, N2)

r = 2
seq = 1
for row in RAW2:
    no_, rev = row[0], row[1]
    tags = expand_all_tags(row[2])
    for tag in tags:
        vals = [seq, rev, tag, row[3], row[4], row[5], row[6], row[7], row[8], 1,
                row[9], row[10], row[11], row[12], row[13], row[14], row[15],
                row[16], row[17], row[18], row[19], row[20]]
        for c, v in enumerate(vals, 1):
            ws2.cell(r, c, v)
        style_cell(ws2, r, N2)
        r += 1
        seq += 1

# 합계
ws2.cell(r, 1, 'TOTAL').font = Font(bold=True, size=9)
ws2.cell(r, 1).border = BORDER
ws2.cell(r, 10, seq-1).font = Font(bold=True, size=9)
ws2.cell(r, 10).border = BORDER
for c in range(2, N2+1):
    ws2.cell(r, c).border = BORDER

widths2 = [5,5,22,30,52,8,7,6,6,5,8,22,18,10,10,10,10,7,7,7,7,8]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

total_rows2 = seq - 1
wb2.save('Raw File/MOV_Butterfly_Valve_Rev_C01_tags.xlsx')
print(f'MOV_Butterfly_Valve_Rev_C01_tags.xlsx 저장 ({total_rows2}행)')
