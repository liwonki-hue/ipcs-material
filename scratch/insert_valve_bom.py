# Valve BOM Excel → Supabase bom 테이블 INSERT SQL 생성
# 실행: python scratch/insert_valve_bom.py
# 생성된 SQL → Supabase SQL Editor에서 실행

import openpyxl, sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── 시스템 코드 정규화 (update_bom_system_codes.sql 기준) ──────────────────
SYS_MAP = {
    'HWR': 'HW', 'HWS': 'HW',
    'IG': 'GT MISC',
    'BWF': 'FW',
    'CWR': 'CCW', 'CWS': 'CCW',
    'UW': 'SW',
    'LN': 'N2',
    'LS': 'LP',
    'HS': 'HP',
}

def norm_sys(val):
    if not val:
        return None
    v = str(val).strip()
    return SYS_MAP.get(v, v)

def esc(val):
    if val is None:
        return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

# ── 엑셀 읽기 ─────────────────────────────────────────────────────────────
WB_PATH = 'Raw File/BOM Data/Valve BOM.xlsx'
wb = openpyxl.load_workbook(WB_PATH)
ws = wb.active

rows = []
skipped = 0
for r in range(2, ws.max_row + 1):
    mc   = ws.cell(r, 1).value
    cat  = ws.cell(r, 2).value
    sys_ = ws.cell(r, 3).value
    iso  = ws.cell(r, 4).value
    line = ws.cell(r, 5).value
    desc = ws.cell(r, 6).value
    tag  = ws.cell(r, 7).value
    uom  = ws.cell(r, 13).value
    qty  = ws.cell(r, 14).value

    if not mc:
        skipped += 1
        continue

    rows.append({
        'mat_code':        str(mc).strip(),
        'category':        'Valve',
        'tag':             str(tag).strip() if tag else None,
        'system':          norm_sys(sys_),
        'iso_dwg_no':      str(iso).strip() if iso else None,
        'line_no':         str(line).strip() if line else None,
        'full_description': str(desc).strip() if desc else None,
        'uom':             str(uom).strip() if uom else 'EA',
        'qty':             float(qty) if qty is not None else 1.0,
    })

print(f'처리 대상: {len(rows)}행, 스킵: {skipped}행')

# ── SQL 생성 ──────────────────────────────────────────────────────────────
BATCH = 500
out_dir = 'scratch'
file_idx = 0
lines = []

HEADER = (
    '-- Valve BOM INSERT\n'
    '-- Supabase SQL Editor에서 실행\n'
    '-- 기존 Valve 데이터 삭제 후 재삽입\n\n'
    "DELETE FROM material.bom WHERE category = 'Valve';\n\n"
)

for i, row in enumerate(rows):
    sql = (
        f"INSERT INTO material.bom "
        f"(mat_code, category, tag, system, iso_dwg_no, line_no, full_description, uom, qty) VALUES "
        f"({esc(row['mat_code'])}, 'Valve', {esc(row['tag'])}, {esc(row['system'])}, "
        f"{esc(row['iso_dwg_no'])}, {esc(row['line_no'])}, {esc(row['full_description'])}, "
        f"{esc(row['uom'])}, {row['qty']});"
    )
    lines.append(sql)

    if len(lines) >= BATCH or i == len(rows) - 1:
        fname = os.path.join(out_dir, f'insert_valve_bom_{file_idx:02d}.sql')
        with open(fname, 'w', encoding='utf-8') as f:
            if file_idx == 0:
                f.write(HEADER)
            else:
                f.write('-- Valve BOM INSERT (continued)\n\n')
            f.write('\n'.join(lines) + '\n')
        print(f'  → {fname} ({len(lines)}행)')
        lines = []
        file_idx += 1

print(f'\n완료: SQL 파일 {file_idx}개 생성')
print('실행 순서: insert_valve_bom_00.sql → insert_valve_bom_01.sql → ...')
print('\n[주의] insert_valve_bom_00.sql 첫 실행 시 기존 Valve 데이터 DELETE 후 INSERT')
