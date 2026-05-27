# PGU-DE-0390 & 0391 Packing List → material.receiving INSERT SQL 생성
# - Valve: Tag 기준 BOM 매칭 → mat_code (미매칭 NULL)
# - Accessory (SEAL RING / HINGE PIN / TAG PLATE): mat_code=NULL, category='Accessory'
# - SPARE: 스킵

import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

START_ID = 1024
OUT_PATH = 'scratch/insert_0390_0391_receiving.sql'

ACCESSORY_KEYWORDS = ('TAG PLATE', 'SEAL RING', 'HINGE PIN')

# ── BOM Tag → mat_code / description 매핑 ───────────────────────────────────
bom_wb = openpyxl.load_workbook('Raw File/BOM Data/Valve BOM.xlsx')
bom_ws = bom_wb.active
tag_map = {}
for r in range(2, bom_ws.max_row + 1):
    mc   = bom_ws.cell(r, 1).value
    tag  = bom_ws.cell(r, 7).value
    desc = bom_ws.cell(r, 6).value
    if tag and mc:
        t = str(tag).strip()
        tag_map[t] = {
            'mat_code': str(mc).strip(),
            'desc':     str(desc).strip() if desc else '',
        }

def esc(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def is_accessory(desc_str):
    d = desc_str.upper()
    return any(kw in d for kw in ACCESSORY_KEYWORDS)

def process_pl(fpath, doc_no):
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    rows = []
    pkg_no = None
    stats = {'valve_matched': 0, 'valve_unmatched': 0, 'accessory': 0, 'skip_spare': 0}

    for r in range(2, ws.max_row + 1):
        pkg  = ws.cell(r, 2).value
        desc = ws.cell(r, 3).value
        qty  = ws.cell(r, 4).value
        tag  = ws.cell(r, 6).value

        if pkg:
            pkg_no = str(pkg).strip()

        desc_str = str(desc).strip() if desc else ''
        tag_str  = str(tag).strip() if tag else ''

        # 빈 행 스킵 (desc도 tag도 없음)
        if not desc_str and not tag_str:
            continue

        # SPARE 스킵
        if tag_str.upper() == 'SPARE':
            stats['skip_spare'] += 1
            continue

        rec_qty = float(qty) if qty is not None else 1.0

        if is_accessory(desc_str):
            # Accessory: BOM 매칭 불필요, PL description 그대로
            stats['accessory'] += 1
            rows.append({
                'doc_no':   doc_no,
                'pkg_no':   pkg_no,
                'mat_code': None,
                'unit':     'EA',
                'qty':      rec_qty,
                'category': 'Accessory',
                'desc':     desc_str,
                'tag':      tag_str if tag_str else None,
            })
        else:
            # Valve: Tag 기준 BOM 매칭
            bom = tag_map.get(tag_str) if tag_str else None
            if bom:
                mat_code  = bom['mat_code']
                full_desc = bom['desc']
                stats['valve_matched'] += 1
            else:
                mat_code  = None
                full_desc = desc_str
                stats['valve_unmatched'] += 1
                if tag_str:
                    print(f'  [미매칭] {tag_str} | {desc_str[:50]}')

            rows.append({
                'doc_no':   doc_no,
                'pkg_no':   pkg_no,
                'mat_code': mat_code,
                'unit':     'EA',
                'qty':      rec_qty,
                'category': 'Valve',
                'desc':     full_desc,
                'tag':      tag_str if tag_str else None,
            })

    return rows, stats

# ── 두 PL 처리 ───────────────────────────────────────────────────────────────
all_rows = []
for fpath, doc in [
    ('Raw File/Packing List (편집)/PGU-DE-0390_BOP Piping-Manual Valve(확인).xlsx',  'PGU-DE-0390'),
    ('Raw File/Packing List (편집)/PGU-DE-0391_BOP Piping-Manual Valves(확인).xlsx', 'PGU-DE-0391'),
]:
    print(f'=== {doc} ===')
    rows, st = process_pl(fpath, doc)
    all_rows.extend(rows)
    print(f'  Valve 매칭: {st["valve_matched"]}, 미매칭: {st["valve_unmatched"]}')
    print(f'  Accessory: {st["accessory"]}, SPARE 스킵: {st["skip_spare"]}')
    print(f'  소계: {len(rows)}행')

print(f'\n총 INSERT: {len(all_rows)}행, id {START_ID} ~ {START_ID + len(all_rows) - 1}')

# ── SQL 생성 ─────────────────────────────────────────────────────────────────
with open(OUT_PATH, 'w', encoding='utf-8') as f:
    f.write('-- PGU-DE-0390 & 0391 Valve Receiving INSERT\n')
    f.write('-- Valve: Tag 기준 BOM mat_code 매칭 / Accessory: mat_code=NULL, category=Accessory\n')
    f.write(f'-- id {START_ID} ~ {START_ID + len(all_rows) - 1}, 총 {len(all_rows)}행\n\n')
    for idx, row in enumerate(all_rows):
        row_id = START_ID + idx
        f.write(
            f"INSERT INTO material.receiving "
            f"(id, doc_no, pkg_no, mat_code, unit, qty, category, full_description, tag) VALUES "
            f"({row_id}, {esc(row['doc_no'])}, {esc(row['pkg_no'])}, {esc(row['mat_code'])}, "
            f"{esc(row['unit'])}, {row['qty']}, {esc(row['category'])}, "
            f"{esc(row['desc'])}, {esc(row['tag'])});\n"
        )

print(f'\n→ SQL 저장: {OUT_PATH}')
