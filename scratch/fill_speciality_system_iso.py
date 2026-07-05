# Speciality BOM.xlsx의 Line No를 ipcs-drawing(Supabase drawing.dwg_latest)와 매칭해서
# SYSTM / ISO DRAWING 컬럼을 채운다 (읽기 전용 조회, ipcs-drawing 쪽은 수정하지 않음)
import json
import re
from collections import defaultdict

import openpyxl

DWG_JSON = 'scratch/dwg_latest_raw.json'
TARGET = 'Raw File/Speciality BOM.xlsx'

CORE_RE = re.compile(r'-(B\d)-(\d+/\d+)-')


def base_of(drawing_no):
    return re.sub(r'-\d+$', '', drawing_no)


def core_key(line_no):
    if not line_no:
        return None
    m = CORE_RE.search(line_no)
    return f'{m.group(1)}-{m.group(2)}' if m else None


def build_agg(rows, keyfunc):
    agg = defaultdict(lambda: {'bases': set(), 'systems': set()})
    for r in rows:
        ln = r.get('line_no')
        if not ln:
            continue
        k = keyfunc(ln.strip())
        if not k:
            continue
        agg[k]['bases'].add(base_of(r['drawing_no']))
        if r.get('system'):
            agg[k]['systems'].add(r['system'].strip())
    return agg


dwg_rows = json.load(open(DWG_JSON, encoding='utf-8'))
exact_map = build_agg(dwg_rows, lambda ln: ln)
core_map = build_agg(dwg_rows, core_key)

wb = openpyxl.load_workbook(TARGET)
ws = wb['VALVELIST_BOP']

stats = {'system_filled': 0, 'iso_filled': 0, 'unresolved': []}

for r in range(2, ws.max_row + 1):
    line_no = ws.cell(r, 3).value
    tag = ws.cell(r, 4).value
    if not line_no:
        continue
    line_no = str(line_no).strip()

    entry = exact_map.get(line_no)
    matched_by = 'exact'
    if entry is None:
        ck = core_key(line_no)
        entry = core_map.get(ck) if ck else None
        matched_by = 'core'

    system_val = None
    iso_val = None
    if entry:
        if len(entry['systems']) == 1:
            system_val = next(iter(entry['systems']))
        if len(entry['bases']) == 1:
            iso_val = next(iter(entry['bases']))

    if system_val:
        ws.cell(r, 1, system_val)
        stats['system_filled'] += 1
    if iso_val:
        ws.cell(r, 2, iso_val)
        stats['iso_filled'] += 1
    if not system_val and not iso_val:
        stats['unresolved'].append((tag, line_no))

wb.save(TARGET)

print(f"SYSTEM 채움: {stats['system_filled']}")
print(f"ISO DRAWING 채움: {stats['iso_filled']}")
print(f"둘 다 못 채운 행: {len(stats['unresolved'])}")
with open('scratch/speciality_unresolved.txt', 'w', encoding='utf-8') as f:
    for tag, ln in stats['unresolved']:
        f.write(f'{tag}\t{ln}\n')
