# FF Type Flange WNRF → WNFF matcode 수정 SQL 생성
import openpyxl, requests, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SUPABASE_URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
HEADERS = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}', 'Accept-Profile': 'material'}

def api(path):
    r = requests.get(f'{SUPABASE_URL}/rest/v1/{path}', headers=HEADERS)
    return r.json()

# DN → D-code 매핑
SIZE_MAP = {
    'DN 80': 'D030', 'DN80': 'D030',
    'DN 100': 'D040', 'DN100': 'D040',
    'DN 125': 'D050', 'DN125': 'D050',
    'DN 150': 'D060', 'DN150': 'D060',
    'DN 200': 'D080', 'DN200': 'D080',
    'DN 250': 'D100', 'DN250': 'D100',
    'DN 400': 'D160', 'DN400': 'D160',
}
# 재질 → matl_code
MATL_MAP = {'A105': 'CS05', 'A182-F304': 'SS04'}
# 두께/클래스 → class_code
def get_class(thick):
    t = str(thick).upper()
    if 'CL150' in t or 'C150' in t: return 'C150'
    if 'CL300' in t or 'C300' in t: return 'C300'
    if 'CL1500' in t or 'S-120' in t: return 'S120'
    return t

# ── FF Type Flange List 읽기 ──────────────────────────────────
wb = openpyxl.load_workbook('Raw File/FF Type Flange List.xlsx', data_only=True)
ws = wb.active
ff_demand = {}   # (pkg_no_short, d_code, matl_code, class_code) → qty_ff

for r in range(3, ws.max_row + 1):
    pkg_full = ws.cell(r, 2).value
    matl     = ws.cell(r, 4).value
    size2    = ws.cell(r, 6).value  # DN size
    thick    = ws.cell(r, 7).value
    qty      = ws.cell(r, 9).value
    if not pkg_full or not matl:
        continue
    parts    = str(pkg_full).strip().split('-')
    pkg_no   = '-'.join(parts[3:]) if len(parts) > 3 else str(pkg_full)
    matl_s   = str(matl).strip()
    size_s   = str(size2).strip()
    d_code   = SIZE_MAP.get(size_s, '')
    m_code   = MATL_MAP.get(matl_s, '')
    c_code   = get_class(thick)
    if not d_code or not m_code:
        print(f'  [경고] 매핑 불가: pkg={pkg_no} matl={matl_s} size={size_s} thick={thick}')
        continue
    key = (pkg_no, d_code, m_code, c_code)
    ff_demand[key] = ff_demand.get(key, 0) + (int(qty) if qty else 0)

print('=== FF Demand (합산) ===')
for k, v in sorted(ff_demand.items()):
    print(f'  {k} → {v}EA')

# ── DB receiving 조회 ─────────────────────────────────────────
db_rows = []
for doc in ['PGU-DE-0525']:   # 0503은 이미 FF
    rows = api(f'receiving?select=id,doc_no,pkg_no,mat_code,qty,full_description&doc_no=eq.{doc}&mat_code=like.FLN*')
    db_rows.extend(rows if isinstance(rows, list) else [])

print('\n=== DB 현황 (0525 FLN) ===')
for r in sorted(db_rows, key=lambda x: (x['pkg_no'], x['mat_code'])):
    print(f'  id={r["id"]:4}  pkg={r["pkg_no"]:25}  mc={r["mat_code"]:25}  qty={r["qty"]}')

# ── matcode_master에서 FF 버전 존재 여부 확인 ─────────────────
need_ff_codes = set()
for (pkg_no, d, m, c), qty in ff_demand.items():
    if 'SS04' in m or m == 'SS04':
        need_ff_codes.add(f'FLN-{m}-{d}-{c}-FF')

existing_mc = {r['mat_code'] for r in api('matcode_master?select=mat_code&mat_code=like.FLN-SS04*')}
print(f'\n=== matcode_master SS04 FLN 현황 ===')
for mc in sorted(existing_mc):
    print(f'  {mc}')

missing_ff = sorted(need_ff_codes - existing_mc)
print(f'\n=== matcode_master에 없는 FF 코드 ({len(missing_ff)}건) ===')
for mc in missing_ff:
    print(f'  {mc}')

# ── SQL 생성 ──────────────────────────────────────────────────
sql_lines = ['-- FF Type Flange: WNRF → WNFF 수정 SQL', '']

# PART 1: matcode_master 신규 FF 코드 추가
if missing_ff:
    sql_lines.append('-- PART 1: matcode_master FF 코드 추가')
    # RF 기준 데이터로 FF 복사
    rf_master = {r['mat_code']: r for r in api('matcode_master?select=*&mat_code=like.FLN-SS04*')}
    for ff_mc in missing_ff:
        rf_mc = ff_mc.replace('-FF', '-RF')
        base  = rf_master.get(rf_mc, {})
        cols  = 'mat_code, category, item_desc, matl_desc, size1, size2, class_desc, et_desc'
        vals  = (
            f"'{ff_mc}', 'Fitting', "
            f"'{base.get('item_desc','FLANGE')}', '{base.get('matl_desc','')}', "
            f"'{base.get('size1','')}', '{base.get('size2','')}', "
            f"'{base.get('class_desc','')}', 'WNFF'"
        )
        sql_lines.append(f"INSERT INTO material.matcode_master ({cols}) VALUES ({vals}) ON CONFLICT (mat_code) DO NOTHING;")
    sql_lines.append('')

# PART 2: receiving 업데이트 / 분할
sql_lines.append('-- PART 2: receiving WNRF → WNFF 수정')

# DB rows를 key로 인덱싱
def parse_mc(mc):
    p = mc.split('-')
    return {'m': p[1], 'd': p[2], 'c': p[3], 'et': p[4]} if len(p) == 5 else {}

for row in db_rows:
    mc   = row['mat_code']
    info = parse_mc(mc)
    if info.get('et') == 'FF':
        continue   # 이미 FF
    m, d, c = info.get('m',''), info.get('d',''), info.get('c','')
    pkg_no  = row['pkg_no'].replace('PGU-DE-0525-', '').replace('PGU-DE-0503-', '')
    key     = (pkg_no, d, m, c)
    qty_ff  = ff_demand.get(key, 0)
    db_qty  = int(row['qty'])
    rid     = row['id']

    if qty_ff == 0:
        print(f'  [스킵] id={rid} {mc} → FF 수요 없음 (키={key})')
        continue

    ff_mc = f'FLN-{m}-{d}-{c}-FF'

    if qty_ff >= db_qty:
        # 전량 FF로 변경
        sql_lines.append(
            f"UPDATE material.receiving SET mat_code = '{ff_mc}' WHERE id = {rid};  "
            f"-- {row['pkg_no']} | {mc} qty={db_qty} → 전량 FF"
        )
    else:
        # 분할: qty_ff개 → FF (qty 변경), 나머지 → RF 신규 행
        sql_lines.append(
            f"UPDATE material.receiving SET mat_code = '{ff_mc}', qty = {qty_ff} WHERE id = {rid};  "
            f"-- {row['pkg_no']} | {mc} qty={db_qty} → FF {qty_ff}EA"
        )
        rf_qty = db_qty - qty_ff
        sql_lines.append(
            f"INSERT INTO material.receiving (doc_no, pkg_no, mat_code, qty, unit, category, full_description) "
            f"SELECT doc_no, pkg_no, '{mc}', {rf_qty}, unit, category, full_description "
            f"FROM material.receiving WHERE id = {rid};  "
            f"-- 나머지 RF {rf_qty}EA 신규 행"
        )

sql_lines.append('')
sql_out = '\n'.join(sql_lines)
with open('scratch/fix_ff_flange.sql', 'w', encoding='utf-8') as f:
    f.write(sql_out)
print(f'\n=== 생성된 SQL ===')
print(sql_out)
print(f'\nSQL 저장: scratch/fix_ff_flange.sql')
