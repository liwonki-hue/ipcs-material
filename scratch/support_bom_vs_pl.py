# Support BOM vs Packing List 매칭 분석 → 미입고/불일치 항목 엑셀 출력
import requests, json, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Accept-Profile': 'material'}

PL_PATH = 'Raw File/Support Packing List.xlsx'
OUT_PATH = 'Raw File/Support_BOM_vs_PL.xlsx'

# ── 1. Support BOM 전체 로드 (Supabase, 페이지네이션) ──
print('Loading Support BOM from DB...')
bom_rows = []
step = 1000
offset = 0
while True:
    r = requests.get(
        f'{URL}/rest/v1/support_bom',
        headers=H,
        params={
            'select': 'system,iso_dwg_no,support_tag,part_no,id_no,item,matl,size_or_type,length_mm,qty',
            'order': 'support_tag.asc,part_no.asc',
            'limit': step,
            'offset': offset
        }
    )
    batch = r.json()
    if not batch:
        break
    bom_rows.extend(batch)
    offset += step
    if len(batch) < step:
        break
print(f'  BOM rows: {len(bom_rows)}')

# ── 2. Support Packing List 로드 (Excel) ──
print('Loading Packing List...')
wb_pl = openpyxl.load_workbook(PL_PATH, data_only=True)
ws_pl = wb_pl['Packing']
pl_rows = []
for row in ws_pl.iter_rows(min_row=2, values_only=True):
    pkg, pkg_no, tag, part_no, id_no, item, matl, size, length, qty = row
    if not tag:
        continue
    pl_rows.append({
        'pkg': str(pkg).strip() if pkg else '',
        'pkg_no': str(pkg_no).strip() if pkg_no else '',
        'support_tag': str(tag).strip(),
        'part_no': str(part_no).strip() if part_no is not None else '',
        'id_no': str(id_no).strip() if id_no else '',
        'item': str(item).strip() if item else '',
        'matl': str(matl).strip() if matl else '',
        'size_or_type': str(size).strip() if size else '',
        'length_mm': str(length).strip() if length else '',
        'qty': float(qty) if qty is not None else 0,
    })
print(f'  PL rows: {len(pl_rows)}')

# ── 3. PL을 TAG+PART_NO 키로 인덱싱 ──
pl_index = defaultdict(list)
for p in pl_rows:
    key = (p['support_tag'], p['part_no'])
    pl_index[key].append(p)

# ── 4. BOM 각 행 매칭 분류 ──
not_in_pl   = []   # BOM에 있으나 PL에 없음 → 미출하
qty_diff    = []   # PL에 있으나 수량 불일치
matched     = []   # 완전 매칭

for b in bom_rows:
    tag = str(b['support_tag']).strip() if b['support_tag'] else ''
    pno = str(b['part_no']).strip() if b['part_no'] is not None else ''
    bom_qty = float(b['qty']) if b['qty'] is not None else 0
    key = (tag, pno)

    if key not in pl_index:
        not_in_pl.append(b)
    else:
        pl_qty = sum(p['qty'] for p in pl_index[key])
        if abs(pl_qty - bom_qty) > 0.001:
            qty_diff.append({**b, 'pl_qty': pl_qty, 'diff': bom_qty - pl_qty})
        else:
            matched.append(b)

print(f'\n결과:')
print(f'  완전 매칭:     {len(matched):,}건')
print(f'  수량 불일치:   {len(qty_diff):,}건')
print(f'  PL 미등록:     {len(not_in_pl):,}건')

# ── 5. 엑셀 출력 ──
wb = openpyxl.Workbook()

HEADER_FILL   = PatternFill('solid', fgColor='1F3864')
RED_FILL      = PatternFill('solid', fgColor='FFCCCC')
YELLOW_FILL   = PatternFill('solid', fgColor='FFF2CC')
GREEN_FILL    = PatternFill('solid', fgColor='E2EFDA')
SUMMARY_FILL  = PatternFill('solid', fgColor='D6E4F0')

header_font  = Font(bold=True, color='FFFFFF', size=10)
bold         = Font(bold=True, size=10)
normal       = Font(size=10)
center       = Alignment(horizontal='center', vertical='center', wrap_text=False)
thin_border  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def set_header(ws, headers, fills=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fills[col-1] if fills else HEADER_FILL
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border

def write_row(ws, row_no, values, fill=None):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row_no, column=col, value=v)
        cell.font = normal
        cell.alignment = center
        cell.border = thin_border
        if fill:
            cell.fill = fill

BOM_COLS = ['SYSTEM', 'ISO DWG NO', 'SUPPORT TAG', 'PART NO', 'ID NO', 'ITEM', 'MATL', 'SIZE OR TYPE', 'LENGTH(mm)', 'BOM QTY']

def bom_vals(b):
    return [b.get('system'), b.get('iso_dwg_no'), b.get('support_tag'),
            b.get('part_no'), b.get('id_no'), b.get('item'),
            b.get('matl'), b.get('size_or_type'), b.get('length_mm'), b.get('qty')]

# ── Sheet 1: Summary ──
ws_sum = wb.active
ws_sum.title = 'Summary'
ws_sum.column_dimensions['A'].width = 28
ws_sum.column_dimensions['B'].width = 14
ws_sum.column_dimensions['C'].width = 18

summary_data = [
    ('항목', '건수', '비고'),
    ('Support BOM 전체', len(bom_rows), ''),
    ('PL 전체', len(pl_rows), ''),
    ('', '', ''),
    ('완전 매칭 (BOM=PL qty)', len(matched), '입고 가능'),
    ('수량 불일치 (BOM≠PL)', len(qty_diff), '확인 필요'),
    ('PL 미등록 (BOM only)', len(not_in_pl), '미출하 / 누락'),
]
for r_no, row in enumerate(summary_data, 1):
    for c_no, val in enumerate(row, 1):
        cell = ws_sum.cell(row=r_no, column=c_no, value=val)
        cell.alignment = center
        cell.border = thin_border
        if r_no == 1:
            cell.fill = HEADER_FILL
            cell.font = header_font
        elif r_no in (5, 6, 7):
            fills = [GREEN_FILL, YELLOW_FILL, RED_FILL]
            cell.fill = fills[r_no - 5]
            cell.font = bold
        else:
            cell.font = normal

# ── Sheet 2: PL 미등록 (미출하) ──
ws1 = wb.create_sheet('PL 미등록 (미출하)')
set_header(ws1, BOM_COLS)
for i, b in enumerate(not_in_pl, 2):
    write_row(ws1, i, bom_vals(b), RED_FILL)
col_widths = [12, 30, 40, 8, 10, 20, 12, 20, 12, 10]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[1].height = 18

# ── Sheet 3: 수량 불일치 ──
ws2 = wb.create_sheet('수량 불일치')
diff_cols = BOM_COLS + ['PL QTY', 'DIFF (BOM-PL)']
set_header(ws2, diff_cols)
for i, b in enumerate(qty_diff, 2):
    vals = bom_vals(b) + [b['pl_qty'], b['diff']]
    write_row(ws2, i, vals, YELLOW_FILL)
for i, w in enumerate(col_widths + [10, 14], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 18

# ── Sheet 4: 완전 매칭 ──
ws3 = wb.create_sheet('완전 매칭')
set_header(ws3, BOM_COLS)
for i, b in enumerate(matched, 2):
    write_row(ws3, i, bom_vals(b), GREEN_FILL)
for i, w in enumerate(col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 18

wb.save(OUT_PATH)
print(f'\n저장 완료: {OUT_PATH}')
