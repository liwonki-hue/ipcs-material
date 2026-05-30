# Speciality BOM.xlsx 수정: 오기재 TAG 3건 수정 + STEAM TRAP/AIR TRAP 82건 추가
import openpyxl, sys, io, shutil
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BOM_PATH = 'Raw File/Speciality BOM.xlsx'
IL_PATH  = 'Raw File/Speciality Item List.xlsx'

# ── MatCode 매핑 (material → matcode prefix) ──────────────────
def gen_matcode(prefix, material, size, socket_rating, end_type):
    mat  = str(material or '').split('-')[0].strip()
    sz   = str(size or '').replace('"', '').strip()
    rt   = str(socket_rating or '').replace('#', '').strip()
    et   = str(end_type or '').strip()
    return f"{prefix}-{mat}-{sz}-{rt}-{et}"

# ── Item List STEAM TRAP / AIR TRAP 데이터 수집 ───────────────
SHEET_CFG = {
    '2. STEAM TRAP': {'prefix': 'STP', 'cat_label': 'STEAM TRAP'},
    '10.AIR TRAP':   {'prefix': 'ATP', 'cat_label': 'AIR TRAP'},
}
TAG_COL = 20

wb_il = openpyxl.load_workbook(IL_PATH, data_only=True)

# BOM에서 이미 올바른 TAG를 얻어놓기 (수정 후 기준)
WRONG_TO_CORRECT = {
    'B0-SP-28442': 'B0-SP-28445',
    'B0-SP-28443': 'B0-SP-28449',
    'B1-AT-40106': 'B0-AT-40011',
}
correct_tags_in_bom = set(WRONG_TO_CORRECT.values())  # 이미 BOM에 있을 예정

il_rows = []  # {tag, matcode, system, iso, line_no, full_desc, type2, size, matl, rating, end_type, qty}
for sh, cfg in SHEET_CFG.items():
    ws = wb_il[sh]
    for r in range(5, ws.max_row + 1):
        tag = ws.cell(r, TAG_COL).value
        if not tag:
            continue
        tag_s = str(tag).strip()
        if not any(tag_s.startswith(p) for p in ('B0-', 'B1-', 'B2-')):
            continue

        material      = ws.cell(r, 6).value
        size          = ws.cell(r, 7).value
        socket_rating = ws.cell(r, 9).value
        end_type      = ws.cell(r, 24).value
        matcode = gen_matcode(cfg['prefix'], material, size, socket_rating, end_type)

        desc = str(ws.cell(r, 3).value or '').strip()
        il_rows.append({
            'tag':       tag_s,
            'matcode':   matcode,
            'system':    str(ws.cell(r, 19).value or '').strip(),
            'iso':       str(ws.cell(r, 17).value or '').strip(),
            'line_no':   str(ws.cell(r, 2).value or '').strip(),
            'full_desc': f"{cfg['cat_label']}, {desc}",
            'type2':     str(ws.cell(r, 22).value or '').strip(),
            'size':      str(size or '').strip(),
            'matl':      str(material or '').strip(),
            'rating':    str(socket_rating or '').strip(),
            'end_type':  str(end_type or '').strip(),
            'qty':       ws.cell(r, 21).value or 1,
        })

print(f'Item List 수집: {len(il_rows)}건')

# ── BOM 로드 ──────────────────────────────────────────────────
wb_bom = openpyxl.load_workbook(BOM_PATH)
ws_bom = wb_bom.active

# ── STEP 1: 오기재 TAG 3건 수정 ───────────────────────────────
fixed = 0
for r in range(2, ws_bom.max_row + 1):
    tag_cell = ws_bom.cell(r, 3)
    old_tag  = str(tag_cell.value or '').strip()
    if old_tag in WRONG_TO_CORRECT:
        new_tag = WRONG_TO_CORRECT[old_tag]
        new_sys = new_tag.split('-')[0]  # B0 / B1 / B2
        print(f'  [수정] row{r}: {old_tag} → {new_tag} (SYSTEM: {ws_bom.cell(r,4).value} → {new_sys})')
        tag_cell.value         = new_tag
        ws_bom.cell(r, 4).value = new_sys
        fixed += 1

print(f'TAG 수정: {fixed}건\n')

# ── STEP 2: 누락 82건 추가 ────────────────────────────────────
# 현재 BOM에 있는 TAG 집합 (수정 반영 후)
existing_tags = set()
for r in range(2, ws_bom.max_row + 1):
    v = ws_bom.cell(r, 3).value
    if v:
        existing_tags.add(str(v).strip())

to_add = [row for row in il_rows if row['tag'] not in existing_tags]
print(f'추가 대상: {len(to_add)}건')

start_row = ws_bom.max_row + 1
for i, row in enumerate(to_add):
    r = start_row + i
    ws_bom.cell(r, 1).value  = 'Speciality'
    ws_bom.cell(r, 2).value  = row['matcode']
    ws_bom.cell(r, 3).value  = row['tag']
    ws_bom.cell(r, 4).value  = row['system']
    ws_bom.cell(r, 5).value  = row['iso']
    ws_bom.cell(r, 6).value  = row['line_no']
    ws_bom.cell(r, 7).value  = row['full_desc']
    ws_bom.cell(r, 8).value  = row['type2']
    ws_bom.cell(r, 9).value  = row['size']
    ws_bom.cell(r, 10).value = row['matl']
    ws_bom.cell(r, 11).value = row['rating']
    ws_bom.cell(r, 12).value = row['end_type']
    ws_bom.cell(r, 13).value = 'EA'
    ws_bom.cell(r, 14).value = row['qty']

print(f'행 추가 완료: row {start_row} ~ {start_row + len(to_add) - 1}')

# ── 저장 ──────────────────────────────────────────────────────
wb_bom.save(BOM_PATH)
print(f'\n저장 완료: {BOM_PATH}')
print(f'최종 행 수: {ws_bom.max_row - 1}건 (헤더 제외)')

# ── 최종 검증 ──────────────────────────────────────────────────
print('\n=== 최종 검증 ===')
wb_check = openpyxl.load_workbook(BOM_PATH, data_only=True)
ws_check = wb_check.active
final_tags = {str(ws_check.cell(r,3).value).strip()
              for r in range(2, ws_check.max_row+1)
              if ws_check.cell(r,3).value}

il_tag_set = {row['tag'] for row in il_rows}
still_missing = il_tag_set - final_tags
wrong_remain  = set(WRONG_TO_CORRECT.keys()) & final_tags

print(f'Item List 총: {len(il_tag_set)}개  |  BOM 최종 TAG 수: {len(final_tags)}개')
print(f'여전히 누락: {len(still_missing)}개  |  오기재 잔존: {len(wrong_remain)}개')
if still_missing:
    for t in sorted(still_missing):
        print(f'  누락: {t}')
if wrong_remain:
    for t in sorted(wrong_remain):
        print(f'  오기재 잔존: {t}')

# 신규 MatCode 목록 출력
new_matcodes = set()
for row in to_add:
    new_matcodes.add(row['matcode'])
print(f'\n신규 MatCode ({len(new_matcodes)}건) — matcode_master 추가 필요:')
for mc in sorted(new_matcodes):
    print(f'  {mc}')
