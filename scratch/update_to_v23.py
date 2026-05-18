# PO_List_Final_v22 → v23: COUPLING-HALF 사이즈 정규화 + 미매칭 분析 시트 삭제
import sys, re, openpyxl
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

IN_FILE  = 'c:/Users/PCLOVE/Downloads/ipcs-material/Raw File/PO_List_Final_v22.xlsx'
OUT_FILE = 'c:/Users/PCLOVE/Downloads/ipcs-material/Raw File/PO_List_Final_v23.xlsx'

# verify_v16_detail.py 최신 결과 (MISSING=0 확정)
VERIFY_SUMMARY = {
    '0443': (210, 210, 0, 0),
    '0503': (62,  62,  0, 0),
    '0524': (81,  81,  0, 0),
    '0525': (25,  25,  0, 0),
    '0554': (5,   5,   0, 0),
    '0555': (20,  20,  0, 0),
    '0565': (10,  10,  0, 0),
    '0574': (10,  10,  0, 0),
}
PL_ORDER = ['0443','0503','0524','0525','0554','0555','0565','0574']

# ── 로드 ─────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(IN_FILE, data_only=True)

# 미매칭 시트 삭제
for sname in list(wb.sheetnames):
    if '미매칭' in sname:
        del wb[sname]
        print(f'시트 삭제: {sname}')

ws = wb['매핑검토표']

# ── 열 인덱스 ─────────────────────────────────────────────────────────────────
C = dict(item=1, matl=2, size=3, thick=4, et=5,
         bom_qty=6, bom_uom=7, po_qty=8, po_uom=9,
         diff=10, stat=11, pkg=12, po_no=13)
NCOLS = 13

SUMMARY_ITEMS = {'PL No', '합계', '0443','0503','0524','0525','0554','0555','0565','0574'}

def get_row(r):
    return {k: ws.cell(r, col).value for k, col in C.items()}

# ── 데이터 읽기 ───────────────────────────────────────────────────────────────
max_row = ws.max_row
all_rows = []
for r in range(2, max_row + 1):
    d = get_row(r)
    if d['item'] is None and d['matl'] is None:
        continue
    d['_r'] = r
    all_rows.append(d)

summary_rows       = [r for r in all_rows if str(r['item'] or '').strip() in SUMMARY_ITEMS]
coupling_rows      = [r for r in all_rows if str(r['item'] or '').strip().upper() == 'COUPLING-HALF']
coupling_full_rows = [r for r in all_rows if str(r['item'] or '').strip().upper() == 'COUPLING-FULL']
PIPE_KEYWORDS = {'PIPE SEAMLESS','PIPE WELDED','PIPE WELDED/SEAMLESS','PIPE SMLS','PIPE NIPPLE','PIPE'}
SMALL_BORE_DN = {15, 20, 25, 32, 40, 50}
MATL_ALIASES  = {
    'A182-F316H': 'A182-F316',  # H suffix = High Temp, 동일 자재
    'A234-WPC':   'A234-WPB',   # WPC ≡ WPB (carbon steel 동급)
}

def is_pipe(item):
    s = str(item or '').strip().upper()
    return any(s == p or s.startswith(p) for p in PIPE_KEYWORDS)

def norm_matl(m):
    s = str(m or '').strip()
    s = re.sub(r'\s+CL\s*\d+\b', '', s, flags=re.IGNORECASE)  # "CL1" class suffix 제거
    s = re.sub(r'((?:TP|WP)[A-Z0-9]+)W\b', r'\1', s)          # WPBW→WPB, TP304W→TP304 (W=Welded)
    return MATL_ALIASES.get(s, s)

def is_coupling(item):
    s = str(item or '').strip().upper()
    return s in ('COUPLING-HALF', 'COUPLING-FULL')

pipe_rows  = [r for r in all_rows
              if str(r['item'] or '').strip() not in SUMMARY_ITEMS
              and not is_coupling(r['item'])
              and is_pipe(r['item'])]
other_rows = [r for r in all_rows
              if str(r['item'] or '').strip() not in SUMMARY_ITEMS
              and not is_coupling(r['item'])
              and not is_pipe(r['item'])]

print(f'원본 — COUPLING-HALF: {len(coupling_rows)}건  COUPLING-FULL: {len(coupling_full_rows)}건  PIPE: {len(pipe_rows)}건  기타: {len(other_rows)}건  요약: {len(summary_rows)}건')

# ── COUPLING-HALF 정규화 ──────────────────────────────────────────────────────
def norm_size(s):
    """DN 700 X DN 25 → DN 25 (두 번째 DN만 유효 사이즈)"""
    if not s:
        return s
    parts = re.split(r'\s*[Xx]\s*', str(s).strip())
    return parts[-1].strip() if len(parts) >= 2 else str(s).strip()

def calc_stat(bom, po):
    b, p = float(bom or 0), float(po or 0)
    if b == 0 and p == 0: return '-'
    if b == 0:            return 'BOM없음'
    if p == 0:            return 'PO없음'
    if abs(b - p) < 0.01: return 'MATCH'
    return 'PO과잉' if p > b else 'PO부족'

merged = defaultdict(lambda: {
    'bom': 0.0, 'po': 0.0,
    'buom': 'EA', 'puom': 'EA',
    'pkgs': [], 'ponos': []
})

for r in coupling_rows:
    key = (
        norm_matl(r['matl']),
        norm_size(r['size']),
        str(r['thick'] or '').strip(),
        str(r['et']    or '').strip(),
    )
    d = merged[key]
    d['bom'] += float(r['bom_qty'] or 0)
    d['po']  += float(r['po_qty']  or 0)
    if r['bom_uom']: d['buom'] = r['bom_uom']
    if r['po_uom']:  d['puom'] = r['po_uom']
    if r['pkg']:     d['pkgs'].append(str(r['pkg']))
    if r['po_no']:   d['ponos'].append(str(r['po_no']))

new_coupling_rows = []
for (matl, size, thick, et), d in sorted(merged.items(), key=lambda x: (x[0][0], x[0][1])):
    bom  = d['bom'] if d['bom'] > 0 else None
    po   = d['po']  if d['po']  > 0 else None
    diff = round(float(bom or 0) - float(po or 0), 3) if (bom or po) else None
    stat = calc_stat(bom, po)
    pkgs  = list(dict.fromkeys(p for p in d['pkgs']  if p))
    ponos = list(dict.fromkeys(p for p in d['ponos'] if p))
    new_coupling_rows.append({
        'item': 'COUPLING-HALF', 'matl': matl, 'size': size,
        'thick': thick, 'et': et,
        'bom_qty': bom, 'bom_uom': d['buom'],
        'po_qty':  po,  'po_uom':  d['puom'],
        'diff': diff, 'stat': stat,
        'pkg':   ', '.join(pkgs)  or None,
        'po_no': ', '.join(ponos) or None,
    })

print(f'COUPLING-HALF 정규화: {len(coupling_rows)}행 → {len(new_coupling_rows)}행')
for nr in new_coupling_rows:
    print(f'  {nr["matl"]:15s} {nr["size"]:8s} {nr["thick"]:6s} '
          f'BOM={nr["bom_qty"] or 0:>6}  PO={nr["po_qty"] or 0:>6}  [{nr["stat"]}]')

# ── PIPE DN≤50 ET 정규화 (blank/BW → PE) ─────────────────────────────────────
def norm_pipe_et(et, size, pipe=True):
    """ET 정규화.
    pipe=True  (PIPE):    DN≤50 → PE,  DN≥65 → BW
    pipe=False (피팅):             DN≥65 → BW 만 적용
    """
    dns = re.findall(r'DN\s*(\d+)', str(size or '').upper())
    if dns:
        dn = int(dns[0])
        et_str = str(et or '').strip().upper()
        if pipe and dn in SMALL_BORE_DN:   # PIPE DN ≤ 50 → PE
            if et_str in ('', 'BW'):
                return 'PE'
        if dn >= 65:                       # DN ≥ 65 → BW (pipe & 피팅 공통)
            if et_str in ('', 'PE'):
                return 'BW'
    return str(et or '').strip()

merged_pipe = defaultdict(lambda: {
    'bom': 0.0, 'po': 0.0,
    'buom': 'M', 'puom': 'M',
    'pkgs': [], 'ponos': []
})
for r in pipe_rows:
    norm_et = norm_pipe_et(r['et'], r['size'])
    key = (
        str(r['item']  or '').strip(),
        norm_matl(r['matl']),
        str(r['size']  or '').strip(),
        str(r['thick'] or '').strip(),
        norm_et,
    )
    d = merged_pipe[key]
    d['bom'] += float(r['bom_qty'] or 0)
    d['po']  += float(r['po_qty']  or 0)
    if r['bom_uom']: d['buom'] = r['bom_uom']
    if r['po_uom']:  d['puom'] = r['po_uom']
    if r['pkg']:     d['pkgs'].append(str(r['pkg']))
    if r['po_no']:   d['ponos'].append(str(r['po_no']))

new_pipe_rows = []
for (item, matl, size, thick, et), d in sorted(merged_pipe.items()):
    bom  = round(d['bom'], 3) if d['bom'] > 0 else None
    po   = round(d['po'],  3) if d['po']  > 0 else None
    diff = round(float(bom or 0) - float(po or 0), 3) if (bom or po) else None
    stat = calc_stat(bom, po)
    pkgs  = list(dict.fromkeys(p for p in d['pkgs']  if p))
    ponos = list(dict.fromkeys(p for p in d['ponos'] if p))
    new_pipe_rows.append({
        'item': item, 'matl': matl, 'size': size,
        'thick': thick, 'et': et,
        'bom_qty': bom, 'bom_uom': d['buom'],
        'po_qty':  po,  'po_uom':  d['puom'],
        'diff': diff, 'stat': stat,
        'pkg':   ', '.join(pkgs)  or None,
        'po_no': ', '.join(ponos) or None,
    })

print(f'PIPE 정규화: {len(pipe_rows)}행 → {len(new_pipe_rows)}행 (DN≤50 ET blank/BW→PE 통합)')

# ── COUPLING-FULL 정규화 (A182-F316H → A182-F316 등 matl 통합) ───────────────
merged_cpf = defaultdict(lambda: {
    'bom': 0.0, 'po': 0.0,
    'buom': 'EA', 'puom': 'EA',
    'pkgs': [], 'ponos': []
})
for r in coupling_full_rows:
    key = (
        norm_matl(r['matl']),
        norm_size(r['size']),
        str(r['thick'] or '').strip(),
        str(r['et']    or '').strip(),
    )
    d = merged_cpf[key]
    d['bom'] += float(r['bom_qty'] or 0)
    d['po']  += float(r['po_qty']  or 0)
    if r['bom_uom']: d['buom'] = r['bom_uom']
    if r['po_uom']:  d['puom'] = r['po_uom']
    if r['pkg']:     d['pkgs'].append(str(r['pkg']))
    if r['po_no']:   d['ponos'].append(str(r['po_no']))

new_coupling_full_rows = []
for (matl, size, thick, et), d in sorted(merged_cpf.items(), key=lambda x: (x[0][0], x[0][1])):
    bom  = d['bom'] if d['bom'] > 0 else None
    po   = d['po']  if d['po']  > 0 else None
    diff = round(float(bom or 0) - float(po or 0), 3) if (bom or po) else None
    stat = calc_stat(bom, po)
    pkgs  = list(dict.fromkeys(p for p in d['pkgs']  if p))
    ponos = list(dict.fromkeys(p for p in d['ponos'] if p))
    new_coupling_full_rows.append({
        'item': 'COUPLING-FULL', 'matl': matl, 'size': size,
        'thick': thick, 'et': et,
        'bom_qty': bom, 'bom_uom': d['buom'],
        'po_qty':  po,  'po_uom':  d['puom'],
        'diff': diff, 'stat': stat,
        'pkg':   ', '.join(pkgs)  or None,
        'po_no': ', '.join(ponos) or None,
    })

print(f'COUPLING-FULL 정규화: {len(coupling_full_rows)}행 → {len(new_coupling_full_rows)}행')
for nr in new_coupling_full_rows:
    print(f'  {nr["matl"]:15s} {nr["size"]:8s} {nr["thick"]:6s} '
          f'BOM={nr["bom_qty"] or 0:>6}  PO={nr["po_qty"] or 0:>6}  [{nr["stat"]}]')

# ── 기타(ELBOW/TEE 등) ET 정규화 및 그룹핑 ──────────────────────────────────
merged_other = defaultdict(lambda: {
    'bom': 0.0, 'po': 0.0,
    'buom': 'EA', 'puom': 'EA',
    'pkgs': [], 'ponos': [],
    '_r': 999999,
})
for r in other_rows:
    try:
        bom_v = float(r['bom_qty'] or 0)
        po_v  = float(r['po_qty']  or 0)
    except (ValueError, TypeError):
        continue  # 중간 헤더행 스킵
    norm_et = norm_pipe_et(r['et'], r['size'], pipe=False)
    key = (
        str(r['item']  or '').strip(),
        norm_matl(r['matl']),
        str(r['size']  or '').strip(),
        str(r['thick'] or '').strip(),
        norm_et,
    )
    d = merged_other[key]
    d['bom'] += bom_v
    d['po']  += po_v
    if r['bom_uom']: d['buom'] = r['bom_uom']
    if r['po_uom']:  d['puom'] = r['po_uom']
    if r['pkg']:     d['pkgs'].append(str(r['pkg']))
    if r['po_no']:   d['ponos'].append(str(r['po_no']))
    if r['_r'] < d['_r']: d['_r'] = r['_r']  # 첫 등장 행번호 기록

new_other_rows = []
for (item, matl, size, thick, et), d in merged_other.items():
    bom  = d['bom'] if d['bom'] > 0 else None
    po   = d['po']  if d['po']  > 0 else None
    diff = round(float(bom or 0) - float(po or 0), 3) if (bom or po) else None
    stat = calc_stat(bom, po)
    pkgs  = list(dict.fromkeys(p for p in d['pkgs']  if p))
    ponos = list(dict.fromkeys(p for p in d['ponos'] if p))
    new_other_rows.append({
        'item': item, 'matl': matl, 'size': size,
        'thick': thick, 'et': et,
        'bom_qty': bom, 'bom_uom': d['buom'],
        'po_qty':  po,  'po_uom':  d['puom'],
        'diff': diff, 'stat': stat,
        'pkg':   ', '.join(pkgs)  or None,
        'po_no': ', '.join(ponos) or None,
        '_r': d['_r'],
    })

print(f'기타 정규화: {len(other_rows)}행 → {len(new_other_rows)}행 (ET 정규화+그룹핑)')

# ── 최종 행 순서 구성 ─────────────────────────────────────────────────────────
# 아이템 유형별 첫 등장 위치 기준으로 삽입 순서 결정
def first_r(rows):
    return min(r['_r'] for r in rows) if rows else 999999

slots = sorted([
    (first_r(coupling_rows),      'CPH'),
    (first_r(coupling_full_rows), 'CPF'),
    (first_r(pipe_rows),          'PIPE'),
    (first_r(other_rows),         'OTHER'),
], key=lambda x: x[0])

# new_other_rows는 _r 기준 정렬
new_other_rows.sort(key=lambda x: x['_r'])

group_map = {
    'CPH':   new_coupling_rows,
    'CPF':   new_coupling_full_rows,
    'PIPE':  new_pipe_rows,
    'OTHER': new_other_rows,
}
final_data = []
for _, gname in slots:
    rows = group_map[gname]
    # _r 제거 후 추가
    for row in rows:
        row.pop('_r', None)
        final_data.append(row)

# ── 시트 재작성 ───────────────────────────────────────────────────────────────
# 헤더 행(1행) 유지, 데이터 행 전체 삭제 후 재기록
ws.delete_rows(2, max_row - 1)

FILLS = {
    'MATCH':   PatternFill('solid', fgColor='C6EFCE'),
    'PO과잉':  PatternFill('solid', fgColor='FFEB9C'),
    'PO부족':  PatternFill('solid', fgColor='FFEB9C'),
    'PO없음':  PatternFill('solid', fgColor='FFC7CE'),
    'BOM없음': PatternFill('solid', fgColor='DDEBF7'),
}
THIN = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

def write_data_row(r_idx, row_data):
    stat = str(row_data.get('stat') or '').strip()
    fill = FILLS.get(stat)
    keys = ('item','matl','size','thick','et','bom_qty','bom_uom','po_qty','po_uom','diff','stat','pkg','po_no')
    vals = [row_data.get(k) for k in keys]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r_idx, c, v)
        cell.font   = Font(size=9)
        cell.border = THIN
        cell.alignment = Alignment(vertical='center', wrap_text=(c in (12, 13)))
        if fill:
            cell.fill = fill
    ws.row_dimensions[r_idx].height = 16

for i, row in enumerate(final_data, 2):
    write_data_row(i, row)

# ── 요약 테이블 (현행 verify 결과로 업데이트) ─────────────────────────────────
HDR_FILL = PatternFill('solid', fgColor='2F5496')
SMRY_HDR = Font(bold=True, color='FFFFFF', size=9)
SMRY_DAT = Font(size=9)

summary_start = len(final_data) + 3   # 데이터 끝 + 빈 행 1개 gap
hdr_labels = ['PL No', '총패키지', 'OK', 'DIFF', 'MISSING']
for c, label in enumerate(hdr_labels, 1):
    cell = ws.cell(summary_start, c, label)
    cell.fill = HDR_FILL
    cell.font = SMRY_HDR
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN

total = [0, 0, 0, 0]  # pkg, ok, diff, missing
for i, pl in enumerate(PL_ORDER, summary_start + 1):
    pkg, ok, diff, missing = VERIFY_SUMMARY[pl]
    for j in range(4): total[j] += [pkg, ok, diff, missing][j]
    for c, v in enumerate([pl, pkg, ok, diff, missing], 1):
        cell = ws.cell(i, c, v)
        cell.font = SMRY_DAT
        cell.border = THIN
        cell.alignment = Alignment(horizontal='center')

total_row = summary_start + len(PL_ORDER) + 1
for c, v in enumerate(['합계'] + total, 1):
    cell = ws.cell(total_row, c, v)
    cell.font = Font(bold=True, size=9)
    cell.border = THIN
    cell.alignment = Alignment(horizontal='center')

# ── 저장 ─────────────────────────────────────────────────────────────────────
wb.save(OUT_FILE)

print(f'\n저장 완료: {OUT_FILE}')
print(f'[매핑검토표] 총 {len(final_data)}행')
print(f'  COUPLING-HALF: {len(new_coupling_rows)}행 (정규화, 기존 {len(coupling_rows)}행)')
print(f'[요약] MISSING=0 반영 완료')
print(f'[삭제] 미매칭 분析 시트')
