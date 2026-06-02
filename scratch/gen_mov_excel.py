# PDF 13 (MOV List) / PDF 12 (Butterfly Valve) → Excel 변환

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 공통 스타일 ─────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
HDR_FONT   = Font(bold=True, color='FFFFFF', size=9)
GRP_FILL   = PatternFill('solid', fgColor='D6E4F0')
GRP_FONT   = Font(bold=True, size=9)
CELL_FONT  = Font(size=9)
CELL_ALIGN = Alignment(wrap_text=True, vertical='top')
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill   = HDR_FILL
        cell.font   = HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

def style_group(ws, row, cols, text):
    ws.cell(row, 1, text).fill   = GRP_FILL
    ws.cell(row, 1).font   = GRP_FONT
    ws.cell(row, 1).border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

def style_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font   = CELL_FONT
        cell.border = BORDER
        cell.alignment = CELL_ALIGN

# ══════════════════════════════════════════════════════════════════════
# 1. PDF 13 → MOV_List.xlsx
# ══════════════════════════════════════════════════════════════════════
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = 'MOV List'

COLS1 = ['System','P&ID No.','Description','GA Dwg\nSheet No.',
         'Tag No.','Q\'ty','Size\n[DN]','Valve Type','Valve Detail Type',
         'ANSI\nClass','Medium','Design\nPress.\n[bar.a]','Design\nTemp.\n[°C]',
         'Body Material\n(ASTM)']
NCOLS1 = len(COLS1)

# 헤더
ws1.row_dimensions[1].height = 40
for c, h in enumerate(COLS1, 1):
    ws1.cell(1, c, h)
style_header(ws1, 1, NCOLS1)

# 데이터
DATA1 = [
    # System, P&ID, Description, GA Sheet, Tags, Qty, DN, Type, Detail, Class, Medium, DP, DT, Material
    # ── HP & LP Steam System
    ('HP & LP Steam System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B126-PR-007-0001-001', 'HRSG #1 HP Steam Drain MOV (upstream NRV)',       5,  'B1-MOV-26001\nB2-MOV-26001', 2, 25,  'GLOBE', 'T-type, Plug Disc', '1500#', 'Steam',      96.4, 530, 'SA182-F91'),
    ('', 'CCP-PW-B126-PR-007-0001-001', 'HRSG #2 HP Steam Drain MOV (upstream NRV)',       5,  'B1-MOV-26002\nB2-MOV-26002', 2, 25,  'GLOBE', 'T-type, Plug Disc', '1500#', 'Steam',      96.4, 530, 'SA182-F91'),
    ('', 'CCP-PW-B126-PR-007-0001-001', 'HRSG #1 HP Steam Drain MOV (downstream NRV)',     6,  'B1-MOV-26003\nB2-MOV-26003', 2, 25,  'GLOBE', 'T-type, Plug Disc', '1500#', 'Steam',      96.4, 530, 'A182-F91'),
    ('', 'CCP-PW-B126-PR-007-0001-001', 'HRSG #2 HP Steam Drain MOV (downstream NRV)',     6,  'B1-MOV-26004\nB2-MOV-26004', 2, 25,  'GLOBE', 'T-type, Plug Disc', '1500#', 'Steam',      96.4, 530, 'A182-F91'),
    ('', 'CCP-PW-B126-PR-007-0001-001', 'Auxiliary Steam Supply MOV',                      7,  'B1-MOV-26011\nB2-MOV-26011', 2, 100, 'GATE',  'Parallel Slide Disc','1500#', 'Steam',      96.4, 530, 'A182-F91 or A217-C12A'),
    ('', 'CCP-PW-B126-PR-007-0001-001', 'HP Steam Dynamic Strainer Drain MOV',             8,  'B1-MOV-26012\nB2-MOV-26012', 2, 50,  'GLOBE', 'T-type, Plug Disc', '1500#', 'Steam',      96.4, 530, 'A182-F91'),
    ('', 'CCP-PW-B126-PR-007-0001-001', 'HP Steam Start-up Warming Drain MOV',             9,  'B1-MOV-26021\nB2-MOV-26021', 2, 80,  'GLOBE', 'T-type, Plug Disc', '1500#', 'Steam',      96.4, 530, 'A182-F91 or A217-C12A'),
    ('', 'CCP-PW-B127-PR-007-0001-002', 'HRSG #1 LP Steam Drain MOV (upstream NRV)',       10, 'B1-MOV-27001\nB2-MOV-27001', 2, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',      10,   292, 'SA105'),
    ('', 'CCP-PW-B127-PR-007-0001-002', 'HRSG #2 LP Steam Drain MOV (upstream NRV)',       10, 'B1-MOV-27002\nB2-MOV-27002', 2, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',      10,   292, 'SA105'),
    ('', 'CCP-PW-B127-PR-007-0001-002', 'HRSG #1 LP Steam Drain MOV (downstream NRV)',     10, 'B1-MOV-27003\nB2-MOV-27003', 2, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',      10,   292, 'SA105'),
    ('', 'CCP-PW-B127-PR-007-0001-002', 'HRSG #2 LP Steam Drain MOV (downstream NRV)',     10, 'B1-MOV-27004\nB2-MOV-27004', 2, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',      10,   292, 'SA105'),
    ('', 'CCP-PW-B127-PR-007-0001-002', 'LP Steam Common Drain MOV',                       29, 'B1-MOV-27011\nB2-MOV-27011', 2, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',      10,   292, 'A105'),
    # ── Steam Blowing System
    ('Steam Blowing System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'N/A', 'HP STEAM BLOWING MOV (Martyr valve)',                                      11, 'Later',                       1, 400, 'GATE',  'Flexible Wedge Disc','900#',  'Steam',      51,   400, 'A216-WCC'),
    ('', 'N/A', 'LP STEAM BLOWING MOV (Martyr valve)',                                      12, 'Later',                       1, 550, 'GATE',  'Flexible Wedge Disc','600#',  'Steam',      10,   292, 'A216-WCC'),
    # ── Auxiliary Steam System
    ('Auxiliary Steam System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B128-PR-007-0001-001', 'Start-up Auxiliary Boiler Supply MOV',            13, 'B0-MOV-28021',                1, 150, 'GATE',  'Flexible Wedge Disc','300#',  'Steam',      21,   326, 'A216-WCB'),
    ('', 'CCP-PW-B128-PR-007-0001-001', 'Neighboring Block Auxiliary Steam Header Supply MOV',13,'B1-MOV-28011\nB2-MOV-28011',2, 150, 'GATE',  'Flexible Wedge Disc','300#',  'Steam',      21,   326, 'A216-WCB'),
    ('', 'CCP-PW-B128-PR-007-0001-001', 'Turbine Gland Seal Steam Header Supply MOV',      14, 'B1-MOV-28012\nB2-MOV-28012', 2, 50,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',      21,   326, 'A105'),
    ('', 'CCP-PW-B128-PR-007-0001-001', 'Turbine Gland Seal Steam Header Warming Drain MOV',15,'B1-MOV-28013\nB2-MOV-28013', 2, 50,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',      21,   326, 'A105'),
    ('', 'CCP-PW-B128-PR-007-0001-001', 'Aux. PRDS Pressure Control Valve Drain MOV',      16, 'B1-MOV-28001\nB2-MOV-28001', 2, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',      21,   326, 'A105'),
    # ── Condensate System
    ('Condensate System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B129-PR-007-0001-001', 'CEP DISCHARGE MOV',                               17, 'B1-MOV-29001~29003\nB2-MOV-29001~29003', 6, 200, 'GATE', 'Flexible Wedge Disc','300#', 'Condensate', 36, 100, 'A216-WCB'),
    # ── Feedwater System
    ('Feedwater System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B130-PR-007-0001-001', 'HRSG #11 BFP A IP Discharge MOV',                 18, 'B1-MOV-30001A/31001A\nB2-MOV-30001A/31001A', 4, 80,  'GATE',  'Parallel Slide Disc','600#',  'Feedwater',  60,  195, 'A216-WCB'),
    ('', 'CCP-PW-B130-PR-007-0001-001', 'HRSG #11 BFP A HP Discharge MOV',                 19, 'B1-MOV-30002A/31002A\nB2-MOV-30002A/31002A', 4, 150, 'GATE',  'Parallel Slide Disc','1500#', 'Feedwater', 143, 198, 'A216-WCC'),
    ('', 'CCP-PW-B130-PR-007-0001-001', 'BYPASS MOV FOR HRSG #11 BFP A HP Discharge MOV',  20, 'B1-MOV-30003A/31003A\nB2-MOV-30003A/31003A', 4, 25,  'GLOBE', 'T-type, Plug Disc', '1500#', 'Feedwater', 143, 198, 'A105'),
    ('', 'CCP-PW-B130-PR-007-0001-001', 'HRSG #11 BFP B IP Discharge MOV',                 18, 'B1-MOV-30001B/31001B\nB2-MOV-30001B/31001B', 4, 80,  'GATE',  'Parallel Slide Disc','600#',  'Feedwater',  60,  195, 'A216-WCB'),
    ('', 'CCP-PW-B130-PR-007-0001-001', 'HRSG #11 BFP B HP Discharge MOV',                 19, 'B1-MOV-30002B/31002B\nB2-MOV-30002B/31002B', 4, 150, 'GATE',  'Parallel Slide Disc','1500#', 'Feedwater', 143, 198, 'A216-WCC'),
    ('', 'CCP-PW-B130-PR-007-0001-001', 'BYPASS MOV FOR HRSG #11 BFP B HP Discharge MOV',  20, 'B1-MOV-30003B/31003B\nB2-MOV-30003B/31003B', 4, 25,  'GLOBE', 'T-type, Plug Disc', '1500#', 'Feedwater', 143, 198, 'A105'),
    # ── Closed Cooling Water System
    ('Closed Cooling Water System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B132-PR-007-0001-001', 'Closed Cooling Water Head Tank Make-up Line',     21, 'B1-MOV-32005\nB2-MOV-32005', 2, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'DM water',   13,   55,  'A182-F304'),
    # ── Hot Water Distribution System
    ('Hot Water Distribution System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B134-PR-007-0001-001', 'HOT WATER HEAD TANK MAKE UP LINE MOV',            22, 'B1-MOV-34004\nB2-MOV-34004', 2, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'DM water',   13,   55,  'A182-F304'),
    ('', 'CCP-PW-B134-PR-007-0001-001', 'HOT WATER SUPPLY PUMP DISCHARGE MOV',             23, 'B1-MOV-34001~34003\nB2-MOV-34001~34003', 6, 150, 'GLOBE', 'T-type, Plug Disc', '150#',  'Hot Water',  16,  120, 'A216-WCB'),
    ('', 'CCP-PW-B134-PR-007-0001-002', 'HOT WATER HEATER TUBE SIDE INLET MOV',            24, 'B1-MOV-34005/34008\nB2-MOV-34005/34008', 4, 150, 'GATE',  'Flexible Wedge Disc','150#',  'Hot Water',  16,  120, 'A216-WCB'),
    ('', 'CCP-PW-B134-PR-007-0001-002', 'HOT WATER HEATER TUBE SIDE OUTLET MOV',           24, 'B1-MOV-34011/34013\nB2-MOV-34011/34013', 4, 150, 'GATE',  'Flexible Wedge Disc','150#',  'Hot Water',  16,  120, 'A216-WCB'),
    ('', 'CCP-PW-B134-PR-007-0001-002', 'HOT WATER HEATER SHELL SIDE INLET MOV',           25, 'B1-MOV-34006/34009\nB2-MOV-34006/34009', 4, 250, 'GATE',  'Flexible Wedge Disc','300#',  'Steam',       5,  326, 'A216-WCB'),
    ('', 'CCP-PW-B134-PR-007-0001-002', 'HOT WATER HEATER SHELL SIDE VENT LINE BYPASS MOV',26,'B1-MOV-34007/34010\nB2-MOV-34007/34010', 4, 25,  'GLOBE', 'T-type, Plug Disc', '600#',  'Steam',       5,  326, 'A105'),
    # ── Demineralized Water System
    ('Demineralized Water System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B035-PR-007-0001-001', 'Demi plant to Demi tank A inlet MOV',             27, 'B0-MOV-35001',                1, 200, 'GATE',  'Flexible Wedge Disc','150#',  'DM water',    5,   55, 'A351-CF8'),
    ('', 'CCP-PW-B035-PR-007-0001-001', 'Demi plant to Demi tank B inlet MOV',             27, 'B0-MOV-35002',                1, 200, 'GATE',  'Flexible Wedge Disc','150#',  'DM water',    5,   55, 'A351-CF8'),
    # ── Fuel Oil System
    ('Fuel Oil System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B046-PR-007-0001-002', 'Fuel Oil Transfer Pump-A Discharge MOV',          28, 'B0-MOV-46001',                1, 400, 'GATE',  'Flexible Wedge Disc','150#',  'Diesel Oil', 16,   55, 'A216-WCB'),
    ('', 'CCP-PW-B046-PR-007-0001-002', 'Fuel Oil Transfer Pump-B Discharge MOV',          28, 'B0-MOV-46002',                1, 400, 'GATE',  'Flexible Wedge Disc','150#',  'Diesel Oil', 16,   55, 'A216-WCB'),
    # ── Service Water System
    ('Service Water System', None, None, None, None, None, None, None, None, None, None, None, None, None),
    ('', 'CCP-PW-B133-PR-007-0001-001', 'Drain Quenching water MOV',                       30, 'B1-MOV-33001\nB2-MOV-33001', 2, 50,  'GLOBE', 'T-type, Plug Disc', '600#',  'Service Water',10,  50,  'A105'),
]

r = 2
for row in DATA1:
    system = row[0]
    if system and row[1] is None:
        style_group(ws1, r, NCOLS1, system)
    else:
        for c, val in enumerate(row, 1):
            ws1.cell(r, c, val)
        style_row(ws1, r, NCOLS1)
    r += 1

# 열 너비
widths1 = [22, 30, 50, 8, 28, 5, 7, 8, 18, 7, 12, 8, 7, 20]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'

wb1.save('Raw File/MOV_List_Rev01.xlsx')
print('MOV_List_Rev01.xlsx 저장 완료')

# ══════════════════════════════════════════════════════════════════════
# 2. PDF 12 → MOV_Butterfly_Valve.xlsx
# ══════════════════════════════════════════════════════════════════════
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Butterfly Valve'

COLS2 = ['NO.','REV.','Tag No.','P&ID No.','Valve Description',
         'Valve\nModel','Rating','Size\n(inch)','Size\n(DN)','Q\'ty',
         'Type','Drilling','End','Body','Disc','Seat','Stem',
         'Design\nPress.\n[bar]','Work\nPress.\n[bar]',
         'Design\nTemp.\n[°C]','Work\nTemp.\n[°C]','Operator']
NCOLS2 = len(COLS2)

ws2.row_dimensions[1].height = 45
for c, h in enumerate(COLS2, 1):
    ws2.cell(1, c, h)
style_header(ws2, 1, NCOLS2)

DATA2 = [
    # NO, REV, Tags, P&ID, Description, Model, Rating, Inch, DN, Qty, Type, Drilling, End, Body, Disc, Seat, Stem, DP, WP, DT, WT, Operator
    (1,'C01','B1-MOV-32001/002/003\nB2-MOV-32001/002/003','CCP-PW-B132-PR-007-0001-001','Each CCWP Outlet Line','SRS912','150LB',14,350,6,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A351 CF8','EPDM','A479 T410',10,7,80,55,'MOTOR'),
    (2,'C01','B1-MOV-32004\nB2-MOV-32004','CCP-PW-B132-PR-007-0001-001','Each CCW FFC Bypass Line','SRS912','150LB',20,500,2,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A351 CF8','EPDM','A479 T410',10,7,80,48,'MOTOR'),
    (3,'C01','B1-MOV-34015/014\nB2-MOV-34015/017','CCP-PW-B134-PR-007-0001-003','Each GT Air Preheater Supply MOV','SRS912','150LB',6,150,4,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A216 WCB','EPDM','A479 T410',16,10,120,90,'MOTOR'),
    (4,'C01','B0-MOV-38001/002','CCP-PW-B038-PR-007-0001-001','External raw water supply to raw water tank inlet MOV','SRS912','150LB',10,250,2,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A351 CF8','A351 CF8','EPDM','A479 T410',10,3,50,44.2,'MOTOR'),
    (5,'C01','B0-MOV-38003/004/005/006','CCP-PW-B038-PR-007-0001-001','Raw Water Discharge to Raw Water Distribution Line Common','SRS912','150LB',6,150,4,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A351 CF8','EPDM','A479 T410',10,4.1,55,44.2,'MOTOR'),
    (6,'C01','B0-MOV-36001/002','CCP-PW-B036-PR-007-0001-001','External Service Water Supply to Service Water Tank Inlet MOV','SRS913','150LB',3,80,2,'FLANGE','ASME B16.5 150LB RF','MSS SP67 NARROW','A351 CF8','A351 CF8','EPDM','A479 T410',10,3,50,44.2,'MOTOR'),
    (7,'C01','B0-MOV-36003/004/005/006','CCP-PW-B036-PR-007-0001-001','Service Water Discharge to Service Water Distribution Line Common','SRS913','150LB',4,100,4,'FLANGE','ASME B16.5 150LB RF','MSS SP67 NARROW','A216 WCB','A351 CF8','EPDM','A479 T410',10,4.7,55,44.2,'MOTOR'),
    (8,'C01','LATER','LATER','N/A','SRS912','150LB',8,200,2,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A216 WCB','A351 CF8','EPDM','A479 T410',10,3,55,35,'MOTOR'),
    (9,'C01','B0-MOV-37003/004/005/006','CCP-PW-B037-PR-007-0001-001','Potable Water Discharge to Potable Water Distribution Line Common','SRS912','150LB',6,150,4,'FLANGE','ASME B16.5 150LB RF','JIS F 7480 FLANGE','A351 CF8','A351 CF8','EPDM','A479 T410',10,6.5,55,35,'MOTOR'),
]

for i, row in enumerate(DATA2, 2):
    for c, val in enumerate(row, 1):
        ws2.cell(i, c, val)
    style_row(ws2, i, NCOLS2)

# 합계 행
total_row = len(DATA2) + 2
ws2.cell(total_row, 1, 'TOTAL')
ws2.cell(total_row, 10, sum(r[9] for r in DATA2))
for c in range(1, NCOLS2 + 1):
    ws2.cell(total_row, c).font = Font(bold=True, size=9)
    ws2.cell(total_row, c).border = BORDER

widths2 = [5, 5, 28, 30, 55, 8, 7, 6, 6, 5, 8, 22, 18, 10, 10, 10, 10, 7, 7, 7, 7, 8]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

wb2.save('Raw File/MOV_Butterfly_Valve_Rev_C01.xlsx')
print('MOV_Butterfly_Valve_Rev_C01.xlsx 저장 완료')
