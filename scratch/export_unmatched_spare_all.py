# 전체 PL 미매칭 Valve & SPARE 항목 통합 엑셀 출력 (PL 추가 시 PL_LIST에 append)
import openpyxl, sys, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

OUT_PATH = 'scratch/Unmatched_Spare_All.xlsx'

# ── 처리할 PL 목록 (신규 PL 등록 시 여기에만 추가) ───────────────────────────
PL_LIST = [
    ('PGU-DE-0390', 'Raw File/Packing List (편집)/Valve/PGU-DE-0390_BOP Piping-Manual Valve(확인).xlsx'),
    ('PGU-DE-0474', 'Raw File/Packing List (편집)/Valve/PGU-DE-0474_BOP Piping_Manual Valve(편집).xlsx'),
]

ACCESSORY_KEYWORDS = ('TAG PLATE', 'SEAL RING', 'HINGE PIN')

# ── BOM tag_map 구성 ─────────────────────────────────────────────────────────
bom_wb = openpyxl.load_workbook('Raw File/BOM Data/Valve BOM.xlsx')
bom_ws = bom_wb.active
tag_map = {str(bom_ws.cell(r, 7).value).strip()
           for r in range(2, bom_ws.max_row + 1)
           if bom_ws.cell(r, 7).value and bom_ws.cell(r, 1).value}

# ── 전체 PL 순회 ─────────────────────────────────────────────────────────────
unmatched, spares = [], []

for doc_no, fpath in PL_LIST:
    pl_wb = openpyxl.load_workbook(fpath, data_only=True)
    pl_ws = pl_wb.active
    pkg_no = None
    u_cnt = s_cnt = 0

    for r in range(2, pl_ws.max_row + 1):
        pkg  = pl_ws.cell(r, 2).value
        desc = pl_ws.cell(r, 3).value
        qty  = pl_ws.cell(r, 4).value
        unit = pl_ws.cell(r, 5).value
        tag  = pl_ws.cell(r, 6).value
        if pkg: pkg_no = str(pkg).strip()

        desc_str = str(desc).strip() if desc else ''
        tag_str  = str(tag).strip() if tag else ''
        if not desc_str and not tag_str:
            continue

        # Accessory 스킵 (BOM 미매칭 대상 아님)
        if any(kw in desc_str.upper() for kw in ACCESSORY_KEYWORDS):
            continue

        row = {
            'doc_no': doc_no,
            'pkg_no': pkg_no,
            'tag':    tag_str,
            'desc':   desc_str,
            'qty':    qty if qty is not None else 1,
            'unit':   str(unit).strip() if unit else 'EA',
            'remark': '',
        }

        if tag_str.upper() == 'SPARE':
            row['tag'] = 'SPARE'
            spares.append(row)
            s_cnt += 1
        elif tag_str and tag_str not in tag_map:
            unmatched.append(row)
            u_cnt += 1

    print(f'{doc_no}: 미매칭 {u_cnt}건, SPARE {s_cnt}건')

print(f'\n합계 — 미매칭: {len(unmatched)}건 / SPARE: {len(spares)}건')

# ── 스타일 정의 ──────────────────────────────────────────────────────────────
HDR_FILL_RED  = PatternFill('solid', fgColor='C00000')
HDR_FILL_BLUE = PatternFill('solid', fgColor='1F497D')
HDR_FONT      = Font(bold=True, color='FFFFFF', size=10)
TITLE_FONT    = Font(bold=True, size=12)
EVEN_FILL     = PatternFill('solid', fgColor='F2F2F2')
CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT          = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin   = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS   = ['No', 'DOC NO', 'Package No.', 'Tag No.', 'Description', "Q'ty", 'Unit', 'Remark']
WIDTHS = [5, 16, 30, 22, 42, 7, 7, 25]

def write_sheet(ws, title, data, hdr_fill):
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value, c.font, c.alignment = title, TITLE_FONT, CENTER

    for ci, (col, w) in enumerate(zip(COLS, WIDTHS), 1):
        ws.column_dimensions[ws.cell(2, ci).column_letter].width = w
        cell = ws.cell(2, ci, col)
        cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, hdr_fill, CENTER, BORDER

    for ri, row in enumerate(data, 3):
        fill = EVEN_FILL if ri % 2 == 0 else PatternFill()
        vals = [ri - 2, row['doc_no'], row['pkg_no'], row['tag'],
                row['desc'], row['qty'], row['unit'], row['remark']]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(ri, ci, val)
            cell.fill, cell.border = fill, BORDER
            cell.alignment = CENTER if ci in (1, 2, 6, 7) else LEFT
        ws.row_dimensions[ri].height = 15

# ── 워크북 생성 ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = 'Unmatched Valve'
write_sheet(ws1,
    f'전체 PL 미매칭 Valve 항목 (총 {len(unmatched)}건 — BOM DB Tag 없음)',
    unmatched, HDR_FILL_RED)

ws2 = wb.create_sheet('SPARE')
write_sheet(ws2,
    f'전체 PL SPARE 항목 (총 {len(spares)}건)',
    spares, HDR_FILL_BLUE)

wb.save(OUT_PATH)
print(f'\n저장 완료: {OUT_PATH}')
