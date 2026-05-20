# Large Bore HP/LP Steam 자재 삭제 SQL 생성
# - LB BOM 전용 ISO (134개): 전체 삭제
# - LB+SB 공통 ISO (416개): SB BOM에 없는 LB 전용 항목만 삭제
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import warnings
warnings.filterwarnings("ignore")
import openpyxl, re

# ─── 헬퍼 함수 (insert_missing_bom.py 동일) ──────────────────────────────

DN_TO_INCH = {
    'DN6': '1/4"', 'DN8': '3/8"', 'DN10': '3/8"',
    'DN15': '1/2"', 'DN20': '3/4"', 'DN25': '1"',
    'DN32': '1-1/4"', 'DN40': '1-1/2"', 'DN50': '2"',
    'DN65': '2-1/2"', 'DN80': '3"', 'DN100': '4"',
    'DN125': '5"', 'DN150': '6"', 'DN200': '8"',
    'DN250': '10"', 'DN300': '12"', 'DN350': '14"',
    'DN400': '16"', 'DN450': '18"', 'DN500': '20"',
    'DN600': '24"', 'DN700': '28"', 'DN750': '30"',
    'DN800': '32"', 'DN900': '36"', 'DN1000': '40"',
}

def dn_to_inch(size_str):
    s = re.sub(r'\s+', '', str(size_str).upper())
    if s in DN_TO_INCH:
        return DN_TO_INCH[s]
    m = re.match(r'(DN\d+)\s*[Xx]\s*(DN\d+)', s)
    if m:
        a = DN_TO_INCH.get(m.group(1), m.group(1))
        b = DN_TO_INCH.get(m.group(2), m.group(2))
        return f"{a} X {b}"
    return size_str

def norm_matl(matl):
    m = str(matl).strip()
    m = re.sub(r'^SA-?', 'A', m)
    m = re.sub(r'^SA(\d)', r'A\1', m)
    return m

def build_full_description(item, matl1, matl2, size, thick, et):
    parts = []
    parts.append(str(item).strip())
    matl = norm_matl(matl1) if matl1 else ''
    if matl: parts.append(matl)
    size_norm = dn_to_inch(size) if size else ''
    if size_norm: parts.append(size_norm)
    if thick: parts.append(str(thick).strip())
    if et: parts.append(str(et).strip())
    return ', '.join(filter(None, parts))

# DN65+ = Large Bore
LARGE_BORE_DNs = {
    'DN65', 'DN80', 'DN100', 'DN125', 'DN150', 'DN200',
    'DN250', 'DN300', 'DN350', 'DN400', 'DN450', 'DN500',
    'DN600', 'DN700', 'DN750', 'DN800', 'DN900', 'DN1000',
}

def is_large_bore(size_str):
    """DN65 이상이면 대구경(Large Bore)으로 판단"""
    s = re.sub(r'\s+', '', str(size_str).upper())
    if s in LARGE_BORE_DNs:
        return True
    # Reducer: DN80 X DN50 → 대구경으로 처리
    m = re.match(r'(DN\d+)[Xx](DN\d+)', s)
    if m:
        return m.group(1) in LARGE_BORE_DNs
    return False

def esc(s):
    return str(s).replace("'", "''")

# ─── 1. LB BOM 스캔: HP/LP Steam ISO 및 BOM 키 수집 ─────────────────────

print("=== LB BOM 스캔 중... ===")
LB_FILE = "Raw File/LARGE BORE BOM(251223).xlsm"
LB_SHEET = "Piping&Fitting"
LB_HEADER = 8
LB_UOM_COL = 23
LB_QTY_COL = 24

lb_hp_lp_isos = set()   # HP/LP Steam인 ISO 목록
lb_all_isos = set()      # LB BOM 전체 ISO
lb_keys_by_iso = {}      # iso -> set of (line, full_desc, uom)

wb = openpyxl.load_workbook(LB_FILE, read_only=True, data_only=True)
ws = wb[LB_SHEET]

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < LB_HEADER: continue
    if row[1] is None or not isinstance(row[1], (int, float)): continue
    iso = str(row[8] or '').strip()
    if not iso: continue
    sys_col = str(row[6] or '').strip().upper()

    lb_all_isos.add(iso)
    if 'HP STEAM' in sys_col or 'LP STEAM' in sys_col:
        lb_hp_lp_isos.add(iso)

    # BOM 키 (HP/LP 여부와 무관하게 수집)
    line = str(row[9] or '').strip().replace('\r', '').replace('\n', '')
    item = str(row[14] or '').strip()
    et   = str(row[15] or '').strip()
    matl1 = str(row[16] or '').strip()
    matl2 = str(row[17] or '').strip()
    size  = str(row[18] or '').strip()
    thick = str(row[19] or '').strip()
    uom_raw = str(row[LB_UOM_COL] or '').strip() if len(row) > LB_UOM_COL else ''
    if uom_raw == 'MM': uom_raw = 'M'
    if not item: continue

    full_desc = build_full_description(item, matl1, matl2, size, thick, et)
    if iso not in lb_keys_by_iso:
        lb_keys_by_iso[iso] = set()
    lb_keys_by_iso[iso].add((line, full_desc, uom_raw, size))

wb.close()
print(f"  LB 전체 ISO: {len(lb_all_isos)}")
print(f"  LB HP/LP Steam ISO: {len(lb_hp_lp_isos)}")

# ─── 2. SB BOM 스캔: 전체 ISO 및 BOM 키 수집 ────────────────────────────

print("\n=== SB BOM 스캔 중... ===")
SB_FILE = "Raw File/SB BOM(20260128) (002).xlsx"
SB_SHEET = "MERGED_Piping&Fitting"
SB_HEADER = 1
SB_UOM_COL = 24
SB_QTY_COL = 25

sb_all_isos = set()
sb_keys_by_iso = {}

wb = openpyxl.load_workbook(SB_FILE, read_only=True, data_only=True)
ws = wb[SB_SHEET]

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < SB_HEADER: continue
    if row[1] is None or not isinstance(row[1], (int, float)): continue
    iso = str(row[8] or '').strip()
    if not iso: continue
    sb_all_isos.add(iso)

    line = str(row[9] or '').strip().replace('\r', '').replace('\n', '')
    item = str(row[14] or '').strip()
    et   = str(row[15] or '').strip()
    matl1 = str(row[16] or '').strip()
    matl2 = str(row[17] or '').strip()
    size  = str(row[18] or '').strip()
    thick = str(row[19] or '').strip()
    uom_raw = str(row[SB_UOM_COL] or '').strip() if len(row) > SB_UOM_COL else ''
    if uom_raw == 'MM': uom_raw = 'M'
    if not item: continue

    full_desc = build_full_description(item, matl1, matl2, size, thick, et)
    if iso not in sb_keys_by_iso:
        sb_keys_by_iso[iso] = set()
    sb_keys_by_iso[iso].add((line, full_desc, uom_raw))

wb.close()
print(f"  SB 전체 ISO: {len(sb_all_isos)}")

# ─── 3. 삭제 대상 분류 ───────────────────────────────────────────────────

lb_only_hp_lp = lb_hp_lp_isos - sb_all_isos   # SB에 없는 LB HP/LP ISO
both_hp_lp    = lb_hp_lp_isos & sb_all_isos   # SB에도 있는 LB HP/LP ISO

print(f"\n=== 삭제 대상 분류 ===")
print(f"  [A] LB 전용 HP/LP ISO (SB BOM에 없음): {len(lb_only_hp_lp)}개 -> 전체 삭제")
print(f"  [B] LB+SB 공통 HP/LP ISO: {len(both_hp_lp)}개 -> LB 전용 항목만 삭제")

# [B] 공통 ISO에서 LB 전용 키 추출 (대구경 DN65+ 만 삭제 대상)
lb_unique_keys = []
lb_unique_sb_excluded = []  # 소구경 → 삭제 제외
for iso in sorted(both_hp_lp):
    lb_k = lb_keys_by_iso.get(iso, set())
    sb_k = sb_keys_by_iso.get(iso, set())
    for line, full_desc, uom, size in sorted(lb_k):
        if (line, full_desc, uom) not in sb_k:
            if is_large_bore(size):
                lb_unique_keys.append((iso, line, full_desc, uom))
            else:
                lb_unique_sb_excluded.append((iso, line, full_desc, uom, size))

print(f"  [B] LB 전용 항목 (대구경, 삭제): {len(lb_unique_keys)}건")
print(f"  [B] LB 전용 항목 (소구경, 보존): {len(lb_unique_sb_excluded)}건")

# 소구경 보존 항목 샘플
if lb_unique_sb_excluded:
    print("\n  [B] 소구경 보존 항목 샘플 (최대 5건):")
    for iso, line, fd, uom, sz in lb_unique_sb_excluded[:5]:
        print(f"    ISO={iso[:40]}, SIZE={sz}, DESC={fd[:40]}")

# ─── 4. SQL 생성 ─────────────────────────────────────────────────────────

sql_lines = [
    "-- Large Bore HP/LP Steam 자재 삭제 SQL",
    "-- 기준: Excel LB BOM의 SYSTEM 컬럼이 HP STEAM SYSTEM 또는 LP STEAM SYSTEM",
    "-- Small Bore BOM 항목은 보존 (Spool 납품 아님)",
    "",
]

# [A] LB 전용 ISO 전체 삭제
sql_lines.append("-- ────────────────────────────────────────────────────────────────")
sql_lines.append(f"-- [A] LB 전용 HP/LP Steam ISO {len(lb_only_hp_lp)}개 전체 삭제")
sql_lines.append("-- ────────────────────────────────────────────────────────────────")
if lb_only_hp_lp:
    iso_list = ",\n  ".join(f"'{esc(iso)}'" for iso in sorted(lb_only_hp_lp))
    sql_lines.append(f"DELETE FROM material.bom\nWHERE iso_dwg_no IN (\n  {iso_list}\n);")
else:
    sql_lines.append("-- (해당 없음)")
sql_lines.append("")

# [B] LB 전용 항목 삭제 (temp table 활용)
sql_lines.append("-- ────────────────────────────────────────────────────────────────")
sql_lines.append(f"-- [B] LB+SB 공통 ISO에서 LB 전용 항목 {len(lb_unique_keys)}건 삭제")
sql_lines.append("-- ────────────────────────────────────────────────────────────────")
if lb_unique_keys:
    sql_lines.append("CREATE TEMP TABLE _lb_unique_del (iso text, line_no text, full_description text, uom text);")
    # INSERT in batches of 500 to avoid overly long lines
    batch_size = 500
    for batch_start in range(0, len(lb_unique_keys), batch_size):
        batch = lb_unique_keys[batch_start:batch_start+batch_size]
        vals = ",\n  ".join(
            f"('{esc(iso)}', '{esc(line)}', '{esc(fd)}', '{esc(uom)}')"
            for iso, line, fd, uom in batch
        )
        sql_lines.append(f"INSERT INTO _lb_unique_del VALUES\n  {vals};")
    sql_lines.append("")
    sql_lines.append("DELETE FROM material.bom b")
    sql_lines.append("USING _lb_unique_del d")
    sql_lines.append("WHERE b.iso_dwg_no = d.iso")
    sql_lines.append("  AND b.line_no = d.line_no")
    sql_lines.append("  AND b.full_description = d.full_description")
    sql_lines.append("  AND b.uom = d.uom;")
    sql_lines.append("")
    sql_lines.append("DROP TABLE _lb_unique_del;")
else:
    sql_lines.append("-- (LB 전용 항목 없음 - SB BOM이 해당 ISO의 모든 항목을 포함)")

sql_lines.append("")
sql_lines.append("-- 삭제 후 확인")
sql_lines.append("SELECT COUNT(*) AS bom_total FROM material.bom;")

out_path = "scratch/delete_lb_hp_lp_bom.sql"
with open(out_path, "w", encoding="utf-8") as f:
    f.write('\n'.join(sql_lines))

print(f"\nSQL 저장 완료: {out_path}")
print(f"  [A] DELETE 1건 (ISO {len(lb_only_hp_lp)}개 전체)")
print(f"  [B] LB 전용 항목 {len(lb_unique_keys)}건")
