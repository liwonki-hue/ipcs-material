# Speciality BOM.xlsx(VALVELIST_BOP) -> bom 테이블 등록 (category='Speciality')
# Valve와 동일하게 mat_code는 생성하지 않고 NULL, Tag로만 매칭.
# full_description 형식: "{ITEM}, {MAT2}, {SIZE}, {RATING}, {END TYPE}" (콤마 구분, app.js에서 split으로 파싱)
import math
import sys
import io
import requests
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}', 'Content-Type': 'application/json', 'Prefer': 'return=minimal'}


def clean(v):
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    s = str(v).strip()
    return s if s else None


wb = openpyxl.load_workbook('Raw File/Speciality BOM.xlsx', data_only=True)
ws = wb['VALVELIST_BOP']
header = [c.value for c in ws[1]]
assert header == ['SYSTM', 'ISO DRAWING', 'LINE NO', 'TAG', 'ITEM', 'MAT 1', 'MAT 2', 'SIZE', 'RATING', 'END TYPE', 'QTY'], header

rows = []
errors = []
for r in range(2, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 12)]
    system, iso_dwg, line_no, tag, item, mat1, mat2, size, rating, end_type, qty = vals
    tag = clean(tag)
    if not tag:
        continue
    item = clean(item) or '-'
    mat2c = clean(mat2) or '-'
    sizec = clean(size) or '-'
    ratingc = clean(rating) or '-'
    endc = clean(end_type) or '-'
    full_desc = f'{item}, {mat2c}, {sizec}, {ratingc}, {endc}'

    rows.append({
        'mat_code': None,
        'category': 'Speciality',
        'tag': tag,
        'system': clean(system),
        'iso_dwg_no': clean(iso_dwg),
        'line_no': clean(line_no),
        'full_description': full_desc,
        'uom': 'EA',
        'qty': float(qty) if qty else 1.0,
        'mat1': clean(mat1),
        'mat2': clean(mat2),
    })

print(f'변환 완료: {len(rows)}행')

tags = [r['tag'] for r in rows]
dup = set(t for t in tags if tags.count(t) > 1)
if dup:
    print(f'중복 TAG {len(dup)}건:', list(dup)[:10])
else:
    print('Tag 중복 없음 확인 완료')

if '--dry-run' in sys.argv:
    print('\n[DRY RUN] 샘플 3행:')
    for r in rows[:3]:
        print(r)
    sys.exit(0)

# 기존 Speciality BOM 삭제 (재실행 대비)
del_r = requests.delete(f'{URL}/rest/v1/bom', headers=H, params={'category': 'eq.Speciality'})
print('기존 Speciality BOM 삭제:', del_r.status_code)

BATCH = 500
ok, fail = 0, 0
for i in range(0, len(rows), BATCH):
    chunk = rows[i:i + BATCH]
    resp = requests.post(f'{URL}/rest/v1/bom', headers=H, json=chunk)
    if resp.status_code in (200, 201):
        ok += len(chunk)
    else:
        fail += len(chunk)
        print(f'  실패 batch {i}: {resp.status_code} {resp.text[:300]}')

print(f'\n삽입 완료: ok={ok} fail={fail}')

cnt_r = requests.get(f'{URL}/rest/v1/bom', headers={**H, 'Prefer': 'count=exact'},
                      params={'select': 'tag', 'category': 'eq.Speciality', 'limit': 1})
print('DB 실제 Speciality 행 수:', cnt_r.headers.get('Content-Range'))
