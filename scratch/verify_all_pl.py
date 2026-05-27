# 확인 폴더 전체 PL 파일 일괄 검증 후 pl_db_comparison.xlsx 업데이트
import sys, io, re, requests
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

SUPABASE_URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
HEADERS = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}', 'Accept-Profile': 'material'}
COMPARE_PATH = 'scratch/pl_db_comparison.xlsx'

# 패키지별 설정
CONFIGS = {
    '0443': {
        'path':    '확인/PGU-DE-0443_BOP Piping_Field Run(확인).xlsx',
        'doc_no':  'PGU-DE-0443',
        'fdr2frd': False,
    },
    '0503': {
        'path':    '확인/PGU-DE-0503_BOP Piping_Field Run(확인).xlsx',
        'doc_no':  'PGU-DE-0503',
        'fdr2frd': False,
        'norm_pkg': lambda p: re.sub(r'-PIP-0(\d{3})$', r'-PIP-\1', p),  # 4자리→3자리
    },
    '0524': {
        'path':    '확인/PGU-DE-0524_BOP Piping_General System(확인).xlsx',
        'doc_no':  'PGU-DE-0524',
        'fdr2frd': False,
    },
    '0525': {
        'path':    '확인/PGU-DE-0525_BOP Piping_General System(확인).xlsx',
        'doc_no':  'PGU-DE-0525',
        'fdr2frd': False,
    },
    '0554': {
        'path':    '확인/PGU-DE-0554_BOP Piping_General System_batch(확인).xlsx',
        'doc_no':  'PGU-DE-0554',
        'fdr2frd': False,
    },
    '0555': {
        'path':    '확인/PGU-DE-0555_BOP Piping_General System(확인).xlsx',
        'doc_no':  'PGU-DE-0555',
        'fdr2frd': False,
    },
    '0574': {
        'path':    '확인/PGU-DE-0574_BOP Piping_General System_(확인).xlsx',
        'doc_no':  'PGU-DE-0574',
        'fdr2frd': True,   # FDR → FRD
    },
}

ITEM_PREFIX = {
    'PIPE':          ['PIS', 'PIW', 'PIN'],
    'SWAGED CON':    ['SWC', 'SCN'],   # NIPPLE보다 먼저 (SWAGED CON NIPPLE 오탐 방지)
    'SWAGE-CON':     ['SWC', 'SCN'],
    'SWAGE-ECC':     ['SWE'],
    'SWAGE':         ['SWC', 'SWE', 'SCN'],
    'NIPPLE':        ['PIN', 'PIS'],
    'ELBOW LR 90':   ['EL9L'],
    'ELBOW SR 90':   ['EL9S'],
    'ELBOW LR 45':   ['EL4L'],
    'ELBOW SR 45':   ['EL4L'],
    'ELBOW 90':      ['EL9L', 'EL9S'],
    'ELBOW 45':      ['EL4L'],
    'ELBOW':         ['EL9L', 'EL9S', 'EL4L'],
    'TEE-RED':       ['TER'],
    'TEE':           ['TEE', 'TER'],
    'REDUCER-CON':   ['RDC'],
    'REDUCER-ECC':   ['RDE'],
    'REDUCER':       ['RDC', 'RDE'],
    'CAP':           ['CAP'],
    'FLANGE-BLIND':  ['FLB'],
    'FLANGE':        ['FLN', 'FLA', 'FLB', 'FLS'],
    'COUPLING-FULL': ['CPF'],
    'COUPLING-HALF': ['CPH'],
    'COUPLING':      ['CPF', 'CPH'],
    'WELDOLET':      ['WOL'],
    'SOCKOLET':      ['SOL'],
    'STUD BOLT':     ['STB'],
    'GASKET':        ['GSKT'],
    'NUT':           ['NUT'],
    'NOZZLE':        ['NOZ'],
}

FILLS = {
    'OK':       PatternFill('solid', fgColor='C6EFCE'),
    'DB↑':      PatternFill('solid', fgColor='FFEB9C'),
    'DB↓':      PatternFill('solid', fgColor='FFC7CE'),
    'DB없음':   PatternFill('solid', fgColor='FFC7CE'),
    'PL없음':   PatternFill('solid', fgColor='DDEBF7'),
    'NULL':     PatternFill('solid', fgColor='FFC7CE'),
    'MISMATCH': PatternFill('solid', fgColor='FF9900'),
}

def check_matcode(desc, mat_code):
    if not mat_code:
        return 'NULL'
    desc_u = desc.upper()
    for kw, prefixes in ITEM_PREFIX.items():
        if kw in desc_u:
            if any(mat_code.startswith(p) for p in prefixes):
                return 'OK'
            return 'MISMATCH'
    return 'OK'

def norm_desc(s):
    return re.sub(r'\s+', ' ', str(s).upper().strip())

def norm_unit(u):
    if not u: return ''
    s = str(u).strip().upper()
    if 'EA' in s: return 'EA'
    if s == 'M':  return 'M'
    return s

def is_pipe(desc):
    d = str(desc).upper().strip()
    return d.startswith('PIPE') and 'NIPPLE' not in d

def row_fill(status):
    s = str(status)
    if s == 'OK':        return FILLS['OK']
    if 'DB↑' in s:      return FILLS['DB↑']
    if 'DB↓' in s:      return FILLS['DB↓']
    if 'NULL' in s:      return FILLS['NULL']
    if 'MISMATCH' in s:  return FILLS['MISMATCH']
    if s == 'DB없음':   return FILLS['DB없음']
    if s == 'PL없음':   return FILLS['PL없음']
    return None

def fetch_db(doc_no):
    all_rows, offset = [], 0
    while True:
        r = requests.get(
            f'{SUPABASE_URL}/rest/v1/receiving?select=id,pkg_no,mat_code,qty,unit,full_description'
            f'&doc_no=eq.{doc_no}&order=pkg_no,id&limit=1000&offset={offset}',
            headers=HEADERS
        )
        data = r.json()
        if not data: break
        all_rows.extend(data)
        offset += 1000
        if len(data) < 1000: break
    return all_rows

def parse_pl(path, fdr2frd, norm_pkg_fn=None):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    pl_grouped = defaultdict(lambda: {'desc': '', 'qty_pl': 0.0, 'unit_pl': 'EA'})
    for row in ws.iter_rows(min_row=2, values_only=True):
        pkg_raw = row[1]
        desc    = row[2]
        qty     = row[3]
        if not pkg_raw or not desc or not isinstance(qty, (int, float)):
            continue
        pkg_db = str(pkg_raw).strip()
        if fdr2frd:
            pkg_db = pkg_db.replace('-BOP-FDR-', '-BOP-FRD-')
        if norm_pkg_fn:
            pkg_db = norm_pkg_fn(pkg_db)
        key = (pkg_db, norm_desc(desc))
        pl_grouped[key]['desc']    = str(desc).strip()
        pl_grouped[key]['qty_pl'] += float(qty)
        pl_grouped[key]['unit_pl'] = norm_unit(row[4]) or 'EA'
    return pl_grouped

def compare(pl_grouped, db_rows):
    db_by_key = {}
    for row in db_rows:
        key = (row['pkg_no'], norm_desc(row['full_description'] or ''))
        db_by_key[key] = row

    results = []
    pl_matched, db_matched = set(), set()

    for (pkg_db, pl_desc_n), pl_info in sorted(pl_grouped.items()):
        pl_qty    = pl_info['qty_pl']
        pl_unit   = pl_info['unit_pl']
        pipe      = is_pipe(pl_info['desc'])
        pl_qty_db = pl_qty * 6.0 if pipe and pl_unit == 'EA' else pl_qty

        exact_key = (pkg_db, pl_desc_n)
        if exact_key in db_by_key:
            dr = db_by_key[exact_key]
            db_qty   = dr['qty'] or 0
            db_unit  = norm_unit(dr['unit'])
            mat_code = dr['mat_code'] or ''
            mc_status = check_matcode(pl_info['desc'], mat_code)
            qty_diff  = round(db_qty - pl_qty_db, 4)
            status    = 'OK' if abs(qty_diff) < 0.01 else ('DB↑' if qty_diff > 0 else 'DB↓')
            if mc_status != 'OK':
                status = mc_status + ('/' + status if status != 'OK' else '')
            pl_matched.add(exact_key)
            db_matched.add(exact_key)
            results.append({'pkg_no': pkg_db, 'pl_desc': pl_info['desc'],
                'pl_qty': pl_qty, 'pl_unit': pl_unit, 'pl_qty_db': pl_qty_db,
                'db_mat_code': mat_code, 'db_qty': db_qty, 'db_unit': db_unit,
                'qty_diff': qty_diff, 'mc_status': mc_status, 'status': status})
            continue

        # 단어 유사도 fallback (같은 pkg_no 내)
        pl_words = set(pl_desc_n.split())
        best_key, best_score = None, 0
        for dk in db_by_key:
            if dk in db_matched or dk[0] != pkg_db: continue
            score = len(pl_words & set(dk[1].split()))
            if score > best_score:
                best_score, best_key = score, dk

        if best_key and best_score >= 3:
            dr = db_by_key[best_key]
            db_qty   = dr['qty'] or 0
            db_unit  = norm_unit(dr['unit'])
            mat_code = dr['mat_code'] or ''
            mc_status = check_matcode(pl_info['desc'], mat_code)
            qty_diff  = round(db_qty - pl_qty_db, 4)
            status    = 'OK' if abs(qty_diff) < 0.01 else ('DB↑' if qty_diff > 0 else 'DB↓')
            if mc_status != 'OK':
                status = mc_status + ('/' + status if status != 'OK' else '')
            pl_matched.add((pkg_db, pl_desc_n))
            db_matched.add(best_key)
            results.append({'pkg_no': pkg_db, 'pl_desc': pl_info['desc'],
                'pl_qty': pl_qty, 'pl_unit': pl_unit, 'pl_qty_db': pl_qty_db,
                'db_mat_code': mat_code, 'db_qty': db_qty, 'db_unit': db_unit,
                'qty_diff': qty_diff, 'mc_status': mc_status, 'status': status})
        else:
            results.append({'pkg_no': pkg_db, 'pl_desc': pl_info['desc'],
                'pl_qty': pl_qty, 'pl_unit': pl_unit, 'pl_qty_db': pl_qty_db,
                'db_mat_code': '', 'db_qty': '', 'db_unit': '',
                'qty_diff': '', 'mc_status': '', 'status': 'DB없음'})

    for dk, dr in sorted(db_by_key.items(), key=lambda x: x[0]):
        if dk in db_matched: continue
        mat_code  = dr['mat_code'] or ''
        mc_status = check_matcode(dr['full_description'] or '', mat_code)
        results.append({'pkg_no': dr['pkg_no'], 'pl_desc': '(PL 미매칭)',
            'pl_qty': '', 'pl_unit': '', 'pl_qty_db': '',
            'db_mat_code': mat_code, 'db_qty': dr['qty'], 'db_unit': norm_unit(dr['unit']),
            'qty_diff': '', 'mc_status': mc_status, 'status': 'PL없음'})

    return results

# ── 메인 처리 ──────────────────────────────────────────────────────────
wb_cmp = openpyxl.load_workbook(COMPARE_PATH)
ws_cmp = wb_cmp['항목별 비교']

summary = {}
for code, cfg in CONFIGS.items():
    doc_no = cfg['doc_no']
    print(f'\n=== {doc_no} ===')

    pl_grouped = parse_pl(cfg['path'], cfg['fdr2frd'], cfg.get('norm_pkg'))
    print(f'  PL: {len(pl_grouped)}개 항목')

    db_rows = fetch_db(doc_no)
    print(f'  DB: {len(db_rows)}행')

    results = compare(pl_grouped, db_rows)

    ok    = sum(1 for r in results if r['status'] == 'OK')
    diff  = sum(1 for r in results if 'DB↑' in str(r['status']) or 'DB↓' in str(r['status']))
    null_ = sum(1 for r in results if 'NULL' in str(r['status']))
    mism  = sum(1 for r in results if 'MISMATCH' in str(r['status']))
    dbm   = sum(1 for r in results if r['status'] == 'DB없음')
    plm   = sum(1 for r in results if r['status'] == 'PL없음')
    summary[code] = {'OK': ok, 'diff': diff, 'null': null_, 'mismatch': mism, 'DB없음': dbm, 'PL없음': plm, 'total': len(results)}
    print(f'  결과: OK={ok}, 수량불일치={diff}, NULL={null_}, MISMATCH={mism}, DB없음={dbm}, PL없음={plm}')
    for r in results:
        if r['status'] != 'OK':
            print(f'    [{r["status"]}] {r["pkg_no"]} | {str(r["pl_desc"])[:50]} | PL_DB={r["pl_qty_db"]} / DB={r["db_qty"]} {r["db_unit"]} | mat={r["db_mat_code"]}')

    # 기존 행 삭제
    tag = code if code != '0574' else '0574'
    delete_rows = [i for i in range(2, ws_cmp.max_row + 1)
                   if ws_cmp.cell(row=i, column=1).value and doc_no in str(ws_cmp.cell(row=i, column=1).value)]
    for i in reversed(delete_rows):
        ws_cmp.delete_rows(i)
    print(f'  기존 {len(delete_rows)}행 삭제, 신규 {len(results)}행 추가')

    for r in results:
        ws_cmp.append([r['pkg_no'], r['pl_desc'], r['pl_qty'], r['pl_unit'],
                       r['pl_qty_db'], r['db_mat_code'], r['db_qty'], r['db_unit'],
                       r['qty_diff'], r['mc_status'], r['status']])
        fill = row_fill(r['status'])
        if fill:
            for cell in ws_cmp[ws_cmp.max_row]:
                cell.fill = fill

wb_cmp.save(COMPARE_PATH)
print(f'\n=== 전체 완료: {COMPARE_PATH} ===')
print(f'{"패키지":<8} {"OK":>5} {"수량불일치":>10} {"NULL":>6} {"MISMATCH":>9} {"DB없음":>7} {"PL없음":>7} {"합계":>6}')
print('-' * 65)
for code, s in summary.items():
    print(f'{code:<8} {s["OK"]:>5} {s["diff"]:>10} {s["null"]:>6} {s["mismatch"]:>9} {s["DB없음"]:>7} {s["PL없음"]:>7} {s["total"]:>6}')
