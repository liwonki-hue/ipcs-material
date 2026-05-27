# PL 파일 항목별 vs DB receiving 비교 스크립트
# 기준: PL Summary PKG NO 합계 정확, PL 파일 항목 상세 vs DB 비교
import sys, io, re, requests
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from collections import defaultdict

SUPABASE_URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
HEADERS = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}', 'Accept-Profile': 'material'}

# ── mat_code 검증 키워드 매핑 ─────────────────────────────────────────
# desc 키워드 → 기대 mat_code 접두어 목록
ITEM_PREFIX = {
    'PIPE':         ['PIS', 'PIW', 'PIN'],
    'ELBOW LR 90':  ['EL9L'],
    'ELBOW SR 90':  ['EL9S'],
    'ELBOW LR 45':  ['EL4L'],
    'ELBOW 90':     ['EL9L', 'EL9S'],
    'ELBOW 45':     ['EL4L'],
    'ELBOW':        ['EL9L', 'EL9S', 'EL4L'],
    'TEE-RED':      ['TER'],
    'TEE':          ['TEE', 'TER'],
    'REDUCER':      ['RDC', 'RDE'],
    'CAP':          ['CAP'],
    'FLANGE-BLIND': ['FLB'],
    'FLANGE':       ['FLN', 'FLA', 'FLB', 'FLS'],
    'COUPLING-HALF':['CPH'],
    'COUPLING-FULL':['CPF'],
    'COUPLING':     ['CPF', 'CPH'],
    'WELDOLET':     ['WOL'],
    'SOCKOLET':     ['SOL'],
    'STUD BOLT':    ['STB'],
    'GASKET':       ['GSKT'],
    'NUT':          ['NUT'],
    'NIPPLE':       ['PIN'],
    'SWAGE':        ['SCN'],
    'NOZZLE':       ['NOZ'],
}

def check_matcode(desc, mat_code):
    """mat_code가 description과 일치하는지 검증. 반환: 'OK'/'NULL'/'MISMATCH'"""
    if not mat_code:
        return 'NULL'
    desc_u = desc.upper()
    for kw, prefixes in ITEM_PREFIX.items():
        if kw in desc_u:
            if any(mat_code.startswith(p) for p in prefixes):
                return 'OK'
            return 'MISMATCH'
    return 'OK'  # 키워드 없으면 미검증 → OK

def norm_unit(u):
    if not u:
        return ''
    s = str(u).strip().upper()
    if 'EA' in s or 'ШТ' in s:
        return 'EA'
    if s == 'M':
        return 'M'
    return s

def is_pipe_desc(desc):
    d = str(desc).upper()
    return d.startswith('PIPE')

# ── PL 파일 파서 ─────────────────────────────────────────────────────
PL_FILES = {
    '0443': 'Raw File/Packing List/PGU-DE-0443_BOP Piping_Field Run(1차분)_CLPL(Rev.1).xlsx',
    '0503': 'Raw File/Packing List/PGU-DE-0503_BOP Piping_Field Run_CIPL(Rev.4).xlsx',
    '0524': 'Raw File/Packing List/PGU-DE-0524_BOP Piping_General System_batch 3_CIPL(Rev.4).xlsx',
    '0525': 'Raw File/Packing List/PGU-DE-0525_BOP Piping_General System_batch 4_CIPL(Rev.2).xlsx',
    '0554': 'Raw File/Packing List/PGU-DE-0554_BOP Piping_General System_batch 5_CIPL(Rev.2).xlsx',
    '0555': 'Raw File/Packing List/PGU-DE-0555_BOP Piping_General System_batch 6_CIPL(Rev.2).xlsx',
    '0565': 'Raw File/Packing List/PGU-DE-0565_BOP Piping_General System_batch 7_CIPL(Rev.3).xlsx',
    '0574': 'Raw File/Packing List/PGU-DE-0574_BOP Piping_General System_batch 8_CIPL(Rev.3).xlsx',
}
# 헤더 행 번호 (1-based)
HEADER_ROW = {'0443': 10, '0555': 10}  # 나머지는 12

def parse_pl_0503(path):
    """0503 전용 파서: qty=col3(D), pkg=col1(B), desc=col2(C)"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Detail PL']
    items = defaultdict(list)  # pkg_no → [{desc, qty, unit}]
    current_pkg = None
    for row in ws.iter_rows(min_row=14, values_only=True):
        pkg  = row[1]
        desc = row[2]
        qty  = row[3]
        # 수식 문자열 처리 (="PGU-DE-0503-BOP-PIP-00"&"X" 형태)
        if pkg and isinstance(pkg, str) and pkg.startswith('="'):
            m = re.findall(r'"([^"]*)"', pkg)
            pkg = ''.join(m) if m else None
        if pkg and isinstance(pkg, str) and 'BOP-PIP-' in pkg:
            current_pkg = pkg.strip()
        if current_pkg and desc and isinstance(qty, (int, float)) and qty > 0:
            unit_out = 'M' if is_pipe_desc(desc) else 'EA'
            # PL qty는 EA(파이프 개수), DB는 M → 변환 표시용
            items[current_pkg].append({'desc': str(desc).strip(), 'qty_pl': float(qty),
                                        'unit_pl': 'EA', 'unit_db': unit_out,
                                        'pipe_len': 6.0 if is_pipe_desc(desc) else 1.0})
    return items

def parse_pl_standard(path, pkg_code):
    """표준 파서: pkg=col1, desc=col2, qty=col4, unit=col5"""
    header_row = HEADER_ROW.get(pkg_code, 12)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Detail PL']
    items = defaultdict(list)
    current_pkg = None
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        pkg  = str(row[1]).strip() if row[1] else None
        desc = row[2]
        qty  = row[4]
        unit = norm_unit(row[5])
        if pkg and 'BOP-' in pkg and not pkg.startswith('='):
            current_pkg = pkg
        if current_pkg and desc and isinstance(qty, (int, float)) and qty > 0:
            if not unit:  # unit 없으면 EA 가정
                unit = 'EA'
            # 파이프이고 EA 단위면 DB는 M (×6)
            if is_pipe_desc(desc) and unit == 'EA':
                items[current_pkg].append({'desc': str(desc).strip(), 'qty_pl': float(qty),
                                            'unit_pl': 'EA', 'unit_db': 'M', 'pipe_len': 6.0})
            else:
                items[current_pkg].append({'desc': str(desc).strip(), 'qty_pl': float(qty),
                                            'unit_pl': unit, 'unit_db': unit, 'pipe_len': 1.0})
    return items

# ── DB 조회 ───────────────────────────────────────────────────────────
print('DB receiving 전체 조회...')
all_rows, offset = [], 0
while True:
    r = requests.get(f'{SUPABASE_URL}/rest/v1/receiving?select=id,doc_no,pkg_no,mat_code,qty,unit,full_description,category&limit=1000&offset={offset}', headers=HEADERS)
    data = r.json()
    if not data: break
    all_rows.extend(data)
    offset += 1000
    if len(data) < 1000: break

# DB dict: pkg_no → list of rows
db_by_pkg = defaultdict(list)
for row in all_rows:
    db_by_pkg[row['pkg_no']].append(row)
print(f'  → {len(all_rows)}행, {len(db_by_pkg)} pkg_no\n')

# ── PL 파일 파싱 ──────────────────────────────────────────────────────
print('PL 파일 파싱...')
pl_all = {}  # pkg_no → items
for code, path in PL_FILES.items():
    if code == '0503':
        items = parse_pl_0503(path)
    else:
        items = parse_pl_standard(path, code)
    pl_all.update(items)
    print(f'  0{code}: {len(items)} pkg_no, {sum(len(v) for v in items.values())} items')

# ── 비교 ──────────────────────────────────────────────────────────────
results = []  # 최종 출력 행

def norm_desc(s):
    return re.sub(r'\s+', ' ', str(s).upper().strip())

for pkg_no in sorted(db_by_pkg.keys()):
    db_rows = db_by_pkg[pkg_no]
    pl_items = pl_all.get(pkg_no, [])

    # DB 항목: 설명 정규화 → qty/unit
    db_matched = [False] * len(db_rows)
    pl_matched = [False] * len(pl_items)

    for di, dr in enumerate(db_rows):
        db_desc = norm_desc(dr['full_description'] or '')
        db_qty  = dr['qty'] or 0
        db_unit = norm_unit(dr['unit'])
        mat_code = dr['mat_code'] or ''
        mc_status = check_matcode(dr['full_description'] or '', mat_code)

        best_pi, best_score = -1, 0
        for pi, pl in enumerate(pl_items):
            if pl_matched[pi]: continue
            pl_desc_n = norm_desc(pl['desc'])
            # 공통 단어 수로 유사도 측정
            db_words = set(db_desc.split())
            pl_words = set(pl_desc_n.split())
            score = len(db_words & pl_words)
            if score > best_score:
                best_score, best_pi = score, pi

        if best_pi >= 0 and best_score >= 2:
            pl = pl_items[best_pi]
            pl_qty_db = pl['qty_pl'] * pl['pipe_len']  # DB 단위로 변환
            qty_diff  = round(db_qty - pl_qty_db, 4)
            status    = 'OK' if abs(qty_diff) < 0.01 else ('DB↑' if qty_diff > 0 else 'DB↓')
            if mc_status != 'OK': status = mc_status + '/' + status if status != 'OK' else mc_status
            pl_matched[best_pi] = True
            db_matched[di] = True
            results.append({
                'pkg_no': pkg_no, 'pl_desc': pl['desc'],
                'pl_qty': pl['qty_pl'], 'pl_unit': pl['unit_pl'],
                'pl_qty_db': pl_qty_db, 'db_mat_code': mat_code,
                'db_qty': db_qty, 'db_unit': db_unit,
                'qty_diff': qty_diff, 'mc_status': mc_status, 'status': status
            })
        else:
            results.append({
                'pkg_no': pkg_no, 'pl_desc': '(PL 미매칭)',
                'pl_qty': '', 'pl_unit': '',
                'pl_qty_db': '', 'db_mat_code': mat_code,
                'db_qty': db_qty, 'db_unit': db_unit,
                'qty_diff': '', 'mc_status': mc_status,
                'status': 'PL없음' if not pl_items else ('MAT_NULL' if not mat_code else mc_status)
            })

    # PL에만 있고 DB에 없는 항목
    for pi, pl in enumerate(pl_items):
        if not pl_matched[pi]:
            results.append({
                'pkg_no': pkg_no, 'pl_desc': pl['desc'],
                'pl_qty': pl['qty_pl'], 'pl_unit': pl['unit_pl'],
                'pl_qty_db': pl['qty_pl'] * pl['pipe_len'],
                'db_mat_code': '', 'db_qty': '', 'db_unit': '',
                'qty_diff': '', 'mc_status': '', 'status': 'DB없음'
            })

# ── Excel 출력 ────────────────────────────────────────────────────────
OUT = 'scratch/pl_db_comparison.xlsx'
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '항목별 비교'

FILLS = {
    'OK':       PatternFill('solid', fgColor='C6EFCE'),
    'DB↑':      PatternFill('solid', fgColor='FFEB9C'),
    'DB↓':      PatternFill('solid', fgColor='FFC7CE'),
    'DB없음':   PatternFill('solid', fgColor='FFC7CE'),
    'PL없음':   PatternFill('solid', fgColor='DDEBF7'),
    'NULL':     PatternFill('solid', fgColor='FFC7CE'),
    'MISMATCH': PatternFill('solid', fgColor='FF9900'),
}

headers = ['PKG NO', 'PL 설명', 'PL 수량(EA)', 'PL 단위',
           'PL→DB 환산수량', 'DB mat_code', 'DB 수량', 'DB 단위',
           '수량 차이', 'MatCode 상태', '결과']
ws_out.append(headers)
for cell in ws_out[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill('solid', fgColor='4472C4')
    cell.font = Font(bold=True, color='FFFFFF')

ok_cnt = diff_cnt = null_cnt = mismatch_cnt = db_miss_cnt = 0
for r in results:
    row = [r['pkg_no'], r['pl_desc'],
           r['pl_qty'], r['pl_unit'],
           r['pl_qty_db'], r['db_mat_code'],
           r['db_qty'], r['db_unit'],
           r['qty_diff'], r['mc_status'], r['status']]
    ws_out.append(row)
    status = r['status']
    fill = None
    if status == 'OK': ok_cnt += 1; fill = FILLS['OK']
    elif status in ('DB↑', 'DB↓'): diff_cnt += 1; fill = FILLS.get(status)
    elif 'NULL' in str(status): null_cnt += 1; fill = FILLS['NULL']
    elif 'MISMATCH' in str(status): mismatch_cnt += 1; fill = FILLS['MISMATCH']
    elif status == 'DB없음': db_miss_cnt += 1; fill = FILLS['DB없음']
    else: fill = FILLS.get(status)
    if fill:
        for cell in ws_out[ws_out.max_row]:
            cell.fill = fill

# 컬럼 폭 조정
ws_out.column_dimensions['A'].width = 38
ws_out.column_dimensions['B'].width = 45
ws_out.column_dimensions['F'].width = 28
for col in ['C','D','E','G','H','I','J','K']:
    ws_out.column_dimensions[col].width = 14

# 요약 시트
ws_sum = wb_out.create_sheet('요약')
ws_sum.append(['구분', '건수'])
ws_sum.append(['OK (일치)', ok_cnt])
ws_sum.append(['수량 불일치 (DB↑/DB↓)', diff_cnt])
ws_sum.append(['MatCode NULL', null_cnt])
ws_sum.append(['MatCode MISMATCH', mismatch_cnt])
ws_sum.append(['DB없음 (PL항목 미등록)', db_miss_cnt])
ws_sum.append(['합계', len(results)])

wb_out.save(OUT)
print(f'\n출력 완료: {OUT}')
print(f'OK={ok_cnt}, 수량불일치={diff_cnt}, MatCode NULL={null_cnt}, MISMATCH={mismatch_cnt}, DB없음={db_miss_cnt}')
