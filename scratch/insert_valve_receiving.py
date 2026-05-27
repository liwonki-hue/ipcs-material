# Valve Packing List → material.receiving INSERT SQL 생성
# Tag No 기준으로 BOM에서 mat_code 조회 후 receiving 등록
# 실행: python scratch/insert_valve_receiving.py

import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

PL_PATH  = 'Raw File/Packing List (편집)/PGU-DE-0510_BOP Piping_Manual Valve(확인).xlsx'
BOM_PATH = 'Raw File/BOM Data/Valve BOM.xlsx'
OUT_PATH  = 'scratch/insert_valve_receiving.sql'
DOC_NO    = 'PGU-DE-0510'
START_ID  = 1010   # 기존 최대 id(1009) 다음

# ── BOM에서 Tag → mat_code / full_description 매핑 ──────────────────────────
bom_wb = openpyxl.load_workbook(BOM_PATH)
bom_ws = bom_wb.active

tag_map = {}   # tag → {mat_code, desc}
for r in range(2, bom_ws.max_row + 1):
    mc  = bom_ws.cell(r, 1).value
    tag = bom_ws.cell(r, 7).value
    desc = bom_ws.cell(r, 6).value
    if tag and mc:
        t = str(tag).strip()
        tag_map[t] = {
            'mat_code': str(mc).strip(),
            'desc':     str(desc).strip() if desc else '',
        }

# ── PL 읽기 ──────────────────────────────────────────────────────────────────
pl_wb = openpyxl.load_workbook(PL_PATH)
pl_ws = pl_wb.active

def esc(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

rows = []
pkg_no = None
for r in range(2, pl_ws.max_row + 1):
    pkg = pl_ws.cell(r, 2).value
    desc_pl = pl_ws.cell(r, 3).value
    qty  = pl_ws.cell(r, 4).value
    tag  = pl_ws.cell(r, 6).value

    if not tag:
        continue

    # Package No 이월 (빈 셀이면 직전 값 유지)
    if pkg:
        pkg_no = str(pkg).strip()

    tag_str = str(tag).strip()
    pl_desc = str(desc_pl).strip() if desc_pl else ''
    rec_qty = float(qty) if qty else 1.0

    bom = tag_map.get(tag_str)
    if bom:
        mat_code = bom['mat_code']
        full_desc = bom['desc']   # BOM 기준 Description 사용
    else:
        mat_code  = None
        full_desc = pl_desc
        print(f'[WARNING] Tag 미매칭: {tag_str}')

    rows.append({
        'doc_no':   DOC_NO,
        'pkg_no':   pkg_no,
        'mat_code': mat_code,
        'unit':     'EA',
        'qty':      rec_qty,
        'category': 'Valve',
        'desc':     full_desc,
        'tag':      tag_str,
    })

print(f'처리 대상: {len(rows)}행')

# ── SQL 생성 ──────────────────────────────────────────────────────────────────
with open(OUT_PATH, 'w', encoding='utf-8') as f:
    f.write('-- Valve Receiving INSERT (PGU-DE-0510 Manual Valve PL)\n')
    f.write('-- Tag No 기준 BOM mat_code 매칭, Supabase SQL Editor에서 실행\n\n')
    for idx, row in enumerate(rows):
        row_id = START_ID + idx
        mc_sql = esc(row['mat_code'])
        f.write(
            f"INSERT INTO material.receiving "
            f"(id, doc_no, pkg_no, mat_code, unit, qty, category, full_description, tag) VALUES "
            f"({row_id}, {esc(row['doc_no'])}, {esc(row['pkg_no'])}, {mc_sql}, "
            f"{esc(row['unit'])}, {row['qty']}, {esc(row['category'])}, "
            f"{esc(row['desc'])}, {esc(row['tag'])});\n"
        )

print(f'→ SQL 저장: {OUT_PATH}')
print('\n등록 내용:')
for r in rows:
    print(f"  {r['tag']:20s} | pkg={r['pkg_no']} | mc={r['mat_code']}")
