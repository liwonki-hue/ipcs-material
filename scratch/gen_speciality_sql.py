# Speciality BOM DB 등록 SQL 생성 (matcode_master 4건 + bom 전체 Speciality 505건)
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BOM_PATH = 'Raw File/BOM Data/Speciality BOM.xlsx'
OUT_MC   = 'scratch/insert_speciality_matcodes.sql'
OUT_BOM  = 'scratch/insert_speciality_bom.sql'

def sq(v):
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

# ── PART 1: matcode_master 4건 ────────────────────────────────
MATCODE_ROWS = [
    ('STP-A106-1-3000-SW', 'Speciality', 'STEAM TRAP', 'ASTM A106 GR.B',   '1"', None, '3000#', 'SW'),
    ('STP-A335-1-3000-SW', 'Speciality', 'STEAM TRAP', 'ASTM A335 GR.P91', '1"', None, '3000#', 'SW'),
    ('ATP-A312-1-3000-SW', 'Speciality', 'AIR TRAP',   'ASTM A312 TP304',  '1"', None, '3000#', 'SW'),
    ('ATP-A106-1-3000-SW', 'Speciality', 'AIR TRAP',   'ASTM A106 GR.B',   '1"', None, '3000#', 'SW'),
]

mc_lines = [
    "-- Speciality matcode_master 등록 (STEAM TRAP 2건 + AIR TRAP 2건)",
    "-- Supabase SQL Editor에서 실행",
    "",
]
for row in MATCODE_ROWS:
    vals = ', '.join(sq(v) for v in row)
    mc_lines.append(
        f"INSERT INTO material.matcode_master"
        f" (mat_code, category, item_desc, matl_desc, size1, size2, class_desc, et_desc)"
        f" VALUES ({vals})"
        f" ON CONFLICT (mat_code) DO NOTHING;"
    )

with open(OUT_MC, 'w', encoding='utf-8') as f:
    f.write('\n'.join(mc_lines))
print(f'[PART 1] matcode_master SQL: {OUT_MC}  ({len(MATCODE_ROWS)}건)')

# ── PART 2: bom 전체 Speciality 505건 ────────────────────────
wb = openpyxl.load_workbook(BOM_PATH, data_only=True)
ws = wb.active

bom_lines = [
    "-- Speciality BOM 전체 등록 (505건)",
    "-- 기존 Speciality 행 전체 삭제 후 재삽입",
    "DELETE FROM material.bom WHERE category = 'Speciality';",
    "",
]

cnt_by_cat = {}
total = 0
for r in range(2, ws.max_row + 1):
    tag  = ws.cell(r, 3).value
    if not tag:
        continue

    mc   = ws.cell(r, 2).value
    sys_ = ws.cell(r, 4).value
    iso  = ws.cell(r, 5).value
    lno  = ws.cell(r, 6).value
    desc = ws.cell(r, 7).value
    uom  = ws.cell(r, 13).value
    qty  = ws.cell(r, 14).value

    tag_s = str(tag).strip()
    vals = ', '.join([
        sq(mc), sq('Speciality'), sq(tag_s),
        sq(sys_), sq(iso), sq(lno), sq(desc),
        sq(uom or 'EA'), sq(qty or 1)
    ])
    bom_lines.append(
        f"INSERT INTO material.bom"
        f" (mat_code, category, tag, system, iso_dwg_no, line_no, full_description, uom, qty)"
        f" VALUES ({vals});"
    )

    # 카테고리별 카운트 (tag prefix 기준)
    parts = tag_s.split('-')
    item_code = parts[1] if len(parts) > 1 else 'UNK'
    cnt_by_cat[item_code] = cnt_by_cat.get(item_code, 0) + 1
    total += 1

with open(OUT_BOM, 'w', encoding='utf-8') as f:
    f.write('\n'.join(bom_lines))

print(f'[PART 2] bom SQL: {OUT_BOM}  (총 {total}건)')
for code, cnt in sorted(cnt_by_cat.items()):
    print(f'  {code}: {cnt}건')
