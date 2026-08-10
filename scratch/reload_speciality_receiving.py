# Speciality (Receiving)_Format_Template.xlsx(Speciality/Untagged Items) 기준 receiving category='Speciality' 전체 재적재
import openpyxl
import psycopg2
import psycopg2.extras

DB_URL = open('.env').read().strip().split('=', 1)[1]
XLSX_PATH = 'Raw File/Speciality (Receiving)_Format_Template.xlsx'

# B0-FJ-36004/B0-FJ-36005(PGU-DE-0516) — 같은 Tag가 10"/8" 두 사이즈로 중복 등록,
# BOM(8", 150#, FF)과 일치하는 8" 행만 반영하기로 사용자 확인, 10" 행은 제외
DUP_TAG_PKG = 'PGU-DE-0516'
DUP_TAGS = {'B0-FJ-36004', 'B0-FJ-36005'}
DUP_DROP_SIZE = '10"'

conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# ============ Speciality 시트 ============
ws = wb['Speciality']
sp_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
print(f'Speciality 시트: {len(sp_rows)}행')

out_rows = []
dropped = 0

for r in sp_rows:
    pkg, pkg_no, tag, item, mat1, mat2, size, rating, conn_type, unit, qty, _ref = r
    tag = tag.strip() if isinstance(tag, str) else tag
    item = (item or '').strip() or None

    if pkg == DUP_TAG_PKG and tag in DUP_TAGS and (size or '').strip() == DUP_DROP_SIZE:
        dropped += 1
        continue

    full_desc = f'{item or "-"}, {mat2 or "-"}, {size or "-"}, {rating or "-"}, {conn_type or "-"}'

    out_rows.append({
        'doc_no': pkg, 'pkg_no': pkg_no, 'tag': tag, 'parent_tag': tag,
        'op_type': None, 'valve_type': item,
        'mat1': mat1, 'mat2': mat2, 'size': size, 'rating': rating,
        'full_description': full_desc, 'unit': unit or 'EA', 'qty': qty or 1,
    })

print(f'중복 태그(10" 행) 제외: {dropped}건')
print(f'Speciality 반영 대상: {len(out_rows)}행')

# ============ Untagged Items 시트 ============
ws2 = wb['Untagged Items']
untagged_rows = [r for r in ws2.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r[:8])]
print(f'Untagged Items 시트: {len(untagged_rows)}행')

acc_seq = {}  # pkg_no -> counter (Tag 유니크화용)

for r in untagged_rows:
    pkg, pkg_no, ref_tag, seq_no, item, desc, unit, qty = r[:8]
    item = (item or '').strip() or None
    ref_tag = (ref_tag or '').strip() or None

    n = acc_seq.get(pkg_no, 0) + 1
    acc_seq[pkg_no] = n
    infix = 'SPARE' if item and item.upper() == 'SPARE' else 'ACC'
    tag = f'{pkg_no}-{infix}-{n:03d}'
    parent_tag = ref_tag or pkg_no

    out_rows.append({
        'doc_no': pkg, 'pkg_no': pkg_no, 'tag': tag, 'parent_tag': parent_tag,
        'op_type': None, 'valve_type': item,
        'mat1': None, 'mat2': None, 'size': None, 'rating': None,
        'full_description': desc, 'unit': unit or 'EA', 'qty': qty or 1,
    })

print(f'총 적재 대상 행: {len(out_rows)}')

# --- Tag 유니크성 최종 검증 ---
tags = [r['tag'] for r in out_rows]
assert len(tags) == len(set(tags)), f'Tag 중복 발견: {len(tags) - len(set(tags))}건'
print('Tag 100% 유니크 확인 완료')

# ============ DB 반영 ============
cur.execute("SELECT COUNT(*) FROM receiving WHERE category='Speciality'")
before = cur.fetchone()[0]
print(f'기존 receiving category=Speciality 행 수: {before}')

cur.execute("DELETE FROM receiving WHERE category='Speciality'")
print(f'삭제된 행: {cur.rowcount}')

cur.execute("SELECT COALESCE(MAX(id), 0) FROM receiving")
next_id = cur.fetchone()[0] + 1

insert_sql = """
    INSERT INTO receiving (id, doc_no, pkg_no, category, tag, parent_tag, op_type, valve_type,
                            mat1, mat2, size, rating, full_description, unit, qty)
    VALUES %s
"""
values = []
for r in out_rows:
    values.append((
        next_id, r['doc_no'], r['pkg_no'], 'Speciality', r['tag'], r['parent_tag'],
        r['op_type'], r['valve_type'], r['mat1'], r['mat2'], r['size'], r['rating'],
        r['full_description'], r['unit'], r['qty']
    ))
    next_id += 1

psycopg2.extras.execute_values(cur, insert_sql, values)
conn.commit()
print(f'삽입된 행: {len(values)}')

cur.execute("SELECT COUNT(*), COUNT(DISTINCT tag) FROM receiving WHERE category='Speciality'")
total, distinct = cur.fetchone()
print(f'최종 확인: 총 {total}행, Tag 유니크 {distinct}개')

cur.close()
conn.close()
