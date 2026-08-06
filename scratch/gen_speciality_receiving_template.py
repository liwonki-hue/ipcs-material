# Speciality (Receiving) 분류 작업용 Excel 템플릿 생성 — Valve (Receiving) 포맷과 동일 구조
# (Speciality/Untagged Items 2시트), 기존 receiving 데이터로 사전 채움
import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

DB_URL = open('.env').read().strip().split('=', 1)[1]
conn = psycopg2.connect(DB_URL)
cur = conn.cursor()

cur.execute("""
    SELECT doc_no, pkg_no, tag, unit, qty, full_description
    FROM receiving WHERE category='Speciality' AND parent_tag IS NULL
    ORDER BY doc_no, pkg_no, tag
""")
tagged = cur.fetchall()

cur.execute("""
    SELECT doc_no, pkg_no, tag, parent_tag, unit, qty, full_description
    FROM receiving WHERE category='Speciality' AND parent_tag IS NOT NULL
    ORDER BY doc_no, pkg_no, parent_tag, tag
""")
untagged = cur.fetchall()
cur.close()
conn.close()

print(f'Tagged(Speciality): {len(tagged)}, Untagged Items: {len(untagged)}')

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill('solid', fgColor='0A2540')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'


# ── Sheet 1: Speciality (Tag 있는 관리 대상) ──────────────────────────
ws1 = wb.active
ws1.title = 'Speciality'
headers1 = ['PKG', 'PKG NO', 'TAG NO', 'ITEM', 'MAT 1', 'MAT 2', 'SIZE', 'RATING', 'CONN.', 'UNIT', 'QTY', 'REFERENCE (Original Description)']
ws1.append(headers1)
for doc_no, pkg_no, tag, unit, qty, desc in tagged:
    ws1.append([doc_no, pkg_no, tag, None, None, None, None, None, None, unit, float(qty), desc])
style_header(ws1, len(headers1))
widths1 = [16, 24, 18, 20, 10, 12, 10, 10, 8, 8, 8, 45]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: Untagged Items (부속품/소모품/공구 — 비관리, 참고용) ──────
ws2 = wb.create_sheet('Untagged Items')
headers2 = ['PKG', 'PKG NO', 'TAG NO (참조)', 'SEQ NO', 'ITEM', 'DESCRIPTION', 'UNIT', 'QTY']
ws2.append(headers2)
seq_counter = defaultdict(int)
for doc_no, pkg_no, tag, parent_tag, unit, qty, desc in untagged:
    seq_counter[parent_tag] += 1
    ws2.append([doc_no, pkg_no, parent_tag, seq_counter[parent_tag], None, desc, unit, float(qty)])
style_header(ws2, len(headers2))
widths2 = [16, 24, 20, 8, 20, 45, 8, 8]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 3: Instructions ─────────────────────────────────────────────
ws3 = wb.create_sheet('Instructions')
ws3.column_dimensions['A'].width = 110
lines = [
    'Speciality (Receiving) 분류 작업 안내 — Valve (Receiving)와 동일한 원칙을 따릅니다.',
    '',
    '[기본 원칙]',
    '- 재고/부족자재 관리는 실제 기기 Tag(계기번호)가 있는 항목만 대상입니다.',
    '- Tag 없는 부속품/소모품(가스켓, 볼트, 홀더, 슬리브 파이프 등)은 관리 대상이 아니라 참고 정보입니다.',
    '',
    '[Speciality 시트 — Tag 있는 관리 대상, 414행 기존 데이터로 사전 채움]',
    '- PKG / PKG NO / TAG NO / UNIT / QTY / REFERENCE(원본 Description)는 기존 DB 값으로 이미 채워져 있습니다.',
    '- ITEM / MAT 1 / MAT 2 / SIZE / RATING / CONN. 은 비어 있습니다. REFERENCE 열을 참고해 직접 채워주세요.',
    '  · ITEM 예시: FLEXIBLE HOSE, FLEXIBLE JOINT, STEAM TRAP, EXPANSION JOINT, STRAINER, SIGHT GLASS, RESTRICTION ORIFICE, AIR TRAP, BIRD SCREEN, EDUCTOR, SPRAY NOZZLE 등',
    '  · MAT 1: 재질 등급(CS/SS/ALLOY 등), MAT 2: 실제 규격(A106-B, A351-CF8 등)',
    '  · CONN.: 연결 방식(BW/SW/RF/FF 등, 확인 안 되면 공란)',
    '- 작업 완료 후 REFERENCE 열은 삭제하지 않아도 됩니다(업로드 시 참고용으로만 쓰이고 무시됩니다).',
    '',
    '[Untagged Items 시트 — Tag 없는 부속품/소모품, 334행 기존 데이터로 사전 채움]',
    '- PKG / PKG NO / TAG NO(참조) / DESCRIPTION / UNIT / QTY는 기존 DB 값으로 이미 채워져 있습니다.',
    '  (TAG NO(참조)는 이 부속품이 어느 Speciality Tag에 딸린 것인지를 나타냅니다.)',
    '- SEQ NO는 같은 TAG NO(참조) 안에서의 순번으로 자동 채워져 있습니다. 필요시 수정 가능합니다.',
    '- ITEM은 비어 있습니다. DESCRIPTION을 참고해 채워주세요(예: GASKET, CONNECTION BOLT, HOLDER, SLEEVE PIPE, TIE ROD 등).',
    '- 이 시트의 항목들은 Valve 예시와 마찬가지로 재고 계산에는 들어가지 않고, 검색으로만 조회됩니다.',
    '',
    '[분류 재검토가 필요한 경우]',
    '- 만약 Speciality 시트의 어떤 Tag가 사실은 특정 Tag에 딸린 부속품이라고 판단되면,',
    '  그 행을 Untagged Items 시트로 옮기고 TAG NO(참조)에 해당 Tag를 입력해주세요.',
    '- 반대로 Untagged Items의 어떤 항목이 사실은 독립된 관리 대상 기기라면 Speciality 시트로 옮겨주세요.',
]
for i, line in enumerate(lines, start=1):
    ws3.cell(row=i, column=1, value=line)
ws3.cell(row=1, column=1).font = Font(bold=True, size=13)

OUT_PATH = 'Raw File/Speciality (Receiving)_Format_Template.xlsx'
wb.save(OUT_PATH)
print(f'저장 완료: {OUT_PATH}')
