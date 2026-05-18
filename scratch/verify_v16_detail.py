# -*- coding: utf-8 -*-
"""
PO_List_Final_v17 vs 8개 Packing List 파일 상세 검증
- 패키지별 Item 수, 수량 비교
"""
import openpyxl, glob, os, sys, re
from collections import defaultdict

BASE = 'c:/Users/PCLOVE/Downloads/ipcs-material/Raw File/'

# Others 카테고리 (통합 시트 미추적 항목) → PL 파싱 시 제외
SKIP_KEYWORDS = ['STUD', 'BOLT', 'NUT', 'GASKET', 'SPIRAL WOUND', 'WASHER',
                 'INSULATION KIT', 'INSUL', 'ORIFICE', 'INSTRUMENT']

def is_others_item(desc):
    if not desc: return False
    d = str(desc).upper()
    return any(kw in d for kw in SKIP_KEYWORDS)

# ----------------------------------------------------------------
# 1. v16 로드 - 패키지별 데이터 수집
# ----------------------------------------------------------------
wb16 = openpyxl.load_workbook(BASE + 'PO_List_Final_v22.xlsx', data_only=True)  # v22 = 최종
# 시트명 확인 - '자재별 수량 현황' (괄호 없는 정확한 이름) 우선
sheet_name = None
for s in wb16.sheetnames:
    if '수량' in s and '매핑' not in s:
        sheet_name = s
        break
if not sheet_name:
    sheet_name = wb16.sheetnames[0]

ws16 = wb16[sheet_name]

# {pkg_no: [{po, qty_ea, item}]}
v16_data = defaultdict(list)
cur_item = ''
for r in range(4, ws16.max_row + 1):
    c1  = ws16.cell(row=r, column=1).value
    c9  = ws16.cell(row=r, column=9).value    # PO NO
    c10 = ws16.cell(row=r, column=10).value   # Pkg No
    c11 = ws16.cell(row=r, column=11).value   # Pkg Qty(ea)
    c12 = ws16.cell(row=r, column=12).value   # Tag No
    if c1:
        cur_item = str(c1)
    if c10:
        v16_data[str(c10).strip()].append({
            'po':   str(c9).strip()  if c9  else '',
            'qty':  float(c11) if c11 and isinstance(c11, (int, float)) else 0,
            'tag':  str(c12).strip() if c12 else '',
            'item': cur_item,
        })

# ----------------------------------------------------------------
# 2. PL 파싱
# ----------------------------------------------------------------
def parse_pl(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Detail PL' not in wb.sheetnames:
        return []
    ws = wb['Detail PL']
    # PO 컬럼 / qty 컬럼 자동 감지 (헤더 행 스캔)
    po_col  = None
    qty_col = None
    for r in range(1, 20):
        for c in range(1, 26):
            v = ws.cell(r, c).value
            if not v or not isinstance(v, str):
                continue
            if re.search(r'\bPO\b', v, re.I) and po_col is None:
                po_col = c - 1   # 0-based
            if ("Q'TY" in v.upper() or 'QTY' in v.upper()) and qty_col is None:
                qty_col = c - 1  # 0-based
    if po_col  is None: po_col  = 18
    if qty_col is None: qty_col = 4
    # Tag 컬럼 자동 감지 - 데이터 행에서 'SHOP' 포함 열 스캔
    tag_col = None
    for r in range(10, min(30, ws.max_row + 1)):
        row = [ws.cell(r, c).value for c in range(1, 26)]
        if not row[1] or 'PGU' not in str(row[1]):
            continue
        for ci in range(2, 20):
            v = row[ci]
            if v and isinstance(v, str) and 'SHOP' in v.upper():
                tag_col = ci
                break
        if tag_col is not None:
            break
    if tag_col is None: tag_col = 13
    # 데이터 행: col B(idx 1)에 'PGU' 포함 여부로 판별
    items = []
    for r in range(10, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 26)]
        col_b = row[1] if len(row) > 1 else None
        if not col_b or not isinstance(col_b, str):
            continue
        col_b = col_b.strip()
        if 'PGU' not in col_b:
            continue
        desc_raw = row[2] if len(row) > 2 else None
        desc = str(desc_raw).strip() if desc_raw else ''
        if is_others_item(desc):
            continue  # STB/NUT/GSKT 등 통합 시트 미추적 항목 스킵
        qty_raw = row[qty_col] if qty_col < len(row) else None
        qty = float(qty_raw) if isinstance(qty_raw, (int, float)) else 0
        po_raw  = row[po_col]  if po_col  < len(row) else None
        po = str(po_raw).strip() if po_raw else ''
        tag_raw = row[tag_col] if tag_col < len(row) else None
        tag = str(tag_raw).strip() if tag_raw else ''
        items.append({
            'pkg': col_b,
            'qty': qty,
            'po':  po,
            'tag': tag,
        })
    return items

# ----------------------------------------------------------------
# 3. 검증 실행
# ----------------------------------------------------------------
PL_NUMS = ['0443', '0503', '0524', '0525', '0554', '0555', '0565', '0574']
pl_files = sorted(glob.glob(BASE + 'Packing List/*.xlsx'))

out = []

for pl_path in pl_files:
    fname = os.path.basename(pl_path)
    pl_num = next((p for p in PL_NUMS if p in fname), None)
    if not pl_num:
        continue

    try:
        pl_items = parse_pl(pl_path)
    except Exception as e:
        out.append('ERROR %s: %s\n' % (pl_num, e))
        continue

    # 패키지별 그룹
    pl_pkg = defaultdict(list)
    for it in pl_items:
        if it['pkg']:
            pl_pkg[it['pkg']].append(it)

    ok = diff = miss = 0
    pkg_lines = []

    for pkg_no in sorted(pl_pkg.keys()):
        pl_its  = pl_pkg[pkg_no]
        v16_its = v16_data.get(pkg_no, [])

        pl_total  = sum(it['qty'] for it in pl_its)
        v16_total = sum(it['qty'] for it in v16_its)

        # Step 1: Tag 기반 매칭 - 동일 Tag+qty → OK (PO 달라도 동일 Item)
        matched_pl  = set()
        matched_v16 = set()
        for i, pl_it in enumerate(pl_its):
            if not pl_it.get('tag'):
                continue
            for j, v16_it in enumerate(v16_its):
                if j in matched_v16:
                    continue
                if (v16_it.get('tag') == pl_it['tag'] and
                        abs(v16_it['qty'] - pl_it['qty']) < 0.01):
                    matched_pl.add(i)
                    matched_v16.add(j)
                    break

        # Step 2: Tag 미매칭 항목만 PO 기반 비교
        unmatched_pl  = [pl_its[i]  for i in range(len(pl_its))  if i not in matched_pl]
        unmatched_v16 = [v16_its[j] for j in range(len(v16_its)) if j not in matched_v16]

        pl_po_qty = defaultdict(float)
        for it in unmatched_pl:
            pl_po_qty[it['po']] += it['qty']

        v16_po_qty = defaultdict(float)
        for it in unmatched_v16:
            v16_po_qty[it['po']] += it['qty']

        # 불일치 PO 찾기
        all_pos = sorted(set(pl_po_qty) | set(v16_po_qty))
        mism = [(po, pl_po_qty.get(po, 0), v16_po_qty.get(po, 0))
                for po in all_pos
                if abs(pl_po_qty.get(po, 0) - v16_po_qty.get(po, 0)) > 0.01]

        if not v16_its:
            st = 'MISSING'
            miss += 1
        elif mism:
            st = 'DIFF'
            diff += 1
        else:
            st = 'OK'
            ok += 1

        flag = '' if st == 'OK' else ('  *** ' + st)

        short_pkg = pkg_no
        # 짧은 표기 (패키지 번호 마지막 부분만)
        parts = pkg_no.split('-')
        short = '-'.join(parts[-3:]) if len(parts) >= 3 else pkg_no

        line = '  %-28s  PL:%3d건/%7.0fEA  v16:%3d행/%7.0fEA%s' % (
            short, len(pl_its), pl_total, len(v16_its), v16_total, flag)
        pkg_lines.append(line)

        # 불일치 상세
        if mism:
            for po, pq, vq in mism:
                pkg_lines.append('      [DIFF] PO=%-35s  PL=%.0f  v16=%.0f  (차이=%+.0f)' % (
                    po, pq, vq, vq - pq))
        if st == 'MISSING':
            # PL 내용 요약 출력
            for it in pl_its[:3]:
                pkg_lines.append('      [미반영] %s | qty=%.0f | po=%s' % (
                    it.get('desc', '')[:50], it['qty'], it['po']))
            if len(pl_its) > 3:
                pkg_lines.append('      ... 외 %d건' % (len(pl_its) - 3))

    header = '=' * 80
    title  = 'PL %s  |  %s' % (pl_num, fname)
    summary = '  >>> 결과: 총 %d패키지  OK=%d  DIFF=%d  MISSING=%d' % (
        len(pl_pkg), ok, diff, miss)

    out.append(header)
    out.append(title)
    out.append(header)
    out += pkg_lines
    out.append(summary)
    out.append('')

# ----------------------------------------------------------------
# 4. 출력
# ----------------------------------------------------------------
sys.stdout.reconfigure(encoding='utf-8', errors='replace') \
    if hasattr(sys.stdout, 'reconfigure') else None

for line in out:
    try:
        print(line)
    except UnicodeEncodeError:
        print(line.encode('ascii', 'replace').decode())
