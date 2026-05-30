# Speciality Item List vs Speciality BOM 비교 검증 스크립트
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SHEET_TAG_COL = {
    '1.STRAINER':             20,
    '2. STEAM TRAP':          20,
    '3. EXPANSION JOINT':     20,
    '4. SIGHT GLASS':         20,
    '5. FLEXIBLE JOINT':      20,
    '6. RESTRICTION ORIFICE': 21,
    '7. FLEXIBLE HOSE':       20,
    '8. SPRAY NOZZLE':        20,
    '9.EDUCTOR':              2,
    '10.AIR TRAP':            20,
    '11.BIRD SCREEN':         19,
}
DATA_START = 5

# ── Item List 태그 수집 ───────────────────────────────────────
wb_il = openpyxl.load_workbook('Raw File/Speciality Item List.xlsx', data_only=True)
il_tags = {}       # {tag: sheet_name}
il_by_sheet = {}

for sh, tag_col in SHEET_TAG_COL.items():
    ws = wb_il[sh]
    tags = []
    for r in range(DATA_START, ws.max_row + 1):
        v = ws.cell(r, tag_col).value
        if not v:
            continue
        tag = str(v).strip()
        skip_vals = {'TAG NO', 'TAG NO.', 'TAG', 'REMARK', 'NOTE', 'NONE', ''}
        if tag.upper() in skip_vals:
            continue
        if not any(tag.startswith(p) for p in ('B0-', 'B1-', 'B2-')):
            continue
        tags.append(tag)
        il_tags[tag] = sh
    il_by_sheet[sh] = tags
    print(f'[{sh}] {len(tags)}개')

print(f'\nItem List 총 TAG 수: {len(il_tags)}개\n')

# ── BOM 태그 수집 ──────────────────────────────────────────────
wb_bom = openpyxl.load_workbook('Raw File/Speciality BOM.xlsx', data_only=True)
ws_bom = wb_bom.active
bom_tags = {}          # {tag: {matcode, desc}}
bom_null_matcode = []

for r in range(2, ws_bom.max_row + 1):
    tag  = ws_bom.cell(r, 3).value
    mat  = ws_bom.cell(r, 2).value
    desc = ws_bom.cell(r, 7).value
    if not tag and not mat:
        continue
    tag_str = str(tag).strip() if tag else ''
    mat_str = str(mat).strip() if mat else ''
    if tag_str:
        bom_tags[tag_str] = {'matcode': mat_str, 'desc': str(desc)[:50] if desc else ''}
    if not mat_str or mat_str == 'None':
        bom_null_matcode.append({'row': r, 'tag': tag_str, 'desc': str(desc)[:50] if desc else ''})

print(f'BOM 총 TAG 수: {len(bom_tags)}개  (전체 데이터 행: {ws_bom.max_row - 1}개)')
print(f'BOM NULL MatCode: {len(bom_null_matcode)}개\n')

# ── 비교 ──────────────────────────────────────────────────────
il_set  = set(il_tags.keys())
bom_set = set(bom_tags.keys())

missing_in_bom = sorted(il_set - bom_set)
extra_in_bom   = sorted(bom_set - il_set)

print('=== 비교 결과 ===')
print(f'Item List에 있으나 BOM 누락: {len(missing_in_bom)}개')
print(f'BOM에 있으나 Item List 없음: {len(extra_in_bom)}개')
print()

if missing_in_bom:
    print('--- [BOM 누락] Item List → BOM 없음 ---')
    for tag in missing_in_bom:
        print(f'  {tag}  (시트: {il_tags[tag]})')
    print()

if extra_in_bom:
    print('--- [BOM 추가] BOM에만 있고 Item List 없음 ---')
    for tag in extra_in_bom:
        mc = bom_tags[tag]['matcode']
        ds = bom_tags[tag]['desc']
        print(f'  {tag}  matcode={mc}  desc={ds}')
    print()

if bom_null_matcode:
    print('--- [NULL MatCode] BOM에서 MatCode 없는 항목 ---')
    for item in bom_null_matcode:
        print(f'  row{item["row"]}  tag={item["tag"]}  desc={item["desc"]}')
    print()

# ── 시트별 요약 ───────────────────────────────────────────────
print('=== 시트별 일치 현황 ===')
for sh, tags in il_by_sheet.items():
    matched = sum(1 for t in tags if t in bom_set)
    miss    = [t for t in tags if t not in bom_set]
    print(f'  [{sh}] IL={len(tags)}, BOM_일치={matched}, 누락={len(miss)}')
    for t in miss:
        print(f'      누락: {t}')
