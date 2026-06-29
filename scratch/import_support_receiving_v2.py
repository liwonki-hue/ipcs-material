# support_receiving 테이블 교체 스크립트 v2
# Base: [TURKISTAN]CRITICAL BOP SUPPLY LIST_260319.xlsx
# ISO DWG NO: 구파일(251120) 우선, 없으면 ipcs-drawing 보완
# PKG/PKG NO: 구파일(251120) 태그 매칭
import openpyxl, requests, json, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
H = {'apikey': KEY, 'Authorization': f'Bearer {KEY}',
     'Content-Type': 'application/json', 'Prefer': 'return=minimal'}

FILE_NEW  = r'C:\Users\PCLOVE\Downloads\ipcs-material\Raw File\[TURKISTAN]CRITICAL BOP SUPPLY LIST_260319.xlsx'
FILE_OLD  = r'C:\Users\PCLOVE\Downloads\ipcs-material\Raw File\TURKISTAN BOP SUPPLY LIST_251120 - 제출.xlsx'
FILE_DRAW = r'C:\Users\PCLOVE\Downloads\ipcs-drawing\Raw File\TRUKISTAN PJT_PIPE SUPPORT LIST(Total)_260622)_updated.xlsx'

def clean(v):
    if v is None: return None
    s = str(v).strip()
    if s in ('', '-', 'None', 'nan'): return None
    if '?' in s or '�' in s: return None
    return s

def to_int(v):
    try: return int(float(v))
    except: return None

# ── 1. ISO DWG NO 딕셔너리 구축 ─────────────────────────────
print('=== ISO DWG NO 매핑 구축 중... ===')

# 1-A. 구파일(251120) ISO DWG NO
iso_map = {}
wb_old = openpyxl.load_workbook(FILE_OLD, read_only=True, data_only=True)
for sh in ['CRITICAL', 'GENERAL']:
    ws = wb_old[sh]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    t = hdr.index('SUPPORT TAG NO.')
    i = hdr.index('ISO DWG NO.')
    for r in rows[1:]:
        tag = clean(r[t])
        iso = clean(r[i])
        if tag and iso and tag not in iso_map:
            iso_map[tag] = iso
print(f'  구파일 ISO 매핑: {len(iso_map)}건')

# 1-B. ipcs-drawing 보완 (구파일에 없는 태그)
wb_draw = openpyxl.load_workbook(FILE_DRAW, read_only=True, data_only=True)
ws_draw = wb_draw['Total']
rows_draw = list(ws_draw.iter_rows(values_only=True))
hdr_draw = rows_draw[0]
t_draw = hdr_draw.index('Support Tag No.')
i_draw = hdr_draw.index('ISO Drawing No.')
added = 0
for r in rows_draw[1:]:
    tag = clean(r[t_draw])
    iso = clean(r[i_draw])
    if tag and iso and tag not in iso_map:
        iso_map[tag] = iso
        added += 1
print(f'  ipcs-drawing 보완: +{added}건 → 합계 {len(iso_map)}건')

# ── 2. PKG / PKG NO 딕셔너리 구축 ──────────────────────────
print('\n=== PKG 매핑 구축 중... ===')
pkg_map = {}
wb_old2 = openpyxl.load_workbook(FILE_OLD, read_only=True, data_only=True)
for sh in ['CRITICAL', 'GENERAL']:
    ws = wb_old2[sh]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    t = hdr.index('SUPPORT TAG NO.')
    pkg_i   = next(i for i, h in enumerate(hdr) if h == 'PKG')
    pkgno_i = next(i for i, h in enumerate(hdr) if h == 'PKG NO')
    for r in rows[1:]:
        tag = clean(r[t])
        pkg    = clean(r[pkg_i])
        pkg_no = clean(r[pkgno_i])
        if tag and tag not in pkg_map and (pkg or pkg_no):
            pkg_map[tag] = (pkg, pkg_no)
print(f'  PKG 매핑: {len(pkg_map)}건')

# ── 3. 신규파일 파싱 ────────────────────────────────────────
print('\n=== 신규파일 파싱 중... ===')
wb_new = openpyxl.load_workbook(FILE_NEW, read_only=True, data_only=True)
ws_new = wb_new['CRITICAL']
rows_new = list(ws_new.iter_rows(values_only=True))
hdr_new = rows_new[0]

ci = {
    'system':  hdr_new.index('SYSTEM'),
    'iso':     hdr_new.index('ISO DWG NO.'),
    'tag':     hdr_new.index('SUPPORT TAG NO.'),
    'type':    hdr_new.index('TYPE'),
    'part_no': hdr_new.index('PART NO'),
    'id_no':   hdr_new.index('ID NO'),
    'item':    hdr_new.index('ITEM'),
    'matl':    hdr_new.index('MATL'),
    'size':    hdr_new.index('SIZE OR TYPE'),
    'length':  next(i for i, h in enumerate(hdr_new) if h and 'LENGTH' in str(h).upper()),
    'qty':     next(i for i, h in enumerate(hdr_new) if h and "Q'TY" in str(h).upper()),
}

rows_out = []
iso_hit, pkg_hit, no_iso, no_pkg = 0, 0, 0, 0

for r in rows_new[1:]:
    if not any(v is not None for v in r):
        continue
    tag = clean(r[ci['tag']])

    # ISO DWG NO: 신규파일 → 구파일 → ipcs-drawing 순
    iso = clean(r[ci['iso']]) or (iso_map.get(tag) if tag else None)
    if iso: iso_hit += 1
    else:   no_iso += 1

    # PKG
    pkg, pkg_no = (None, None)
    if tag and tag in pkg_map:
        pkg, pkg_no = pkg_map[tag]
        pkg_hit += 1
    else:
        no_pkg += 1

    rows_out.append({
        'system':       clean(r[ci['system']]),
        'iso_dwg_no':   iso,
        'support_tag':  tag,
        'type':         clean(r[ci['type']]),
        'part_no':      to_int(r[ci['part_no']]),
        'id_no':        clean(r[ci['id_no']]),
        'item':         clean(r[ci['item']]),
        'matl':         clean(r[ci['matl']]),
        'size_or_type': clean(r[ci['size']]),
        'length_mm':    clean(r[ci['length']]),
        'qty':          to_int(r[ci['qty']]) or 0,
        'pkg':          pkg,
        'package_no':   pkg_no,
    })

print(f'  파싱 완료: {len(rows_out)}행')
print(f'  ISO DWG NO 있음: {iso_hit}행 ({iso_hit/len(rows_out)*100:.1f}%)')
print(f'  ISO DWG NO 없음: {no_iso}행')
print(f'  PKG 있음: {pkg_hit}행 ({pkg_hit/len(rows_out)*100:.1f}%)')
print(f'  PKG 없음: {no_pkg}행')

# ── 4. 기존 데이터 삭제 → 신규 INSERT ───────────────────────
print('\n=== 기존 support_receiving 삭제 중... ===')
resp = requests.delete(f'{URL}/rest/v1/support_receiving',
    headers={**H, 'Prefer': 'return=minimal'}, params={'id': 'gte.0'})
print(f'  DELETE: {resp.status_code}')
if resp.status_code not in (200, 204):
    print('  오류:', resp.text[:300])
    sys.exit(1)

print('\n=== 배치 INSERT 중... ===')
BATCH = 500
ok, fail = 0, 0
for i in range(0, len(rows_out), BATCH):
    batch = rows_out[i:i+BATCH]
    resp = requests.post(f'{URL}/rest/v1/support_receiving', headers=H, data=json.dumps(batch))
    if resp.status_code in (200, 201):
        ok += len(batch)
        print(f'  배치 {i//BATCH+1}/{(len(rows_out)-1)//BATCH+1}: {len(batch)}행 완료')
    else:
        fail += len(batch)
        print(f'  배치 {i//BATCH+1} 실패 ({resp.status_code}): {resp.text[:200]}')

print(f'\n=== 완료: 성공 {ok}행 / 실패 {fail}행 ===')
