# BOM 비교 스크립트: Excel 원본 vs Supabase DB (ISO Drawing 기준)
# 단위 정규화: Excel MM -> M (÷1000)
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import warnings
warnings.filterwarnings("ignore")
import openpyxl
import requests
from collections import defaultdict
import re

SUPABASE_URL = 'https://ognhvfvlboqblueuldlm.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im9nbmh2ZnZsYm9xYmx1ZXVsZGxtIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI3MzY2NTUsImV4cCI6MjA4ODMxMjY1NX0.paO5jr16M7yTySUAp9LgberoatDds9rTNa_eCU_ET_I'
HEADERS = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}', 'Accept-Profile': 'material'}

# ─── 컬럼 인덱스 (0-based, 두 파일 공통) ──────────────────────────────
COL_ISO  = 8    # ISO DWG NO
COL_LINE = 9    # LINE NO
COL_SYS  = 6    # SYSTEM
COL_ITEM = 14   # ITEM
COL_ET   = 15   # END_TYPE
COL_ML1  = 16   # MATL1
COL_SIZE = 18   # SIZE
# LB: UOM=23, QTY=24  /  SB: UOM=24, QTY=25

def norm(s):
    return ' '.join(str(s).upper().strip().split()) if s else ''

def norm_uom(uom_str):
    """단위 정규화: MM -> M (DB는 M만 사용)"""
    u = norm(uom_str)
    return 'M' if u == 'MM' else u

def norm_qty(qty_raw, uom_raw):
    """수량 정규화: MM 단위이면 /1000 -> M"""
    try:
        q = float(qty_raw) if qty_raw is not None else 0.0
    except (TypeError, ValueError):
        q = 0.0
    if norm(uom_raw) == 'MM':
        q = q / 1000.0
    return q

# ─── Excel 파싱 ────────────────────────────────────────────────────────

def parse_excel(filepath, header_row_idx, uom_col, qty_col, source):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if 'LARGE BORE' in filepath:
        ws = wb["Piping&Fitting"]
    else:
        ws = wb["MERGED_Piping&Fitting"]

    # (iso, line) -> (item, size, uom_norm) -> qty_norm
    result = defaultdict(lambda: defaultdict(float))
    iso_set = set()
    # (iso, line) -> [(item, size, uom_norm, qty_norm, matl1)]  — 상세 보존
    detail = defaultdict(list)
    rows = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row_idx:
            continue
        no_val = row[1] if len(row) > 1 else None
        if no_val is None or not isinstance(no_val, (int, float)):
            continue

        iso  = norm(row[COL_ISO]  if len(row) > COL_ISO  else '')
        line = norm(row[COL_LINE] if len(row) > COL_LINE else '')
        item = norm(row[COL_ITEM] if len(row) > COL_ITEM else '')
        size = norm(row[COL_SIZE] if len(row) > COL_SIZE else '')
        size = re.sub(r'DN\s+', 'DN', size)
        matl = norm(row[COL_ML1]  if len(row) > COL_ML1  else '')
        et   = norm(row[COL_ET]   if len(row) > COL_ET   else '')
        uom_raw = row[uom_col] if len(row) > uom_col else ''
        qty_raw = row[qty_col] if len(row) > qty_col else None

        if not iso or not item:
            continue

        uom = norm_uom(uom_raw)
        qty = norm_qty(qty_raw, uom_raw)

        iso_set.add(iso)
        key = (item, size, uom)
        result[(iso, line)][key] += qty
        detail[(iso, line)].append((item, size, uom, qty, matl, et))
        rows += 1

    wb.close()
    print(f"  {source}: {rows}행, {len(iso_set)} ISO")
    return result, detail, iso_set

print("=== Excel BOM 파싱 중... ===")
lb_agg, lb_det, lb_isos = parse_excel("Raw File/LARGE BORE BOM(251223).xlsm",  8, 23, 24, "LB")
sb_agg, sb_det, sb_isos = parse_excel("Raw File/SB BOM(20260128) (002).xlsx",   1, 24, 25, "SB")

# 합치기
excel_agg = defaultdict(lambda: defaultdict(float))
excel_det = defaultdict(list)
for (iso, line), d in lb_agg.items():
    for k, q in d.items():
        excel_agg[(iso, line)][k] += q
for (iso, line), d in sb_agg.items():
    for k, q in d.items():
        excel_agg[(iso, line)][k] += q
for (iso, line), rows in lb_det.items():
    excel_det[(iso, line)].extend(rows)
for (iso, line), rows in sb_det.items():
    excel_det[(iso, line)].extend(rows)

all_excel_isos = lb_isos | sb_isos
print(f"\nExcel 전체: {len(all_excel_isos)} ISO")

# ─── Supabase DB fetch ─────────────────────────────────────────────────

print("\n=== Supabase BOM fetch 중... ===")
all_db_rows = []
offset = 0
while True:
    resp = requests.get(
        f'{SUPABASE_URL}/rest/v1/bom'
        f'?select=iso_dwg_no,line_no,full_description,uom,qty,mat_code'
        f'&iso_dwg_no=neq.&limit=1000&offset={offset}',
        headers=HEADERS
    )
    rows = resp.json()
    if not rows: break
    all_db_rows.extend(rows)
    offset += 1000
    if len(rows) < 1000: break

print(f"  DB fetch: {len(all_db_rows)}행")

# DB 구조화: (iso, line) -> (item_first_word, uom) -> qty
# item_first_word = full_description에서 첫 번째 콤마 이전
db_agg = defaultdict(lambda: defaultdict(float))
db_detail = defaultdict(list)
db_isos = set()

for row in all_db_rows:
    iso   = norm(row.get('iso_dwg_no', ''))
    line  = norm(row.get('line_no', ''))
    fdesc = row.get('full_description') or ''
    uom   = norm(row.get('uom', ''))
    qty   = float(row.get('qty') or 0)
    mcode = row.get('mat_code') or ''

    db_isos.add(iso)
    item_kw = norm(fdesc.split(',')[0]) if ',' in fdesc else norm(fdesc)
    key = (item_kw, uom)
    db_agg[(iso, line)][key] += qty
    db_detail[(iso, line)].append({'desc': fdesc, 'uom': uom, 'qty': qty, 'mat_code': mcode})

print(f"  DB ISO: {len(db_isos)}개")

# ─── 비교 ─────────────────────────────────────────────────────────────

print("\n=== BOM 비교 ===\n")
excel_only_isos = sorted(all_excel_isos - db_isos)
db_only_isos    = sorted(db_isos - all_excel_isos)
common_isos     = all_excel_isos & db_isos

print(f"전체 고유 ISO: {len(all_excel_isos | db_isos)}")
print(f"Excel에만 있는 ISO (DB 미등록): {len(excel_only_isos)}")
print(f"DB에만 있는 ISO (Excel 없음): {len(db_only_isos)}")
print(f"공통 ISO: {len(common_isos)}")

# Excel-only ISO 시스템 코드 분류
sys_counter = defaultdict(int)
for iso in excel_only_isos:
    # 시스템 코드 추출: PI-140-XX-NNN -> XX
    m = re.search(r'PI-\d+-([A-Z]+)-', iso)
    sys_code = m.group(1) if m else '??'
    sys_counter[sys_code] += 1

print(f"\n[DB 미등록 ISO 시스템 분류]")
for sys_code, cnt in sorted(sys_counter.items(), key=lambda x: -x[1]):
    print(f"  {sys_code}: {cnt}개")

# LINE NO 비교 (공통 ISO 내)
excel_lines_by_iso = defaultdict(set)
for (iso, line) in excel_agg:
    excel_lines_by_iso[iso].add(line)

db_lines_by_iso = defaultdict(set)
for (iso, line) in db_agg:
    db_lines_by_iso[iso].add(line)

missing_lines = []   # (iso, line, items_str)
qty_diffs     = []   # (iso, line, item, uom, excel_q, db_q, diff)

for iso in sorted(common_isos):
    ex_lines = excel_lines_by_iso.get(iso, set())
    db_lines = db_lines_by_iso.get(iso, set())

    # 누락 LINE
    for line in sorted(ex_lines - db_lines):
        items_info = []
        for item, size, uom, qty, matl, et in excel_det.get((iso, line), []):
            if qty > 0:
                items_info.append(f"{item} {size} [{uom}] {qty:.3f}")
        # 중복 제거
        seen = set()
        uniq = []
        for x in items_info:
            if x not in seen:
                seen.add(x)
                uniq.append(x)
        missing_lines.append((iso, line, uniq))

    # 공통 LINE: item별 qty 비교
    for line in sorted(ex_lines & db_lines):
        ex_items = excel_agg[(iso, line)]  # (item, size, uom) -> qty
        db_items = db_agg[(iso, line)]     # (item_kw, uom) -> qty

        # Excel 쪽을 item_kw 기준으로 재집계 (size 무시하고 item keyword 레벨)
        ex_by_kw = defaultdict(float)
        for (item, size, uom), q in ex_items.items():
            ex_by_kw[(item, uom)] += q

        # DB도 동일
        db_by_kw = dict(db_items)

        all_keys = set(ex_by_kw.keys()) | set(db_by_kw.keys())
        for key in all_keys:
            eq = ex_by_kw.get(key, 0.0)
            dq = db_by_kw.get(key, 0.0)
            diff = round(eq - dq, 4)
            if abs(diff) > 0.05:  # 0.05M(=50mm) 이상 차이만 보고
                qty_diffs.append((iso, line, key[0], key[1], eq, dq, diff))

# ─── 출력 ─────────────────────────────────────────────────────────────

print(f"\n누락 LINE NO: {len(missing_lines)}건")
print(f"수량 불일치: {len(qty_diffs)}건 (0.05M/EA 초과)")

if missing_lines:
    print("\n[ 누락 LINE NO 목록 ]")
    for iso, line, items in missing_lines:
        print(f"\n  ISO:  {iso}")
        print(f"  LINE: {line}")
        for it in items[:5]:
            print(f"    - {it}")

print(f"\n{'='*110}")
print(f"[ 수량 불일치 상위 50건 ]  (단위 정규화 후, Excel MM -> M 변환 완료)")
print(f"{'ISO':<45} {'LINE':<35} {'ITEM':<20} {'UOM':<4} {'Excel':>9} {'DB':>9} {'DIFF':>9}")
print("-"*130)
for iso, line, item, uom, eq, dq, diff in sorted(qty_diffs, key=lambda x: abs(x[6]), reverse=True)[:50]:
    print(f"{iso:<45} {line:<35} {item:<20} {uom:<4} {eq:>9.3f} {dq:>9.3f} {diff:>+9.3f}")

# ─── Excel 저장 ────────────────────────────────────────────────────────

print("\n=== 결과 Excel 저장 중... ===")
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb_out = Workbook()

# ── 시트1: DB미등록 ISO ──
ws1 = wb_out.active
ws1.title = "DB미등록_ISO"
ws1.append(["ISO DWG NO", "시스템코드", "비고"])
for iso in sorted(excel_only_isos):
    m = re.search(r'PI-\d+-([A-Z]+)-', iso)
    sys_code = m.group(1) if m else '??'
    ws1.append([iso, sys_code, "Excel에 있으나 DB 미등록"])

# ── 시트2: 누락 LINE NO ──
ws2 = wb_out.create_sheet("누락_LINE_NO")
ws2.append(["ISO DWG NO", "LINE NO", "Excel 항목 (Item Size [UOM] QTY)"])
for iso, line, items in missing_lines:
    ws2.append([iso, line, " | ".join(items)])

# ── 시트3: 수량 불일치 ──
ws3 = wb_out.create_sheet("수량불일치")
ws3.append(["ISO DWG NO", "LINE NO", "ITEM", "UOM", "Excel QTY", "DB QTY", "DIFF(Excel-DB)"])
red_fill   = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
yel_fill   = PatternFill(start_color="FFFFAA", end_color="FFFFAA", fill_type="solid")
for iso, line, item, uom, eq, dq, diff in sorted(qty_diffs, key=lambda x: abs(x[6]), reverse=True):
    ws3.append([iso, line, item, uom, round(eq, 4), round(dq, 4), round(diff, 4)])
    if abs(diff) > 5:
        for cell in ws3[ws3.max_row]:
            cell.fill = red_fill
    elif abs(diff) > 1:
        for cell in ws3[ws3.max_row]:
            cell.fill = yel_fill

# ── 시트4: ISO별 요약 ──
ws4 = wb_out.create_sheet("ISO별_요약")
ws4.append(["ISO DWG NO", "출처", "Excel LINE수", "DB LINE수", "누락LINE수", "수량불일치LINE수", "상태"])

missing_line_by_iso = defaultdict(int)
for iso, line, items in missing_lines:
    missing_line_by_iso[iso] += 1

qdiff_line_by_iso = defaultdict(set)
for iso, line, item, uom, eq, dq, diff in qty_diffs:
    qdiff_line_by_iso[iso].add(line)

for iso in sorted(all_excel_isos):
    src = "LB" if iso in lb_isos else "SB"
    if iso in lb_isos and iso in sb_isos:
        src = "LB+SB"
    ex_l = len(excel_lines_by_iso.get(iso, set()))
    db_l = len(db_lines_by_iso.get(iso, set()))
    miss = missing_line_by_iso.get(iso, 0)
    qdiff_cnt = len(qdiff_line_by_iso.get(iso, set()))

    if iso not in db_isos:
        status = "DB미등록"
    elif miss > 0 or qdiff_cnt > 0:
        status = "불일치"
    else:
        status = "일치"
    ws4.append([iso, src, ex_l, db_l, miss, qdiff_cnt, status])

# ── 시트5: 전체 집계 요약 ──
ws5 = wb_out.create_sheet("집계_요약")
ws5.append(["항목", "수량"])
ws5.append(["Excel 전체 ISO", len(all_excel_isos)])
ws5.append(["DB 전체 ISO (BOM 있는)", len(db_isos)])
ws5.append(["공통 ISO", len(common_isos)])
ws5.append(["DB 미등록 ISO (Excel만)", len(excel_only_isos)])
ws5.append(["DB에만 있는 ISO", len(db_only_isos)])
ws5.append(["", ""])
ws5.append(["공통 ISO 내 누락 LINE NO", len(missing_lines)])
ws5.append(["수량 불일치 (item+uom 기준, 0.05이상)", len(qty_diffs)])
ws5.append(["", ""])
ws5.append(["[DB미등록 ISO 시스템별]", ""])
for sys_code, cnt in sorted(sys_counter.items(), key=lambda x: -x[1]):
    ws5.append([f"  {sys_code} 시스템", cnt])

out_path = "scratch/BOM_비교결과.xlsx"
wb_out.save(out_path)
print(f"저장 완료: {out_path}")
